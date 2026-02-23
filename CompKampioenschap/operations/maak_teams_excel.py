# -*- coding: utf-8 -*-

#  Copyright (c) 2020-2026 Ramon van der Winkel.
#  All rights reserved.
#  Licensed under BSD-3-Clause-Clear. See LICENSE file for details.

from django.conf import settings
from django.utils import timezone
from django.http import HttpResponse
from Competitie.definities import DEEL_RK, DEELNAME_JA
from Competitie.models import Kampioenschap, CompetitieTeamKlasse, CompetitieMatch, KampioenschapSporterBoog, KampioenschapTeam
from Sporter.models import SporterVoorkeuren
from tempfile import NamedTemporaryFile
from copy import copy
import openpyxl
import zipfile
import shutil
import os


TEMPLATE_FPATH = os.path.join('CompKampioenschap', 'files', 'template-excel-teams.xlsx')

CONTENT_TYPE_XLSX = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'     # noqa


class MaakTeamsExcel:

    """ Maak en kopie van de teams excel template en vul deze met een deelnemerslijst """

    def __init__(self, deelkamp: Kampioenschap,
                       team_klasse: CompetitieTeamKlasse,
                       match: CompetitieMatch):
        """
            deelkamp,
            team_klasse,
            match:       Voor deze wedstrijd de excel vullen.
        """
        self.match = match
        self.team_klasse = team_klasse
        self.deelkamp = deelkamp
        self.klasse_str = self.team_klasse.beschrijving
        self.vastgesteld = timezone.localtime(timezone.now())

        comp = deelkamp.competitie
        if deelkamp.deel == DEEL_RK:
            rayon = deelkamp.rayon
            self.titel = 'RK Teams Rayon %s, %s' % (rayon.rayon_nr, comp.beschrijving)
            self.deelkamp_titel = 'Rayonkampioen'
        else:
            self.titel = 'BK Teams, %s' % comp.beschrijving
            self.deelkamp_titel = team_klasse.titel_bk

        self.lid2voorkeuren = dict()  # [lid_nr] = (SporterVoorkeuren: .para_voorwerpen, .opmerking_para_sporter)
        self._laad_lid_voorkeuren()

        self.deelnemers_ver_nrs = list()

    def _laad_lid_voorkeuren(self):
        self.lid2voorkeuren = dict()  # [lid_nr] = (SporterVoorkeuren: .para_voorwerpen, .opmerking_para_sporter)
        for tup in SporterVoorkeuren.objects.select_related('sporter').values_list('sporter__lid_nr',
                                                                                   'para_voorwerpen',
                                                                                   'opmerking_para_sporter'):
            lid_nr, para_voorwerpen, opmerking_para_sporter = tup
            self.lid2voorkeuren[lid_nr] = (para_voorwerpen, opmerking_para_sporter)
        # for

    def _vul_deelnemers(self, prg):
        # maak wijzigingen in het RK programma
        ws = prg['Deelnemers']

        ws['B1'] = self.titel
        ws['B2'] = self.klasse_str

        # organisatie
        if self.match.locatie:
            ws['B3'] = self.match.vereniging.naam + ', ' + self.match.locatie.plaats
        else:
            ws['B3'] = self.match.vereniging.naam

        ws['B4'] = self.match.datum_wanneer.strftime('%Y-%m-%d')

        teams = (KampioenschapTeam
                 .objects
                 .filter(kampioenschap=self.deelkamp,
                         team_klasse=self.team_klasse.pk,
                         deelname=DEELNAME_JA)              # reserve teams staan nog op ?
                 .select_related('vereniging',
                                 'vereniging__regio')
                 .prefetch_related('gekoppelde_leden')
                 .order_by('rank'))

        self.deelnemers_ver_nrs = list()

        volg_nr = 0
        for team in teams[:8]:
            row_nr = 8 + volg_nr * 5
            row = str(row_nr)

            ver = team.vereniging
            if ver.ver_nr not in self.deelnemers_ver_nrs:
                self.deelnemers_ver_nrs.append(ver.ver_nr)

            # team naam
            ws['B' + row] = team.team_naam

            # ver_nr
            ws['C' + row] = str(ver.ver_nr)

            # team sterkte
            # sterkte_str = "%.1f" % (team.aanvangsgemiddelde * aantal_pijlen)
            # sterkte_str = sterkte_str.replace('.', ',')
            # ws['G' + row] = sterkte_str

            # vul de 3 sporters in
            aantal = 0
            for deelnemer in team.gekoppelde_leden.select_related('sporterboog__sporter'):
                row_nr += 1
                row = str(row_nr)

                sporter = deelnemer.sporterboog.sporter

                para_mark = False
                try:
                    voorkeuren_para_voorwerpen, voorkeuren_opmerking_para_sporter = self.lid2voorkeuren[sporter.lid_nr]
                except KeyError:        # pragma: no cover
                    pass
                else:
                    if voorkeuren_para_voorwerpen or voorkeuren_opmerking_para_sporter:
                        para_mark = True

                # volledige naam
                naam_str = sporter.volledige_naam()
                if para_mark:
                    naam_str += ' **'
                ws['B' + row] = naam_str

                # bondsnummer
                ws['C' + row] = str(sporter.lid_nr)

                # regio gemiddelde
                gem_str = '%.3f' % deelnemer.gemiddelde
                gem_str = gem_str.replace('.', ',')         # NL komma
                ws['D' + row] = gem_str

                aantal += 1
            # for

            # altijd 3 regels invullen
            while aantal < 3:
                row_nr += 1
                row = str(row_nr)
                ws['B' + row] = 'n.v.t.'    # naam
                ws['C' + row] = '-'         # bondsnummer
                ws['D' + row] = ''          # gemiddelde
                aantal += 1
            # while

            volg_nr += 1
        # for

        # aantal aangemelde teams
        if volg_nr <= 2:
            aantal = "2"
        elif volg_nr <= 4:
            aantal = "3 of 4"
        elif volg_nr <= 6:
            aantal = "5 of 6"
        else:
            aantal = "7 of 8"
        ws['D5'] = aantal

        while volg_nr < 8:
            row_nr = 8 + volg_nr * 5
            row = str(row_nr)

            # vereniging leeg maken
            ws['B' + row] = 'n.v.t.'     # vereniging
            ws['C' + row] = ''           # regio

            # sporters leegmaken
            aantal = 0
            while aantal < 3:
                row_nr += 1
                row = str(row_nr)
                ws['B' + row] = ''         # naam
                ws['C' + row] = ''         # bondsnummer
                ws['D' + row] = ''         # gemiddelde
                aantal += 1
            # while

            volg_nr += 1
        # while

        ws['A50'] = 'Deze gegevens zijn opgehaald op %s' % self.vastgesteld.strftime('%Y-%m-%d %H:%M:%S')

    def _vul_toegestane_deelnemers(self, prg):
        # alle gerechtigde deelnemers opnemen op een apart tabblad, met gemiddelde en boogtype

        ws = prg['Toegestane deelnemers']

        ws['B1'] = '%s, %s' % (self.titel, self.klasse_str)

        cd_font = ws['C18'].font
        c_align = ws['C18'].alignment
        c_format = ws['C18'].number_format

        d_align = ws['D18'].alignment
        d_format = ws['D18'].number_format

        efgh_font = ws['E18'].font      # noqa
        e_align = ws['E18'].alignment

        f_align = ws['F18'].alignment

        g_align = ws['G18'].alignment
        g_format = ws['G18'].number_format

        boog_typen = self.team_klasse.team_type.boog_typen.all()
        boog_pks = list(boog_typen.values_list('pk', flat=True))

        row_nr = 16
        prev_ver = None
        for deelnemer in (KampioenschapSporterBoog
                          .objects
                          .filter(kampioenschap=self.deelkamp,
                                  bij_vereniging__ver_nr__in=self.deelnemers_ver_nrs,
                                  sporterboog__boogtype__pk__in=boog_pks)       # filter op toegestane boogtypen
                          .select_related('bij_vereniging__regio',
                                          'sporterboog__sporter',
                                          'sporterboog__boogtype')
                          .order_by('bij_vereniging__regio',
                                    'bij_vereniging',
                                    '-gemiddelde')):                            # hoogste eerst

            row_nr += 1
            row = str(row_nr)

            # vereniging
            ver = deelnemer.bij_vereniging
            if ver != prev_ver:
                row_nr += 1     # extra lege regel
                row = str(row_nr)
                ws['C' + row] = ver.regio.regio_nr
                if row_nr != 18:
                    ws['C' + row].font = copy(cd_font)
                    ws['C' + row].alignment = copy(c_align)
                    ws['C' + row].number_format = copy(c_format)

                ws['D' + row] = '[%s] %s' % (ver.ver_nr, ver.naam)
                if row_nr != 18:
                    ws['D' + row].font = copy(cd_font)
                    ws['D' + row].alignment = copy(d_align)
                    ws['D' + row].number_format = copy(d_format)

                prev_ver = ver

            # sporter
            sporter = deelnemer.sporterboog.sporter

            para_notities = ''
            try:
                voorkeuren_para_voorwerpen, voorkeuren_opmerking_para_sporter = self.lid2voorkeuren[sporter.lid_nr]
            except KeyError:        # pragma: no cover
                pass
            else:
                if voorkeuren_para_voorwerpen:
                    para_notities = 'Sporter laat voorwerpen op de schietlijn staan. '

                if voorkeuren_opmerking_para_sporter:
                    para_notities += voorkeuren_opmerking_para_sporter
            para_notities = para_notities.strip()

            naam_str = sporter.volledige_naam()
            if para_notities:
                naam_str += ' **'
            ws['E' + row] = naam_str

            ws['F' + row] = str(sporter.lid_nr)

            gem_str = '%.3f' % deelnemer.gemiddelde
            gem_str = gem_str.replace('.', ',')  # NL komma
            ws['G' + row] = gem_str
            ws['H' + row] = deelnemer.sporterboog.boogtype.beschrijving

            if para_notities:
                ws['I' + row] = para_notities

            if row_nr != 18:
                ws['E' + row].font = copy(efgh_font)
                ws['F' + row].font = copy(efgh_font)
                ws['G' + row].font = copy(efgh_font)
                ws['H' + row].font = copy(efgh_font)
                ws['I' + row].font = copy(efgh_font)

                ws['E' + row].alignment = copy(e_align)
                ws['G' + row].alignment = copy(g_align)
                ws['G' + row].number_format = copy(g_format)
        # for

        row_nr += 2
        row = str(row_nr)
        ws['B' + row] = 'Deze gegevens zijn opgehaald op %s' % self.vastgesteld.strftime('%Y-%m-%d %H:%M:%S')
        ws['B' + row].font = copy(efgh_font)
        ws['B' + row].alignment = copy(f_align)

    def _vul_stand(self, prg):
        # vul de te behalen titel in

        ws = prg['Stand']

        ws['G8'] = self.deelkamp_titel      # Rayonkampioen / Bondskampioen / Nederlands Kampioen

    def vul_excel(self) -> HttpResponse:

        """
            Geeft een HttpResponse terug met de inhoud van het bestand.
            can raise RuntimeError
        """

        template_fpath = os.path.join(settings.INSTALL_PATH, TEMPLATE_FPATH)

        # bepaal de naam van het terug te geven bestand
        if self.deelkamp.deel == DEEL_RK:
            fname = "rk-programma_teams-rayon%s_%s.xlsx" % (self.deelkamp.rayon.rayon_nr,
                                                            self.klasse_str.lower().replace(' ', '-'))
        else:
            fname = "bk-programma_teams_%s.xlsx" % self.klasse_str.lower().replace(' ', '-')

        # make een tijdelijk bestand aan
        tmp_file = NamedTemporaryFile()

        try:
            shutil.copyfile(template_fpath, tmp_file.name)
        except FileNotFoundError:
            raise RuntimeError('Kan teams programma template bestand niet vinden')

        # open de kopie, zodat we die aan kunnen passen
        try:
            prg = openpyxl.load_workbook(tmp_file)
        except (OSError, zipfile.BadZipFile, KeyError):
            raise RuntimeError('Kan zojuist aangemaakte programma niet laden')

        self._vul_deelnemers(prg)
        self._vul_toegestane_deelnemers(prg)
        self._vul_stand(prg)

        # maak een response met daarin het ingevulde programma
        response = HttpResponse(content_type=CONTENT_TYPE_XLSX)
        response['Content-Disposition'] = 'attachment; filename="%s"' % fname
        prg.save(response)      # noqa

        del prg
        tmp_file.close()

        return response


# end of file
