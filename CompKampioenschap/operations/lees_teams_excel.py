# -*- coding: utf-8 -*-

#  Copyright (c) 2022-2026 Ramon van der Winkel.
#  All rights reserved.
#  Licensed under BSD-3-Clause-Clear. See LICENSE file for details.

from openpyxl.utils.exceptions import InvalidFileException
from decimal import Decimal, InvalidOperation
from types import SimpleNamespace
import openpyxl
import zipfile


class LeesTeamsExcel:

    def __init__(self):
        self.issues = list()

        """
            teams = [team, ...]
            
            team.row_nr = int
            team.team_naam = str
            team.ver_nr = int
            team.leden = [lid, ...]
            
            lid.row_nr = int
            lid.lid_nr = int
            lid.lid_ag = Decimal
        """
        self.teams = list()

        """
            eindstand = list[stand, ...]
            
            stand.team_naam = str
            stand.matchpunten = int
            stand.shootoff = int | None
        """
        self.eindstand = list()

    @staticmethod
    def _lees_team_naam(ws, cell):
        team_naam = ws[cell].value
        if team_naam is None:
            team_naam = ''
        else:
            team_naam = str(team_naam)
            if str(team_naam).upper() in ('N.V.T.', 'NVT', 'BYE', '#N/A', '0'):
                team_naam = ''
        return team_naam

    def _lees_teams_en_sporters(self, prg):
        blad = 'Deelnemers'
        try:
            ws = prg[blad]
        except KeyError:        # pragma: no cover
            self.issues.append('[ERROR] Kan blad %s niet vinden' % repr(blad))
            return

        for row_nr in range(8, 43+1, 5):

            # team naam
            team_naam = self._lees_team_naam(ws, 'B' + str(row_nr))
            if not team_naam:
                continue

            # ver_nr
            try:
                ver_nr = int(ws['C' + str(row_nr)].value)
            except ValueError:
                continue

            # lees de deelnemers van dit team
            leden = list()
            for sporter_row in range(row_nr+1, row_nr+4):
                lid_nr_str = ws['C' + str(sporter_row)].value
                if lid_nr_str is None:
                    # silently skip
                    continue
                try:
                    lid_nr = int(lid_nr_str)
                except ValueError:
                    self.issues.append('[ERROR] Geen valide lid_nr %s op regel %s' % (repr(lid_nr_str), sporter_row))
                    continue

                lid_ag_str = ws['D' + str(sporter_row)].value
                lid_ag_str = str(lid_ag_str).replace(',', '.')      # decimale komma naar punt
                try:
                    lid_ag = round(Decimal(lid_ag_str), 3)          # 3 cijfers achter de komma
                except (TypeError, InvalidOperation):
                    self.issues.append('[ERROR] Geen valide AG %s op regel %s' % (repr(lid_ag_str), sporter_row))
                    continue

                lid = SimpleNamespace(
                                lid_nr=lid_nr,
                                lid_ag=lid_ag,
                                row_nr=sporter_row)
                leden.append(lid)
            # for sporter

            team = SimpleNamespace(
                            team_naam=team_naam,
                            ver_nr=ver_nr,
                            leden=leden,
                            row_nr=row_nr)
            self.teams.append(team)
        # for team row

    def _lees_eindstand(self, prg):
        blad = 'Stand'
        try:
            ws = prg[blad]
        except KeyError:        # pragma: no cover
            self.issues.append('[ERROR] Kan blad %s niet vinden' % repr(blad))
            return

        for row_nr in range(8, 15+1):
            # team naam
            team_naam = self._lees_team_naam(ws, 'B' + str(row_nr))
            if not team_naam:
                continue

            matchpunten_str = ws['D' + str(row_nr)].value
            try:
                matchpunten = int(matchpunten_str)
            except ValueError:
                self.issues.append('[ERROR] Geen valide matchpunten %s op regel %s' % (repr(matchpunten_str), row_nr))
                continue

            shootoff_str = str(ws['F' + str(row_nr)].value)
            if shootoff_str.upper() == 'NONE':
                shootoff_str = ''
            if shootoff_str != '':
                try:
                    shootoff = int(shootoff_str)
                except ValueError:
                    self.issues.append('[ERROR] Geen valide shootoff %s op regel %s' % (repr(shootoff_str), row_nr))
                    continue
            else:
                shootoff = None

            stand = SimpleNamespace(
                            team_naam=team_naam,
                            matchpunten=matchpunten,
                            shootoff=shootoff)
            self.eindstand.append(stand)
        # for

    def lees_bestand(self, fname: str):
        try:
            prg = openpyxl.load_workbook(fname,
                                         read_only=True,        # avoids warnings
                                         data_only=True)        # do not evaluate formulas; use last calculated values
        except (OSError, zipfile.BadZipFile, KeyError, InvalidFileException) as exc:
            self.issues.append('[ERROR] Kan het excel bestand niet openen (%s)' % str(exc))
            return

        self._lees_teams_en_sporters(prg)
        self._lees_eindstand(prg)

# end of file
