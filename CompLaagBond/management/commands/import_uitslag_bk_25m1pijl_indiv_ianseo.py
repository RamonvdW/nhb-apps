# -*- coding: utf-8 -*-

#  Copyright (c) 2023 Ramon van der Winkel.
#  All rights reserved.
#  Licensed under BSD-3-Clause-Clear. See LICENSE file for details.

from django.core.management.base import BaseCommand
from Competitie.definities import DEEL_BK, DEELNAME_NEE, KAMP_RANK_NO_SHOW, KAMP_RANK_RESERVE
from Competitie.models import KampioenschapSporterBoog
from openpyxl.utils.exceptions import InvalidFileException
import openpyxl
import zipfile


class Command(BaseCommand):
    help = "Importeer uitslag 25m1pijl BK individueel uit Ianseo download"

    def __init__(self, stdout=None, stderr=None, no_color=False, force_color=False):
        super().__init__(stdout, stderr, no_color, force_color)
        self.ver2deelnemers = dict()     # [ver_nr] = [KampioenschapSporterBoog, ...]
        self.klasse_pk = None
        self.dryrun = True
        self.verbose = False

    def add_arguments(self, parser):
        parser.add_argument('--dryrun', action='store_true')
        parser.add_argument('--verbose', action='store_true')
        parser.add_argument('bestand', type=str,
                            help='Pad naar het HTML bestand')

    def deelnemers_ophalen(self):
        for deelnemer in (KampioenschapSporterBoog
                          .objects
                          .filter(kampioenschap__competitie__afstand='25',
                                  kampioenschap__deel=DEEL_BK)
                          .select_related('kampioenschap',
                                          'sporterboog__sporter',
                                          'sporterboog__boogtype',
                                          'bij_vereniging',
                                          'indiv_klasse')):

            ver_nr = deelnemer.bij_vereniging.ver_nr
            try:
                self.ver2deelnemers[ver_nr].append(deelnemer)
            except KeyError:
                self.ver2deelnemers[ver_nr] = [deelnemer]
        # for

    def lees_uitslagen(self, ws):
        col_lid_nr = 'D'
        col_score1 = 'J'
        col_score2 = 'K'
        col_10s = 'M'
        col_9s = 'N'
        col_8s = 'O'

        # houd bij welke klassen er langs komen, voor het melden van no-shows
        deelkamp_pk = None

        # doorloop het blad opnieuw en koppel de scores aan de sporters (in de bepaalde klasse)
        row_nr = 0
        nix_count = 0
        rank_doorlopend = rank = 0
        prev_totaal = 999
        prev_counts_str = ""
        toon_counts = True
        while nix_count < 10:
            row_nr += 1
            row = str(row_nr)

            lid_nr_str = ws[col_lid_nr + row].value
            if lid_nr_str:
                nix_count = 0
                try:
                    lid_nr = int(lid_nr_str)
                except ValueError:
                    pass
                else:
                    try:
                        deelnemers = self.deelnemers[lid_nr]
                    except KeyError:
                        self.stderr.write('[ERROR] Geen BK deelnemer op regel %s: %s' % (row, lid_nr))
                    else:
                        if len(deelnemers) == 1:
                            deelnemer = deelnemers[0]
                        else:
                            deelnemer = None
                            for kandidaat in deelnemers:
                                if kandidaat.indiv_klasse.pk == self.klasse_pk:
                                    deelnemer = kandidaat
                            # for
                        if deelnemer is None:
                            self.stderr.write('[ERROR] Kan deelnemer niet bepalen voor regel %s. Keuze uit %s' % (row, repr(deelnemers)))
                            continue    # met de while

                        dupe_check = False
                        if deelnemer.result_rank > 0:
                            dupe_check = True

                        if deelnemer.indiv_klasse.pk != self.klasse_pk:
                            self.stderr.write('[INFO] Verkeerde klasse: %s' % deelnemer.indiv_klasse)
                        else:
                            deelkamp_pk = deelnemer.kampioenschap.pk

                        score1 = ws[col_score1 + row].value
                        score2 = ws[col_score2 + row].value
                        c10 = ws[col_10s + row].value
                        c9 = ws[col_9s + row].value
                        c8 = ws[col_8s + row].value

                        try:
                            score1 = int(score1)
                            score2 = int(score2)
                        except (TypeError, ValueError):
                            # if deelnemer.deelname != DEELNAME_NEE:      # afgemeld?
                            if score1 is None and score2 is None:
                                if self.verbose:
                                    self.stdout.write('[WARNING] Regel %s wordt overgeslagen (geen scores)' % row)
                            else:
                                self.stderr.write('[ERROR] Probleem met scores op regel %s: %s en %s' % (row, repr(score1), repr(score2)))
                        else:
                            if score1 > 250 or score2 > 250:
                                self.stderr.write('[ERROR] Te hoge scores op regel %s: %s en %s' % (row, score1, score2))

                            counts = list()
                            try:
                                if c10:
                                    c10 = int(c10)
                                    counts.append('%sx10' % c10)
                                else:
                                    c10 = 0
                                if c9:
                                    c9 = int(c9)
                                    counts.append('%sx9' % c9)
                                else:
                                    c9 = 0
                                if c8:
                                    c8 = int(c8)
                                    counts.append('%sx8' % c8)
                                else:
                                    c8 = 0
                            except (TypeError, ValueError) as err:
                                self.stderr.write('[ERROR] Probleem met 10/9/8 count op regel %s: %s' % (row, str(err)))
                                counts_str = ''
                            else:
                                counts_str = ", ".join(counts)
                                if c10 + c9 + c8 > (2 * 25):
                                    self.stderr.write('[ERROR] Te veel 10/9/8-en op regel %s: %s / %s / %s' % (row, c10, c9, c8))

                            totaal = score1 + score2
                            if totaal > 0:                  # soms wordt 0,0 ingevuld bij niet aanwezig
                                rank_doorlopend += 1
                                if totaal == prev_totaal and counts_str == prev_counts_str:
                                    # zelfde score, zelf rank
                                    pass
                                elif totaal > prev_totaal:
                                    self.stderr.write('[ERROR] Score is niet aflopend op regel %s' % row)
                                else:
                                    rank = rank_doorlopend

                                # counts worden alleen gebruikt voor plaats 1, 2, 3
                                # nog wel tonen voor plaats 3+ als deze relevant was
                                if rank > 3 and totaal != prev_totaal:
                                    toon_counts = False

                                if not toon_counts:
                                    counts_str = ''

                                prev_totaal = totaal
                                prev_counts_str = counts_str

                                if self.verbose:
                                    self.stdout.write('%s: %s, scores: %s %s %s' % (rank, deelnemer, score1, score2, counts_str))

                                opslaan = True
                                if dupe_check:
                                    # allow result_rank change
                                    if deelnemer.result_score_1 != score1 or deelnemer.result_score_2 != score2:
                                        opslaan = False
                                        diff = "%s,%s,%s -> %s,%s,%s" % (deelnemer.result_score_1,
                                                                         deelnemer.result_score_2,
                                                                         deelnemer.result_counts,
                                                                         score1, score2, counts_str)
                                        self.stderr.write('[ERROR] Deelnemer pk=%s heeft al andere resultaten! (%s): %s' % (
                                                            deelnemer.pk, deelnemer, diff))

                                    else:
                                        # allow counts to change from empty to non-empty
                                        if deelnemer.result_counts == '' and deelnemer.result_counts != counts_str:
                                            diff = "%s -> %s" % (deelnemer.result_counts, counts_str)
                                            self.stderr.write('[WARNING] Deelnemer pk=%s krijgt nieuwe result_counts (%s): %s' % (
                                                                deelnemer.pk, deelnemer, diff))
                                        elif deelnemer.result_counts != counts_str:
                                            opslaan = False
                                            diff = "%s,%s,%s -> %s,%s,%s" % (deelnemer.result_score_1,
                                                                             deelnemer.result_score_2,
                                                                             deelnemer.result_counts,
                                                                             score1, score2, counts_str)
                                            self.stderr.write('[ERROR] Deelnemer pk=%s heeft al andere resultaten! (%s): %s' % (
                                                                deelnemer.pk, deelnemer, diff))

                                if opslaan:
                                    deelnemer.result_rank = rank
                                    deelnemer.result_score_1 = score1
                                    deelnemer.result_score_2 = score2
                                    deelnemer.result_counts = counts_str

                        if not self.dryrun:
                            deelnemer.save(update_fields=['result_rank',
                                                          'result_score_1', 'result_score_2',
                                                          'result_counts'])
            else:
                nix_count += 1
        # while

        # zet deelnemers die niet meegedaan hebben op een hoog rank nummer
        for deelnemers in self.deelnemers.values():
            for deelnemer in deelnemers:
                if deelnemer.kampioenschap.pk == deelkamp_pk and deelnemer.indiv_klasse.pk == self.klasse_pk:
                    if deelnemer.result_rank in (0, KAMP_RANK_NO_SHOW, KAMP_RANK_RESERVE):
                        if deelnemer.deelname != DEELNAME_NEE:
                            # noteer: we weten niet welke reserves opgeroepen waren
                            # noteer: nog geen oplossing voor reserves die toch niet mee kunnen doen
                            if self.verbose:
                                self.stdout.write('[WARNING] Mogelijke no-show voor deelnemer %s' % deelnemer)
                            deelnemer.result_rank = KAMP_RANK_NO_SHOW
                        else:
                            deelnemer.result_rank = KAMP_RANK_RESERVE

                        if not self.dryrun:
                            deelnemer.save(update_fields=['result_rank'])
            # for
        # for

    def _read_html(self, fname):
        self.stdout.write('[INFO] Lees bestand %s' % repr(fname))
        try:
            data = open(fname).read()
        except OSError as exc:
            self.stderr.write('[ERROR] Kan het  bestand niet openen (%s)' % str(exc))
            return

        rank_doorlopend = 0
        rank_deze = 0
        prev_totaal = 0
        prev_counts_str = ''
        toon_counts = True
        klasse_titel = ''
        tabel = data[data.find('<table'):data.find('</table>')]
        for regel in tabel.split('</tr>'):
            if '</td>' in regel:
                parts = list()
                for part in regel.split('</td>'):
                    pos = part.rfind('>')
                    parts.append(part[pos + 1:])
                # for
                if len(parts) == 9:
                    # print('data: %s' % repr(parts))
                    rank, naam, ver, score1, score2, _, count10, count9, _ = parts
                    ver = ver[:4]
                    if self.verbose:
                        self.stdout.write('   %s %s %s %s %s %s %s' % (rank, ver, naam, score1, score2, count10, count9))

                    ver_nr = int(ver)
                    score1 += '/'
                    score1 = int(score1[:score1.find('/')])
                    score2 += '/'
                    score2 = int(score2[:score2.find('/')])
                    count10 = int(count10)
                    count9 = int(count9)
                    totaal = score1 + score2

                    matches = list()
                    up_naam = naam.upper()
                    for deelnemer in self.ver2deelnemers[ver_nr]:
                        match = 0
                        sporter = deelnemer.sporterboog.sporter
                        if sporter.voornaam.upper() in up_naam:
                            # warm
                            match = 3

                            # achternaam kan gehusseld zijn, dus kijk per woord
                            for part in sporter.achternaam.upper().split(' '):
                                if part in up_naam:
                                    match += 2
                                    # print('match: %s' % repr(part))
                            # for

                            if deelnemer.sporterboog.boogtype.beschrijving in klasse_titel:
                                match += 10

                            tup = (match, deelnemer.pk, deelnemer)
                            matches.append(tup)
                    # for

                    if len(matches) == 0:
                        self.stderr.write('[WARNING] Deelnemer %s niet gevonden!' % (repr(naam)))
                        for deelnemer in self.ver2deelnemers[ver_nr]:
                            self.stdout.write('[INFO] kandidaat? %s' % deelnemer)
                        # for
                    else:
                        matches.sort(reverse=True)      # hoogste eerst
                        # for match, _, deelnemer in matches:
                        #     print('   kandidaat: (match %s): %s' % (match, deelnemer))
                        _, _, deelnemer = matches[0]
                        # print(' deelnemer: %s' % deelnemer)

                        counts_str = "%sx10, %sx9" % (count10, count9)

                        rank_doorlopend += 1

                        # print('rank_doorlopend=%s, totaal=%s, prev_totaal=%s, counts_str=%s, prev_counts_str=%s' % (rank_doorlopend, totaal, prev_totaal, counts_str, prev_counts_str))

                        if totaal == prev_totaal:
                            if rank_deze > 3:
                                # we kijken niet meer naar counts_str
                                # zelfde totaal = zelfde rank
                                if self.verbose:
                                    self.stdout.write('[INFO] Zelfde score: %s == %s' % (totaal, prev_totaal))

                            elif counts_str == prev_counts_str:
                                # zelfde score --> zelf rank
                                if self.verbose:
                                    self.stdout.write('[INFO] Zelfde score: %s == %s en counts: %s == %s' % (
                                                        totaal, prev_totaal, counts_str, prev_counts_str))
                            else:
                                rank_deze = rank_doorlopend

                        elif totaal > prev_totaal and rank_doorlopend > 1:
                            self.stderr.write('[ERROR] Score is niet aflopend')

                        else:
                            # geef rank
                            rank_deze = rank_doorlopend

                        # counts worden alleen gebruikt voor plaats 1, 2, 3
                        # nog wel tonen voor plaats 3+ als deze relevant was
                        if rank_deze > 3 and totaal != prev_totaal:
                            toon_counts = False

                        if not toon_counts:
                            counts_str = ''

                        prev_totaal = totaal
                        prev_counts_str = counts_str

                        if rank == 'DNS':
                            deelnemer.result_rank = KAMP_RANK_NO_SHOW
                            deelnemer.result_score_1 = 0
                            deelnemer.result_score_2 = 0
                            deelnemer.result_counts = ''
                        else:
                            if str(rank_deze) != rank:
                                self.stdout.write('[WARNING] Afwijkende rank: %s vs %s' % (rank_deze, rank))

                            deelnemer.result_rank = rank_deze
                            deelnemer.result_score_1 = score1
                            deelnemer.result_score_2 = score2
                            deelnemer.result_counts = counts_str

                        if not self.dryrun:
                            deelnemer.save(update_fields=['result_rank', 'result_score_1', 'result_score_2', 'result_counts'])

            else:
                parts = list()
                for part in regel.split('</th>'):
                    pos = part.rfind('>')
                    parts.append(part[pos + 1:])
                # for
                # print('header: %s' % repr(parts))
                header = parts[0]
                if len(header) > 10:
                    self.stdout.write('[INFO] Klasse: %s' % repr(header))
                    klasse_titel = header
                    rank_doorlopend = rank_deze = 0
                    prev_totaal = 0
                    prev_counts_str = ''
                    toon_counts = True
        # for

        # zet deelnemers die niet meegedaan hebben op een hoog rank nummer
        for ver, deelnemers in self.ver2deelnemers.items():
            for deelnemer in deelnemers:
                if deelnemer.result_rank in (0, KAMP_RANK_NO_SHOW, KAMP_RANK_RESERVE):
                    if deelnemer.deelname != DEELNAME_NEE:
                        # noteer: we weten niet welke reserves opgeroepen waren
                        # noteer: nog geen oplossing voor reserves die toch niet mee kunnen doen
                        if self.verbose:
                            self.stdout.write('[WARNING] Mogelijke no-show voor deelnemer %s' % deelnemer)
                        deelnemer.result_rank = KAMP_RANK_NO_SHOW
                    else:
                        deelnemer.result_rank = KAMP_RANK_RESERVE

                    if not self.dryrun:
                        deelnemer.save(update_fields=['result_rank'])
            # for
        # for

    def handle(self, *args, **options):

        self.dryrun = options['dryrun']
        self.verbose = options['verbose']
        fname = options['bestand']

        self.deelnemers_ophalen()

        self._read_html(fname)

        if False:
            self.bepaal_klasse(ws)
            if self.klasse_pk:
                self.lees_uitslagen(ws)

# end of file
