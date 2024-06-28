# -*- coding: utf-8 -*-

#  Copyright (c) 2022-2024 Ramon van der Winkel.
#  All rights reserved.
#  Licensed under BSD-3-Clause-Clear. See LICENSE file for details.

# verwijder onnodige (oude) data van voorgaande competities

from django.core.management.base import BaseCommand
from Competitie.models import (Competitie, CompetitieIndivKlasse, CompetitieTeamKlasse,
                               KampioenschapSporterBoog, KampioenschapTeam)
import openpyxl
from openpyxl.styles import Alignment, Font, DEFAULT_FONT

"""
    Maak een excel met de BK kandidaten + teams, voor als het RK afgelast is
"""

FONT_NAME = 'Arial'
FONT_SIZE = 11

COL_WIDTH_VERENIGING = 30
COL_WIDTH_TEAM_NAAM = 30
COL_WIDTH_SPORTER_NAAM = 30
COL_WIDTH_GEMIDDELDE = 13


class Command(BaseCommand):
    help = "Maak excel met BK indiv + team (voor als RK afgelast is)"

    def __init__(self, stdout=None, stderr=None, no_color=False, force_color=False):
        super().__init__(stdout, stderr, no_color, force_color)
        self.comp = None

    def add_arguments(self, parser):
        parser.add_argument('afstand', type=int, help='Competitie afstand (18/25)')
        parser.add_argument('--teams', action='store_true', help='Teams klassen toevoegen')
        parser.add_argument('--indiv', action='store_true', help='Individuele klassen toevoegen')

    def _write_teams(self, wb):

        if self.comp.is_indoor():
            aantal_pijlen = 30
        else:
            aantal_pijlen = 25

        left_align = Alignment(horizontal='left')
        right_align = Alignment(horizontal='right')
        font_header = Font(name=FONT_NAME, size=FONT_SIZE, bold=True)
        font_titel = Font(name=FONT_NAME, size=FONT_SIZE + 2, bold=True)

        for klasse in (CompetitieTeamKlasse
                       .objects
                       .filter(competitie=self.comp,
                               is_voor_teams_rk_bk=True)
                       .order_by('volgorde')):

            self.stdout.write('[INFO] Team klasse: %s' % klasse)
            nr2ver = dict()  # [ver_nr] = vereniging

            ws = wb.create_sheet(title=klasse.beschrijving[:31].strip())     # sheet title max length = 31
            ws['A1'] = 'Team klasse: %s' % klasse.beschrijving
            ws['A1'].font = font_titel

            ws.column_dimensions['C'].width = COL_WIDTH_VERENIGING
            ws.column_dimensions['D'].width = COL_WIDTH_TEAM_NAAM
            ws.column_dimensions['F'].width = COL_WIDTH_SPORTER_NAAM
            ws.column_dimensions['G'].width = COL_WIDTH_GEMIDDELDE
            ws.column_dimensions['H'].width = COL_WIDTH_GEMIDDELDE

            ws['A3'] = 'Nr'
            ws['B3'] = 'Ver nr'
            ws['C3'] = 'Vereniging'
            ws['D3'] = 'Team naam'
            ws['E3'] = 'Lid nr'
            ws['F3'] = 'Naam'
            ws['G3'] = 'Gemiddelde'
            ws['H3'] = 'Team sterkte'

            # ws['G3'].alignment = left_align
            ws['H3'].alignment = left_align

            ws['A3'].font = font_header
            ws['B3'].font = font_header
            ws['C3'].font = font_header
            ws['D3'].font = font_header
            ws['E3'].font = font_header
            ws['F3'].font = font_header
            ws['G3'].font = font_header
            ws['H3'].font = font_header

            row_nr = 4
            team_nr = 0
            for team in (KampioenschapTeam
                         .objects
                         .filter(kampioenschap__competitie=self.comp,
                                 team_klasse=klasse)
                         .select_related('vereniging')
                         .prefetch_related('gekoppelde_leden')
                         .order_by('-aanvangsgemiddelde')):

                row_str = str(row_nr)

                team_nr += 1
                ws['A' + row_str] = team_nr

                ver = team.vereniging
                nr2ver[ver.ver_nr] = ver

                ws['B' + row_str] = ver.ver_nr
                ws['C' + row_str] = ver.naam
                ws['D' + row_str] = team.team_naam
                ws['H' + row_str] = team.aanvangsgemiddelde * aantal_pijlen

                ws['A' + row_str].alignment = left_align
                ws['B' + row_str].alignment = left_align
                ws['H' + row_str].number_format = '000.0'
                ws['H' + row_str].alignment = left_align

                row_nr += 1

                # schrijf de sporters op
                for deelnemer in (team
                                  .gekoppelde_leden
                                  .select_related('sporterboog__sporter')
                                  .order_by('-gemiddelde')):        # hoogste eerst

                    row_str = str(row_nr)

                    sporter = deelnemer.sporterboog.sporter

                    ws['E' + row_str] = sporter.lid_nr
                    ws['F' + row_str] = sporter.volledige_naam()
                    ws['G' + row_str] = deelnemer.gemiddelde

                    ws['E' + row_str].alignment = left_align
                    ws['G' + row_str].number_format = '0.000'
                    ws['G' + row_str].alignment = left_align

                    row_nr += 1
                # for

                row_nr += 1     # lege regel
            # for

            row_str = str(row_nr)
            ws['A' + row_str] = 'Toegestane invallers per vereniging'
            ws['A' + row_str].font = font_titel
            row_nr += 1

            row_nr += 1     # lege regel

            # voeg de toegestane invallers toe
            row_str = str(row_nr)

            ws['B' + row_str] = 'Ver nr'
            ws['C' + row_str] = 'Vereniging'
            ws['E' + row_str] = 'Lid nr'
            ws['F' + row_str] = 'Naam'
            ws['G' + row_str] = 'Gemiddelde'
            ws['H' + row_str] = 'Boog type'

            # ws['G' + row_str].alignment = left_align

            ws['B' + row_str].font = font_header
            ws['C' + row_str].font = font_header
            ws['E' + row_str].font = font_header
            ws['F' + row_str].font = font_header
            ws['G' + row_str].font = font_header
            ws['H' + row_str].font = font_header

            row_nr += 1

            for ver_nr, ver in nr2ver.items():

                # vereniging details
                row_str = str(row_nr)

                ws['B' + row_str] = ver.ver_nr
                ws['C' + row_str] = ver.naam

                ws['B' + row_str].alignment = left_align

                row_nr += 1

                # sporter details
                for deelnemer in (KampioenschapSporterBoog
                                  .objects
                                  .filter(kampioenschap__competitie=self.comp,
                                          bij_vereniging=ver)
                                  .select_related('sporterboog__sporter',
                                                  'sporterboog__boogtype')
                                  .order_by('-gemiddelde')):
                    row_str = str(row_nr)

                    sporter = deelnemer.sporterboog.sporter
                    ws['E' + row_str] = sporter.lid_nr
                    ws['F' + row_str] = sporter.volledige_naam()
                    ws['G' + row_str] = deelnemer.gemiddelde
                    ws['H' + row_str] = deelnemer.sporterboog.boogtype.beschrijving

                    ws['E' + row_str].alignment = left_align
                    ws['G' + row_str].number_format = '0.000'
                    ws['G' + row_str].alignment = left_align

                    row_nr += 1
                # for
            # for

        # for (klasse)

    def _write_indiv(self, wb):

        left_align = Alignment(horizontal='left')
        right_align = Alignment(horizontal='right')
        font_header = Font(name=FONT_NAME, size=FONT_SIZE, bold=True)
        font_titel = Font(name=FONT_NAME, size=FONT_SIZE + 2, bold=True)

        for klasse in (CompetitieIndivKlasse
                       .objects
                       .filter(competitie=self.comp)
                       .exclude(is_ook_voor_rk_bk=False)        # aspiranten en klasse onbekend
                       .order_by('volgorde')):

            self.stdout.write('[INFO] Individuele klasse: %s' % klasse)

            ws = wb.create_sheet(title=klasse.beschrijving[:31].strip())     # sheet title max length = 31
            ws['A1'] = 'Individuele klasse: %s' % klasse.beschrijving
            ws['A1'].font = font_titel

            ws['A3'] = 'Nr'
            ws['B3'] = 'Lid nr'
            ws['C3'] = 'Naam'
            ws['D3'] = 'Ver nr'
            ws['E3'] = 'Vereniging'
            ws['F3'] = 'Gemiddelde'
            ws['G3'] = 'Notitie'

            # ws['F3'].alignment = right_align

            ws['A3'].font = font_header
            ws['B3'].font = font_header
            ws['C3'].font = font_header
            ws['D3'].font = font_header
            ws['E3'].font = font_header
            ws['F3'].font = font_header
            ws['G3'].font = font_header

            ws.column_dimensions['C'].width = COL_WIDTH_VERENIGING
            ws.column_dimensions['E'].width = COL_WIDTH_SPORTER_NAAM
            ws.column_dimensions['F'].width = COL_WIDTH_GEMIDDELDE

            row_nr = 4
            nr = 0
            for deelnemer in (KampioenschapSporterBoog
                              .objects
                              .filter(kampioenschap__competitie=self.comp,
                                      indiv_klasse=klasse)
                              .select_related('sporterboog__sporter',
                                              'bij_vereniging')
                              .order_by('-gemiddelde')):

                row_str = str(row_nr)

                nr += 1
                ws['A' + row_str] = nr

                sporter = deelnemer.sporterboog.sporter
                ws['B' + row_str] = sporter.lid_nr
                ws['C' + row_str] = sporter.volledige_naam()

                ver = deelnemer.bij_vereniging
                ws['D' + row_str] = ver.ver_nr
                ws['E' + row_str] = ver.naam

                ws['F' + row_str] = deelnemer.gemiddelde

                ws['A' + row_str].alignment = left_align
                ws['B' + row_str].alignment = left_align
                ws['D' + row_str].alignment = left_align
                ws['F' + row_str].number_format = '0.000'
                ws['F' + row_str].alignment = left_align

                if deelnemer.kampioen_label:
                    ws['G' + row_str] = deelnemer.kampioen_label

                row_nr += 1
            # for
        # for

    def handle(self, *args, **options):
        afstand = options['afstand']

        if not (options['teams'] or options['indiv']):
            self.stderr.write('[ERROR] Verplicht: --indiv en/of --teams')
            return

        qset = Competitie.objects.filter(is_afgesloten=False, afstand=afstand).order_by('begin_jaar')
        if qset.count() == 0:
            self.stderr.write('[ERROR] Geen competitie beschikbaar')
            return

        self.comp = qset[0]
        self.stdout.write('[INFO] Geselecteerde competitie: %s' % self.comp)

        DEFAULT_FONT.name = FONT_NAME
        DEFAULT_FONT.size = FONT_SIZE

        wb = openpyxl.Workbook()
        default_sheet = wb.active

        if options['teams']:
            self._write_teams(wb)

        if options['indiv']:
            self._write_indiv(wb)

        wb.remove(default_sheet)

        fname = 'bk_lijst.xlsx'
        self.stdout.write('[INFO] Schrijf bestand: %s' % fname)
        wb.save(fname)


# end of file
