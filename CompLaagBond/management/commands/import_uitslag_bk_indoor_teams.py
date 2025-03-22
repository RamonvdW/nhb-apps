# -*- coding: utf-8 -*-

#  Copyright (c) 2023-2024 Ramon van der Winkel.
#  All rights reserved.
#  Licensed under BSD-3-Clause-Clear. See LICENSE file for details.

from django.core.management.base import BaseCommand
from Competitie.definities import DEEL_RK, DEEL_BK
from Competitie.models import KampioenschapSporterBoog, KampioenschapTeam
from openpyxl.utils.exceptions import InvalidFileException
from decimal import Decimal
import openpyxl
import zipfile


class Command(BaseCommand):
    help = "Importeer uitslag BK Indoor Teams"

    def __init__(self, stdout=None, stderr=None, no_color=False, force_color=False):
        super().__init__(stdout, stderr, no_color, force_color)
        self.dryrun = True
        self.verbose = False
        self.pk2deelnemer = dict()                  # [lid_nr] = [KampioenschapSporterBoog, ...]
        self.teams_cache = list()                   # [KampioenschapTeam, ...]
        self.team_gekoppelde_deelnemers = dict()    # [team.pk] = [(deelnemer.gemiddelde, deelnemer.pk), ...]
        self.ver_kamp_pks = dict()                  # [ver_nr] = [KampioenschapSporterBoog.pk, ...]
        self.kamp_lid_nrs = list()                  # [lid_nr, ...]   iedereen die geplaatst is voor de kampioenschappen
        self.deelnemende_teams = dict()             # [team naam] = KampioenschapTeam
        self.team_klasse = None
        self.toegestane_bogen = list()

    def add_arguments(self, parser):
        parser.add_argument('--dryrun', action='store_true')
        parser.add_argument('--verbose', action='store_true')
        parser.add_argument('bestand', type=str,
                            help='Pad naar het Excel bestand')

    def _zet_team_klasse(self, klasse):
        self.team_klasse = klasse
        self.toegestane_bogen = list(klasse.boog_typen.values_list('afkorting', flat=True))
        # self.stdout.write('[DEBUG] toegestane bogen: %s' % repr(self.toegestane_bogen))

    def _deelnemers_ophalen(self):
        # alle RK (!) sporters mogen deelnemer aan een BK teams
        for deelnemer in (KampioenschapSporterBoog
                          .objects
                          .filter(kampioenschap__competitie__afstand='18',
                                  kampioenschap__deel=DEEL_RK)
                          .select_related('kampioenschap',
                                          'kampioenschap__rayon',
                                          'sporterboog__sporter',
                                          'sporterboog__boogtype',
                                          'indiv_klasse')):

            self.pk2deelnemer[deelnemer.pk] = deelnemer

            lid_nr = deelnemer.lid_nr = deelnemer.sporterboog.sporter.lid_nr
            ver_nr = deelnemer.bij_vereniging.ver_nr

            try:
                self.ver_kamp_pks[ver_nr].append(deelnemer.pk)
            except KeyError:
                self.ver_kamp_pks[ver_nr] = [deelnemer.pk]

            self.kamp_lid_nrs.append(lid_nr)
        # for

    def _teams_ophalen(self):
        for team in (KampioenschapTeam
                     .objects
                     .filter(kampioenschap__competitie__afstand='18',
                             kampioenschap__deel=DEEL_BK)
                     .select_related('kampioenschap',
                                     'kampioenschap__rayon',
                                     'vereniging',
                                     'team_type',
                                     'team_klasse')
                     .prefetch_related('gekoppelde_leden',
                                       'feitelijke_leden')):

            self.teams_cache.append(team)

            deelnemers = [(deelnemer.gemiddelde, deelnemer.pk)
                          for deelnemer in team.gekoppelde_leden.all()]
            deelnemers.sort(reverse=True)       # hoogste eerst

            self.team_gekoppelde_deelnemers[team.pk] = deelnemers
        # for

    def _sort_op_gemiddelde(self, lid_nrs):
        gem = list()
        for lid_nr in lid_nrs:
            deelnemer_all = [deelnemer
                             for deelnemer in self.deelnemers[lid_nr]
                             if deelnemer.sporterboog.boogtype.afkorting in self.toegestane_bogen]

            if len(deelnemer_all) == 1:
                deelnemer = deelnemer_all[0]
            else:
                self.stderr.write('[WARNING] TODO: bepaal juiste deelnemer uit %s' % repr(deelnemer_all))
                deelnemer = deelnemer_all[0]
            tup = (deelnemer.gemiddelde, lid_nr)
            gem.append(tup)
        # for
        gem.sort(reverse=True)
        return gem

    def _get_deelnemer(self, ver_nr, lid_nr, lid_ag):
        for pk in self.ver_kamp_pks[ver_nr]:
            deelnemer = self.pk2deelnemer[pk]
            if deelnemer.lid_nr == lid_nr and abs(deelnemer.gemiddelde - lid_ag) < 0.0001:
                return deelnemer
        # for

        # niet gevonden

        if lid_nr not in self.kamp_lid_nrs:
            self.stderr.write('[ERROR] Lid %s is niet gekwalificeerd voor dit kampioenschap!' % lid_nr)
            return None

        # misschien van een andere vereniging?
        self.stderr.write('[ERROR] Kan vereniging %s, lid %s met ag=%.3f niet vinden' % (ver_nr, lid_nr, lid_ag))
        return None

    def _get_team(self, team_naam, ver_nr, row_nr):
        if self.verbose:
            self.stdout.write('[DEBUG] get_team: team_klasse=%s' % repr(self.team_klasse))

        up_naam = team_naam.upper()
        sel_teams = list()
        for team in self.teams_cache:
            # self.stdout.write('[DEBUG] cached team: %s' % repr(team))
            if team.vereniging.ver_nr == ver_nr:
                if team.team_naam.upper() == up_naam:
                    if self.team_klasse is None or team.team_klasse == self.team_klasse:
                        sel_teams.append(team)
        # for

        if len(sel_teams) == 0:
            # niet gevonden, dus waarschijnlijk is de naam aangepast
            for team in self.teams_cache:
                # self.stdout.write('[DEBUG] cached team: %s' % repr(team))
                if team.vereniging.ver_nr == ver_nr:
                    team_up_naam = team.team_naam.upper()
                    if team_up_naam in up_naam or up_naam in team_up_naam:
                        if self.team_klasse is None or team.team_klasse == self.team_klasse:
                            sel_teams.append(team)
                            self.stdout.write('[WARNING] Aangepaste team naam: %s --> %s' % (
                                                repr(team.team_naam), repr(team_naam)))
            # for

        kamp_team = None
        if len(sel_teams) == 1:
            kamp_team = sel_teams[0]
        elif len(sel_teams) > 1:
            self.stderr.write('[ERROR] Kan team %s van vereniging %s op regel %s niet kiezen uit\n%s' % (
                repr(team_naam), ver_nr, row_nr, "\n".join([str(team) for team in sel_teams])))

        if kamp_team is None:
            self.stderr.write('[ERROR] Kan team %s van vereniging %s op regel %s niet vinden' % (
                repr(team_naam), ver_nr, row_nr))

        return kamp_team

    def _bepaal_feitelijke_deelnemers(self, ver_nr: int, kamp_team: KampioenschapTeam, gevonden_deelnemers):

        gekoppelde_deelnemers = self.team_gekoppelde_deelnemers[kamp_team.pk]
        aantal_gekoppeld = len(gekoppelde_deelnemers)

        feitelijke_deelnemers = list()

        # haal
        klopt = list()
        for tup in gevonden_deelnemers:
            if tup in gekoppelde_deelnemers:
                klopt.append(tup)
        # for
        for tup in klopt:
            gevonden_deelnemers.remove(tup)
            gekoppelde_deelnemers.remove(tup)

            _, pk = tup
            deelnemer = self.pk2deelnemer[pk]
            feitelijke_deelnemers.append(deelnemer)
        # for

        if len(gekoppelde_deelnemers) == len(gevonden_deelnemers) == 0:
            # alles is verwerkt
            self.stdout.write('[DEBUG]   Feitelijke deelnemers: %s' % repr([deelnemer.lid_nr
                                                                            for deelnemer in feitelijke_deelnemers]))
            return feitelijke_deelnemers

        uitvallers = gekoppelde_deelnemers
        invallers = gevonden_deelnemers

        self.stdout.write('[DEBUG]   Feitelijke deelnemers: %s' % repr([deelnemer.lid_nr
                                                                        for deelnemer in feitelijke_deelnemers]))
        self.stdout.write('[DEBUG]   Uitvallers deelnemers: %s' % repr(uitvallers))
        self.stdout.write('[DEBUG]   Invallers  deelnemers: %s' % repr(invallers))

        if len(invallers) > len(uitvallers):
            self.stderr.write('[ERROR] Te veel invallers voor team %s met max %s sporters (ver: %s)' % (
                                repr(kamp_team.team_naam), aantal_gekoppeld, ver_nr))
            feitelijke_deelnemers = list()
            return feitelijke_deelnemers

        while len(invallers) > 0:
            gemiddelde_in, pk_in = invallers.pop(0)
            gemiddelde_uit, pk_uit = uitvallers[0]
            deelnemer_in = self.pk2deelnemer[pk_in]

            if gemiddelde_in > gemiddelde_uit:
                # mag niet invallen
                self.stderr.write(
                    '[ERROR] Te hoog gemiddelde %s voor invaller %s voor team %s van vereniging %s' % (
                            gemiddelde_in, deelnemer_in.lid_nr, kamp_team.team_naam, ver_nr))
                for gemiddelde, pk in uitvallers:
                    deelnemer_uit = self.pk2deelnemer[pk]
                    self.stderr.write('        Uitvaller %s heeft gemiddelde %s' % (deelnemer_uit.lid_nr, gemiddelde))
            else:
                # mag wel invallen en vervangt de hoogste uitvaller
                feitelijke_deelnemers.append(deelnemer_in)
                uitvallers.pop(0)
        # while

        self.stdout.write('[DEBUG]   Feitelijke deelnemers: %s' % repr([deelnemer.lid_nr
                                                                        for deelnemer in feitelijke_deelnemers]))

        if len(feitelijke_deelnemers) < 3:
            self.stderr.write('[ERROR]   Maar %s deelnemers in team %s' % (len(feitelijke_deelnemers),
                                                                           kamp_team.team_naam))

        return feitelijke_deelnemers

    def _importeer_voorronde(self, prg):
        blad = 'Deelnemers en Scores'
        try:
            ws = prg[blad]
        except KeyError:            # pragma: no cover
            self.stderr.write('[ERROR] Kan blad %s niet vinden' % repr(blad))
            return

        col_ver_naam = 'D'
        col_team_naam = 'F'
        col_lid_nr = 'E'
        col_lid_ag = 'G'
        col_score1 = 'H'
        col_score2 = 'I'

        # doorloop alle regels van het excel blad en ga op zoek naar bondsnummers
        row_nr = 9 - 1
        nix_count = 0
        kamp_teams = list()
        while nix_count < 10:
            row_nr += 1
            row = str(row_nr)

            # vind een vereniging + team naam
            ver_naam = ws[col_ver_naam + row].value
            team_naam = ws[col_team_naam + row].value

            if ver_naam is None:
                nix_count += 1
                continue

            nix_count = 0
            ver_nr = -1

            try:
                if ver_naam[0] == '[' and ver_naam[5:5+2] == '] ':
                    ver_nr = int(ver_naam[1:1+4])
            except ValueError:
                pass

            if self.verbose:
                self.stdout.write('[DEBUG] regel %s: ver_nr=%s, ver_naam=%s, team_naam=%s' % (
                                    row, ver_nr, repr(ver_naam), repr(team_naam)))

            if ver_nr < 0 or team_naam is None:
                continue

            # zoek het team erbij
            kamp_team = self._get_team(team_naam, ver_nr, row_nr)
            if self.verbose:
                self.stdout.write('[DEBUG] row_nr=%s, kamp_team=%s' % (row_nr, kamp_team))
            if kamp_team is None:
                continue

            if self.team_klasse is None:
                self._zet_team_klasse(kamp_team.team_klasse)
            else:
                if self.team_klasse != kamp_team.team_klasse:
                    self.stderr.write('[ERROR] Inconsistente team klasse op regel %s: %s (eerdere teams: %s)' % (
                                        row_nr, kamp_team.team_klasse, self.team_klasse))
                    continue

            self.deelnemende_teams[team_naam] = kamp_team

            # haal de teamleden op
            gevonden_deelnemers = list()  # [(deelnemer.gemiddelde, deelnemer.pk), ..]
            for lp in range(10):            # pragma: no branch
                row_nr += 1
                row = str(row_nr)

                lid_nr = ws[col_lid_nr + row].value
                if self.verbose:
                    self.stdout.write('[DEBUG] row_nr=%s, lid_nr=%s' % (row_nr, repr(lid_nr)))
                if lid_nr is None:
                    break
                if str(lid_nr) == '-':
                    continue

                try:
                    lid_ag = round(Decimal(ws[col_lid_ag + row].value), 3)  # 3 cijfers achter de komma
                    score1 = int(ws[col_score1 + row].value)
                    score2 = int(ws[col_score2 + row].value)
                except (ValueError, TypeError):
                    self.stderr.write('[ERROR] Probleem met de scores op regel %s' % row_nr)
                else:
                    if score1 + score2 == 0:
                        # sporter heeft niet meegedaan
                        self.stdout.write('[WARNING] Geen scores voor sporter %s op regel %s' % (lid_nr, row_nr))
                    else:
                        deelnemer = self._get_deelnemer(ver_nr, lid_nr, lid_ag)
                        if deelnemer:
                            deelnemer.result_bk_teamscore_1 = score1
                            deelnemer.result_bk_teamscore_2 = score2
                            tup = (deelnemer.gemiddelde, deelnemer.pk)
                            gevonden_deelnemers.append(tup)
            # for

            # moeten nu de lijst van feitelijke deelnemers controleren tegen de aan het team gekoppelde deelnemers
            gevonden_deelnemers.sort(reverse=True)        # hoogste eerst

            feitelijke_deelnemers = self._bepaal_feitelijke_deelnemers(ver_nr, kamp_team, gevonden_deelnemers)

            if len(feitelijke_deelnemers) == 0:
                self.stdout.write('[WARNING] Team %s van vereniging %s heeft niet meegedaan (geen scores)' % (
                                        ver_nr, team_naam))
                del self.deelnemende_teams[team_naam]
            else:
                if not self.dryrun:
                    kamp_team.feitelijke_leden.set(feitelijke_deelnemers)

                    # bereken de team score
                    deelnemer_totalen = list()
                    for deelnemer in feitelijke_deelnemers:
                        if not self.dryrun:
                            # uitgestelde save actie
                            deelnemer.save(update_fields=['result_bk_teamscore_1', 'result_bk_teamscore_2'])
                        deelnemer_totaal = deelnemer.result_bk_teamscore_1 + deelnemer.result_bk_teamscore_2
                        deelnemer_totalen.append(deelnemer_totaal)
                    # for

                    deelnemer_totalen.sort(reverse=True)                        # hoogste eerst
                    kamp_team.result_teamscore = sum(deelnemer_totalen[:3])     # de 3 hoogste gebruiken

                    deelnemer_totalen.insert(0, kamp_team.pk)                   # backup voor sorteren
                    deelnemer_totalen.insert(0, kamp_team.result_teamscore)     # resultaat vooraan
                    deelnemer_totalen.append(kamp_team)                         # team record laatst
                    kamp_teams.append(deelnemer_totalen)
        # while

        kamp_teams.sort(reverse=True)               # hoogste eerste
        rank = 0
        for tup in kamp_teams:
            kamp_team = tup[-1]
            if kamp_team.result_teamscore > 0:
                rank += 1
                kamp_team.result_rank = rank
                kamp_team.result_volgorde = rank
            else:
                kamp_team.result_rank = 0
                kamp_team.result_volgorde = 0
            if not self.dryrun:
                kamp_team.save(update_fields=['result_rank', 'result_volgorde', 'result_teamscore'])
        # for

    def _importeer_uitslag(self, prg):
        """ Lees het blad Uitslag in om de correcte volgorde van de teams te krijgen
            Hierop staat het resultaat van de shoot-off voor deelname aan de finale.
        """
        blad = 'Uitslag'
        try:
            ws = prg[blad]
        except KeyError:            # pragma: no cover
            self.stderr.write('[ERROR] Kan blad %s niet vinden' % repr(blad))
            return

        # TODO: kolommen vereniging en team zijn omgedraaid (ook in template en test files)
        col_ver_naam = 'D'
        col_team_naam = 'C'
        col_resultaat = 'V'

        kamp_teams = list()

        # doorloop alle regels van het excel blad en ga op zoek naar bondsnummers
        row_nr = 9 - 1

        # check de header
        header = ws['B' + str(row_nr)].value
        if header != 'Baan':            # pragma: no cover
            self.stderr.write('[ERROR] Blad is niet in orde: cell B8 bevat niet "Baan" maar %s' % repr(header))
            raise ValueError('Uitslag B8')

        nix_count = 0
        while nix_count < 5:
            row_nr += 1
            row = str(row_nr)

            # vind een vereniging + team naam
            ver_naam = ws[col_ver_naam + row].value
            team_naam = ws[col_team_naam + row].value
            resultaat = ws[col_resultaat + row].value

            if ver_naam is None:
                nix_count += 1
                continue

            if self.verbose:
                self.stdout.write('[DEBUG] Uitslag: ver_naam=%s, team_naam=%s, resultaat=%s' % (
                                    repr(ver_naam), repr(team_naam), repr(resultaat)))

            ver_nr = -1
            ver_naam = str(ver_naam)
            try:
                if ver_naam[0] == '[' and ver_naam[5:5+2] == '] ':
                    ver_nr = int(ver_naam[1:1+4])
            except ValueError:          # pragma: no cover
                pass

            if ver_nr < 0:
                continue

            if team_naam is None:       # pragma: no cover
                continue

            nix_count = 0

            # zoek het team erbij
            kamp_team = self._get_team(team_naam, ver_nr, row_nr)
            if kamp_team is None:
                continue

            tup = (resultaat, kamp_team.pk, kamp_team)
            kamp_teams.append(tup)
        # while

        kamp_teams.sort(reverse=True)               # hoogste eerste
        rank = 0
        for tup in kamp_teams:
            kamp_team = tup[-1]
            if kamp_team.result_teamscore > 0:
                rank += 1
                kamp_team.result_rank = rank
                kamp_team.result_volgorde = rank
            else:
                kamp_team.result_rank = 0
                kamp_team.result_volgorde = 0
            if not self.dryrun:
                kamp_team.save(update_fields=['result_rank', 'result_volgorde'])
        # for

    def _zet_rank_en_volgorde(self, goud, zilver, brons, vierde, vijfden):

        if self.verbose:
            self.stdout.write('[DEBUG] goud: %s' % repr(goud))
            self.stdout.write('[DEBUG] zilver: %s' % repr(zilver))
            self.stdout.write('[DEBUG] brons: %s' % repr(brons))
            self.stdout.write('[DEBUG] vierde: %s' % repr(vierde))
            self.stdout.write('[DEBUG] vijfden: %s' % repr(vijfden))

        rank = 1
        for team_naam in (goud, zilver, brons, vierde):
            if team_naam:
                try:
                    kamp_team = self.deelnemende_teams[team_naam]
                except KeyError:
                    self.stderr.write('[ERROR] Kan team %s niet vinden!' % repr(team_naam))
                else:
                    kamp_team.result_rank = rank
                    kamp_team.result_volgorde = rank
                    if not self.dryrun:
                        kamp_team.save(update_fields=['result_rank', 'result_volgorde'])
            rank += 1
        # for

        # de rest (van de 8 finalisten) is 5e
        # result_volgorde is al gezet, gebaseerd op aflopende scores
        for team_naam in vijfden:
            try:
                kamp_team = self.deelnemende_teams[team_naam]
            except KeyError:
                self.stderr.write('[ERROR] Kan team %s niet vinden!' % repr(team_naam))
            else:
                kamp_team.result_rank = 5
                if not self.dryrun:
                    kamp_team.save(update_fields=['result_rank'])
        # for

    @staticmethod
    def _lees_team_naam(ws, cell):
        team_naam = ws[cell].value
        if team_naam in (None, 'n.v.t.', 'nvt', 'n.v.t', 'BYE'):
            team_naam = ''
        return team_naam

    @staticmethod
    def _kies_team_a_b(team_naam_a, team_naam_b, team_a_is_tweede):
        if team_a_is_tweede:
            return team_naam_b, team_naam_a
        return team_naam_a, team_naam_b

    def _importeer_finales_8(self, ws):
        """ Lees de uitslag van het blad 'Finales 8 teams' """

        final_8 = list()
        for row_nr in (12, 14, 18, 20, 24, 26, 30, 32):
            team_naam = self._lees_team_naam(ws, 'B' + str(row_nr))
            if team_naam in self.deelnemende_teams.keys():
                final_8.append(team_naam)
            else:
                self.stdout.write('[WARNING] Kwartfinale team %s op finale blad wordt niet herkend' % repr(team_naam))
        # for
        # self.stdout.write('final_8: %s' % repr(final_8))

        # gouden finale
        team_naam_1 = self._lees_team_naam(ws, 'T19')
        team_naam_2 = self._lees_team_naam(ws, 'T21')
        goud, zilver = self._kies_team_a_b(team_naam_1, team_naam_2, ws['Z19'].value == '2e')

        # bronzen finale
        team_naam_1 = self._lees_team_naam(ws, 'T31')
        team_naam_2 = self._lees_team_naam(ws, 'T33')
        brons, vierde = self._kies_team_a_b(team_naam_1, team_naam_2, ws['Z31'].value != '3e')

        for team_naam in (goud, zilver, brons, vierde):
            if team_naam in final_8:            # pragma: no branch
                final_8.remove(team_naam)
        # for

        self._zet_rank_en_volgorde(goud, zilver, brons, vierde, final_8)

    def _importeer_finales_4(self, ws):
        """ Lees de uitslag van het blad 'Finales 4 teams' """

        # gouden finale
        team_naam_1 = self._lees_team_naam(ws, 'L15')
        team_naam_2 = self._lees_team_naam(ws, 'L17')
        goud, zilver = self._kies_team_a_b(team_naam_1, team_naam_2, ws['R15'].value == '2e')

        # bronzen finale
        team_naam_1 = self._lees_team_naam(ws, 'L25')
        team_naam_2 = self._lees_team_naam(ws, 'L27')
        brons, vierde = self._kies_team_a_b(team_naam_1, team_naam_2, ws['R25'].value != '3e')

        self._zet_rank_en_volgorde(goud, zilver, brons, vierde, [])

    def handle(self, *args, **options):

        self.dryrun = options['dryrun']
        self.verbose = options['verbose']

        # open de kopie, zodat we die aan kunnen passen
        fname = options['bestand']
        self.stdout.write('[INFO] Lees bestand %s' % repr(fname))
        try:
            prg = openpyxl.load_workbook(fname,
                                         data_only=True)        # do not evaluate formulas; use last calculated values
        except (OSError, zipfile.BadZipFile, KeyError, InvalidFileException) as exc:
            self.stderr.write('[ERROR] Kan het excel bestand niet openen (%s)' % str(exc))
            return

        self._deelnemers_ophalen()
        self._teams_ophalen()

        self._importeer_voorronde(prg)
        self._importeer_uitslag(prg)

        if len(self.deelnemende_teams) == 0:
            self.stdout.write('[WARNING] Geen deelnemende teams, dus geen kampioen')
        else:
            winnaar = None
            if "ERE" in self.team_klasse.beschrijving:
                blad = 'Finales 8 teams'
                try:
                    ws = prg[blad]
                except KeyError:        # pragma: no cover
                    self.stderr.write('[ERROR] Kan blad %s niet vinden' % repr(blad))
                    return

                winnaar = ws['Z19'].value
                if winnaar:
                    self.stdout.write('[INFO] Finales worden van blad %s gehaald' % repr(blad))
                    self._importeer_finales_8(ws)

            if not winnaar:
                # None or leeg
                # dit blad is niet gebruikt - probeer het blad met finales met 4 teams
                blad = 'Finales 4 teams'
                try:
                    ws = prg[blad]
                except KeyError:            # pragma: no cover
                    self.stderr.write('[ERROR] Kan blad %s niet vinden' % repr(blad))
                    return

                winnaar = ws['R15'].value
                if not winnaar:     # pragma: no cover
                    # None or leeg
                    self.stderr.write('[ERROR] Kan juiste finale blad niet bepalen (geen WINNAAR)')
                    return

                self.stdout.write('[INFO] Finales worden van blad %s gehaald' % repr(blad))
                self._importeer_finales_4(ws)

            result = list()
            for team_naam, team in self.deelnemende_teams.items():
                tup = (team.result_rank, team.result_volgorde, team_naam, team.result_teamscore)
                result.append(tup)
            # for
            if self.verbose:
                result.sort()
                self.stdout.write('Resultaat:')
                for rank, volgorde, naam, result_teamscore in result:
                    self.stdout.write('%s %s %s (score: %s)' % (rank, volgorde, naam, result_teamscore))

# end of file
