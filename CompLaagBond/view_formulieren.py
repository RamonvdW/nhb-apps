# -*- coding: utf-8 -*-

#  Copyright (c) 2023-2026 Ramon van der Winkel.
#  All rights reserved.
#  Licensed under BSD-3-Clause-Clear. See LICENSE file for details.

from django.conf import settings
from django.urls import reverse
from django.http import Http404, HttpResponse
from django.utils import timezone
from django.views.generic import TemplateView
from django.utils.safestring import mark_safe
from django.contrib.auth.mixins import UserPassesTestMixin
from Competitie.definities import DEEL_BK, DEEL_RK, DEELNAME_NEE
from Competitie.models import (CompetitieMatch, CompetitieTeamKlasse,
                               Kampioenschap, KampioenschapSporterBoog, KampioenschapTeam,
                               KampioenschapTeamKlasseLimiet)
from CompKampioenschap.operations import get_url_wedstrijdformulier
from Functie.definities import Rol
from Functie.rol import rol_get_huidige_functie
from Sporter.models import SporterVoorkeuren
from tempfile import NamedTemporaryFile
from types import SimpleNamespace
from copy import copy
import openpyxl
import zipfile
import shutil
import os


TEMPLATE_DOWNLOAD_BK_FORMULIEREN = 'complaagbond/bko-download-bk-formulieren.dtl'

CONTENT_TYPE_XLSX = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'


class FormulierBkTeamsAlsBestandView(UserPassesTestMixin, TemplateView):

    """ Geef de BKO het ingevulde wedstrijdformulier voor een BK wedstrijd """

    # class variables shared by all instances
    raise_exception = True  # genereer PermissionDefined als test_func False terug geeft
    permission_denied_message = 'Geen toegang'

    def test_func(self):
        """ called by the UserPassesTestMixin to verify the user has permissions to use this view """
        rol_nu, functie_nu = rol_get_huidige_functie(self.request)
        return functie_nu and rol_nu == Rol.ROL_BKO

    def get(self, request, *args, **kwargs):
        """ Afhandelen van de GET request waarmee we een bestand terug geven. """

        try:
            match_pk = int(kwargs['match_pk'][:6])      # afkappen voor de veiligheid
            match = (CompetitieMatch
                     .objects
                     .select_related('vereniging',
                                     'locatie')
                     .get(pk=match_pk))
        except (ValueError, CompetitieMatch.DoesNotExist):
            raise Http404('Wedstrijd niet gevonden')

        try:
            klasse_pk = int(kwargs['klasse_pk'][:6])            # afkappen voor de veiligheid
            team_klasse = (CompetitieTeamKlasse
                           .objects
                           .get(pk=klasse_pk))
        except (ValueError, CompetitieTeamKlasse.DoesNotExist):
            raise Http404('Klasse niet gevonden')

        deelkamps_bk = match.kampioenschap_set.filter(deel=DEEL_BK)
        if len(deelkamps_bk) == 0:
            raise Http404('Geen kampioenschap')

        deelkamp_bk = deelkamps_bk[0]

        comp = deelkamp_bk.competitie
        # TODO: check fase

        if comp.is_indoor():
            aantal_pijlen = 30
        else:
            aantal_pijlen = 25

        lid2voorkeuren = dict()  # [lid_nr] = SporterVoorkeuren
        for voorkeuren in SporterVoorkeuren.objects.select_related('sporter').all():
            lid2voorkeuren[voorkeuren.sporter.lid_nr] = voorkeuren
        # for

        vastgesteld = timezone.localtime(timezone.now())

        klasse_str = team_klasse.beschrijving

        boog_typen = team_klasse.team_type.boog_typen.all()
        boog_pks = list(boog_typen.values_list('pk', flat=True))

        # bepaal de naam van het terug te geven bestand
        fname = "bk-programma_teams_"
        fname += klasse_str.lower().replace(' ', '-')
        fname += '.xlsx'

        if comp.is_indoor():
            excel_name = 'template-excel-bk-indoor-teams.xlsx'
        else:
            excel_name = 'template-excel-bk-25m1pijl-teams.xlsx'

        # make een kopie van het RK programma in een tijdelijk bestand
        fpath = os.path.join(settings.INSTALL_PATH, 'CompLaagBond', 'files', excel_name)
        tmp_file = NamedTemporaryFile()

        try:
            shutil.copyfile(fpath, tmp_file.name)
        except FileNotFoundError:
            raise Http404('Kan BK programma niet vinden')

        # open de kopie, zodat we die aan kunnen passen
        try:
            prg = openpyxl.load_workbook(tmp_file)
        except (OSError, zipfile.BadZipFile, KeyError):
            raise Http404('Kan BK programma niet openen')

        # maak wijzigingen in het BK programma
        ws = prg['Deelnemers en Scores']

        if "ERE" in klasse_str:
            max_teams = 12
        else:
            max_teams = 8

            # maximaal 4 teams naar de finale, dus verwijder het blad voor 8 team finale
            if comp.is_indoor():
                del prg['Finales 8 teams']

            # verwijder 4 regels in Uitslag (voor teams 9..12)
            prg['Uitslag'].delete_rows(17, 4)

        ws['B2'] = 'BK teams %s, Klasse: %s' % (comp.beschrijving, klasse_str)
        ws['H4'] = match.datum_wanneer.strftime('%Y-%m-%d')
        ws['B4'] = match.vereniging.naam         # organisatie
        if match.locatie:
            ws['F4'] = match.locatie.adres       # adres van de locatie
        else:
            ws['F4'] = 'Onbekend'

        try:
            limiet = KampioenschapTeamKlasseLimiet.objects.get(kampioenschap=deelkamp_bk, team_klasse=team_klasse)
            max_teams = limiet.limiet
        except KampioenschapTeamKlasseLimiet.DoesNotExist:
            pass

        teams = (KampioenschapTeam
                 .objects
                 .filter(kampioenschap=deelkamp_bk,
                         team_klasse=team_klasse.pk)
                 .exclude(deelname=DEELNAME_NEE)
                 .exclude(is_reserve=True)
                 .select_related('vereniging')
                 .prefetch_related('gekoppelde_leden')
                 .order_by('volgorde'))

        ver_nrs = list()

        volg_nr = 0
        for team in teams:
            row_nr = 9 + volg_nr * 6
            row = str(row_nr)

            ver = team.vereniging
            if ver.ver_nr not in ver_nrs:
                ver_nrs.append(ver.ver_nr)

            # vereniging
            ws['D' + row] = '[%s] %s' % (ver.ver_nr, ver.naam)

            # team naam
            ws['F' + row] = team.team_naam

            # team sterkte
            sterkte = float(team.aanvangsgemiddelde) * aantal_pijlen
            sterkte_str = "%.1f" % sterkte
            sterkte_str = sterkte_str.replace('.', ',')
            ws['G' + row] = sterkte_str

            # vul de 4 sporters in
            aantal = 0
            for deelnemer in (team
                              .gekoppelde_leden
                              .select_related('sporterboog__sporter')
                              .order_by('-gemiddelde')):        # hoogste gemiddelde bovenaan
                row_nr += 1
                row = str(row_nr)

                sporter = deelnemer.sporterboog.sporter

                para_mark = False
                try:
                    voorkeuren = lid2voorkeuren[sporter.lid_nr]
                except KeyError:        # pragma: no cover
                    pass
                else:
                    if voorkeuren.para_voorwerpen or voorkeuren.opmerking_para_sporter:
                        para_mark = True

                # bondsnummer
                ws['E' + row] = sporter.lid_nr

                # volledige naam
                naam_str = sporter.volledige_naam()
                if para_mark:
                    naam_str += ' **'
                ws['F' + row] = naam_str

                # RK gemiddelde
                ws['G' + row] = deelnemer.gemiddelde

                aantal += 1
            # for

            # bij minder dan 4 sporters de overgebleven regels leegmaken
            while aantal < 4:
                row_nr += 1
                row = str(row_nr)
                ws['E' + row] = '-'         # bondsnummer
                ws['F' + row] = 'n.v.t.'    # naam
                ws['G' + row] = ''          # gemiddelde
                ws['H' + row] = ''          # score 1
                ws['I' + row] = ''          # score 2
                aantal += 1
            # while

            volg_nr += 1
            if volg_nr == max_teams:
                break
        # for

        while volg_nr < 12:
            row_nr = 9 + volg_nr * 6
            row = str(row_nr)

            # vereniging leeg maken
            ws['D' + row] = 'n.v.t.'     # vereniging
            ws['F' + row] = 'n.v.t.'     # team naam
            ws['G' + row] = ''           # team sterkte

            # sporters leegmaken
            aantal = 0
            while aantal < 4:
                row_nr += 1
                row = str(row_nr)
                ws['E' + row] = '-'         # bondsnummer
                ws['F' + row] = '-'         # naam
                ws['G' + row] = ''          # gemiddelde
                ws['H' + row] = ''          # score 1
                ws['I' + row] = ''          # score 2
                aantal += 1
            # while

            volg_nr += 1
        # while

        ws['B82'] = 'Deze gegevens zijn opgehaald op %s' % vastgesteld.strftime('%Y-%m-%d %H:%M:%S')

        if "ERE" not in klasse_str:
            # verwijder teams 9..12 in Deelnemers en Scores
            ws.delete_rows(56, 24)

        # alle gerechtigde deelnemers opnemen op een apart tabblad, met gemiddelde en boogtype
        ws = prg['Toegestane deelnemers']

        cd_font = ws['C18'].font
        c_align = ws['C18'].alignment
        c_format = ws['C18'].number_format

        d_align = ws['D18'].alignment
        d_format = ws['D18'].number_format

        efgh_font = ws['E18'].font          # noqa
        e_align = ws['E18'].alignment

        f_align = ws['F18'].alignment

        g_align = ws['G18'].alignment
        g_format = ws['G18'].number_format

        row_nr = 16
        prev_ver = None
        # alle RK deelnemers mogen schieten in het team
        # (sporter hoeft niet persoonlijk geplaatst te zijn voor het BK)
        for deelnemer in (KampioenschapSporterBoog
                          .objects
                          .filter(kampioenschap__competitie=comp,
                                  kampioenschap__deel=DEEL_RK,
                                  bij_vereniging__ver_nr__in=ver_nrs,
                                  sporterboog__boogtype__pk__in=boog_pks)       # filter op toegestane boogtypen
                          .select_related('bij_vereniging',
                                          'sporterboog__sporter',
                                          'sporterboog__boogtype')
                          .order_by('bij_vereniging',
                                    '-gemiddelde')):                            # hoogste eerst

            row_nr += 1
            row = str(row_nr)

            # vereniging
            ver = deelnemer.bij_vereniging
            if ver != prev_ver:
                row_nr += 1     # extra lege regel
                row = str(row_nr)
                ws['C' + row] = ver.regio.regio_nr
                if row_nr != 18:
                    ws['C' + row].font = copy(cd_font)
                    ws['C' + row].alignment = copy(c_align)
                    ws['C' + row].number_format = copy(c_format)

                ws['D' + row] = '[%s] %s' % (ver.ver_nr, ver.naam)
                if row_nr != 18:
                    ws['D' + row].font = copy(cd_font)
                    ws['D' + row].alignment = copy(d_align)
                    ws['D' + row].number_format = copy(d_format)

                prev_ver = ver

            # sporter
            sporter = deelnemer.sporterboog.sporter

            para_notities = ''
            try:
                voorkeuren = lid2voorkeuren[sporter.lid_nr]
            except KeyError:        # pragma: no cover
                pass
            else:
                if voorkeuren.para_voorwerpen:
                    para_notities = 'Sporter laat voorwerpen op de schietlijn staan'

                if voorkeuren.opmerking_para_sporter:
                    if para_notities != '':
                        para_notities += '\n'
                    para_notities += voorkeuren.opmerking_para_sporter

            ws['E' + row] = sporter.lid_nr

            naam_str = sporter.volledige_naam()
            if para_notities:
                naam_str += ' **'
            ws['F' + row] = naam_str

            ws['G' + row] = deelnemer.gemiddelde
            ws['H' + row] = deelnemer.sporterboog.boogtype.beschrijving

            if para_notities:
                ws['I' + row] = para_notities

            if row_nr != 18:
                ws['E' + row].font = copy(efgh_font)
                ws['F' + row].font = copy(efgh_font)
                ws['G' + row].font = copy(efgh_font)
                ws['H' + row].font = copy(efgh_font)
                ws['I' + row].font = copy(efgh_font)

                ws['E' + row].alignment = copy(e_align)
                ws['G' + row].alignment = copy(g_align)
                ws['G' + row].number_format = copy(g_format)
        # for

        row_nr += 2
        row = str(row_nr)
        ws['B' + row] = 'Deze gegevens zijn opgehaald op %s' % vastgesteld.strftime('%Y-%m-%d %H:%M:%S')
        ws['B' + row].font = copy(efgh_font)
        ws['B' + row].alignment = copy(f_align)

        # geef het aangepaste BK programma aan de client
        response = HttpResponse(content_type=CONTENT_TYPE_XLSX)
        response['Content-Disposition'] = 'attachment; filename="%s"' % fname
        prg.save(response)      # noqa

        del prg
        tmp_file.close()

        return response


# end of file
