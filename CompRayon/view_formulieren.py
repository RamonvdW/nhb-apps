# -*- coding: utf-8 -*-

#  Copyright (c) 2020-2022 Ramon van der Winkel.
#  All rights reserved.
#  Licensed under BSD-3-Clause-Clear. See LICENSE file for details.

from django.conf import settings
from django.urls import reverse
from django.http import Http404, HttpResponse
from django.utils import timezone
from django.views.generic import TemplateView
from django.contrib.auth.mixins import UserPassesTestMixin
from Competitie.models import (CompetitieKlasse, LAAG_RK, DeelcompetitieKlasseLimiet,
                               KampioenschapSchutterBoog, KampioenschapTeam,
                               DEELNAME_NEE, DEELNAME2STR)
from Functie.rol import Rollen, rol_get_huidige_functie
from Plein.menu import menu_dynamics
from Wedstrijden.models import CompetitieWedstrijd
from tempfile import NamedTemporaryFile
from copy import copy
import openpyxl
import zipfile
import shutil
import os


TEMPLATE_DOWNLOAD_RK_FORMULIER = 'comprayon/hwl-download-rk-formulier.dtl'


class DownloadRkFormulierView(UserPassesTestMixin, TemplateView):

    """ Toon de HWL of WL de waarschijnlijke deelnemerslijst voor een wedstrijd bij deze vereniging
    """

    # class variables shared by all instances
    template_name = TEMPLATE_DOWNLOAD_RK_FORMULIER
    raise_exception = True  # genereer PermissionDefined als test_func False terug geeft

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.rol_nu, self.functie_nu = None, None

    def test_func(self):
        """ called by the UserPassesTestMixin to verify the user has permissions to use this view """
        self.rol_nu, self.functie_nu = rol_get_huidige_functie(self.request)
        return self.functie_nu and self.rol_nu in (Rollen.ROL_SEC, Rollen.ROL_HWL, Rollen.ROL_WL)

    def get_context_data(self, **kwargs):
        """ called by the template system to get the context data for the template """
        context = super().get_context_data(**kwargs)

        try:
            wedstrijd_pk = int(kwargs['wedstrijd_pk'][:6])      # afkappen voor de veiligheid
            wedstrijd = (CompetitieWedstrijd
                         .objects
                         .select_related('vereniging')
                         .prefetch_related('indiv_klassen',
                                           'team_klassen')
                         .get(pk=wedstrijd_pk,
                              vereniging=self.functie_nu.nhb_ver))
        except (ValueError, CompetitieWedstrijd.DoesNotExist):
            raise Http404('Wedstrijd niet gevonden')

        plannen = wedstrijd.competitiewedstrijdenplan_set.all()
        if len(plannen) == 0:
            raise Http404('Geen wedstrijden plan')
        plan = plannen[0]
        del plannen

        deelcomp = plan.deelcompetitie_set.select_related('competitie', 'nhb_rayon').all()[0]
        if deelcomp.laag != LAAG_RK:
            raise Http404('Verkeerde competitie')

        context['deelcomp'] = deelcomp
        context['wedstrijd'] = wedstrijd
        context['vereniging'] = wedstrijd.vereniging        # als we hier komen is dit altijd bekend
        context['locatie'] = wedstrijd.locatie
        context['aantal_banen'] = '?'

        comp = deelcomp.competitie
        # TODO: check fase
        if comp.afstand == '18':
            aantal_pijlen = 30
            if wedstrijd.locatie:
                context['aantal_banen'] = wedstrijd.locatie.banen_18m
        else:
            aantal_pijlen = 25
            if wedstrijd.locatie:
                context['aantal_banen'] = wedstrijd.locatie.banen_25m

        if deelcomp.laag == LAAG_RK:
            wedstrijd.is_rk = True
            wedstrijd.beschrijving = "Rayonkampioenschap"
        # else:
        #     wedstrijd.is_bk = True
        #     wedstrijd.beschrijving = "Bondskampioenschap"

        heeft_indiv = heeft_teams = False
        beschr = list()

        klasse_indiv_pks = list()
        klasse_team_pks = list()
        wedstrijd.klassen_lijst = klassen_lijst = list()
        for klasse in wedstrijd.indiv_klassen.all():
            klassen_lijst.append(str(klasse))
            klasse_indiv_pks.append(klasse.pk)
            if not heeft_indiv:
                heeft_indiv = True
                beschr.append('Individueel')
        # for
        for klasse in wedstrijd.team_klassen.all():
            klassen_lijst.append(str(klasse))
            klasse_team_pks.append(klasse.pk)
            if not heeft_teams:
                heeft_teams = True
                beschr.append('Teams')
        # for

        vastgesteld = timezone.localtime(timezone.now())
        context['vastgesteld'] = vastgesteld

        context['heeft_indiv'] = heeft_indiv
        context['heeft_teams'] = heeft_teams
        context['beschrijving'] = "%s %s" % (wedstrijd.beschrijving, " en ".join(beschr))

        # zoek de deelnemers erbij
        if heeft_indiv:
            deelnemers = (KampioenschapSchutterBoog
                          .objects
                          .filter(deelcompetitie=deelcomp,
                                  klasse__indiv__pk__in=klasse_indiv_pks)
                          .exclude(deelname=DEELNAME_NEE)
                          .select_related('sporterboog',
                                          'sporterboog__sporter',
                                          'bij_vereniging',
                                          'klasse')
                          .order_by('klasse',
                                    'rank'))
            context['deelnemers_indiv'] = deelnemers

            prev_klasse = None
            for deelnemer in deelnemers:
                if deelnemer.klasse != prev_klasse:
                    deelnemer.break_before = True
                    deelnemer.url_download_indiv = reverse('CompRayon:formulier-indiv-als-bestand',
                                                           kwargs={'wedstrijd_pk': wedstrijd.pk,
                                                                   'klasse_pk': deelnemer.klasse.pk})
                    prev_klasse = deelnemer.klasse

                deelnemer.ver_nr = deelnemer.bij_vereniging.ver_nr
                deelnemer.ver_naam = deelnemer.bij_vereniging.naam
                deelnemer.lid_nr = deelnemer.sporterboog.sporter.lid_nr
                deelnemer.volledige_naam = deelnemer.sporterboog.sporter.volledige_naam()
                deelnemer.gemiddelde_str = "%.3f" % deelnemer.gemiddelde
                deelnemer.gemiddelde_str = deelnemer.gemiddelde_str.replace('.', ',')
            # for

        if heeft_teams:
            teams = (KampioenschapTeam
                     .objects
                     .filter(deelcompetitie=deelcomp,
                             klasse__team__pk__in=klasse_team_pks)
                     .select_related('vereniging',
                                     'klasse')
                     .prefetch_related('gekoppelde_schutters')
                     .order_by('klasse',
                               '-aanvangsgemiddelde'))      # sterkste team bovenaan
            context['deelnemers_teams'] = teams

            if not comp.klassengrenzen_vastgesteld_rk_bk:
                context['geen_klassengrenzen'] = True

            volg_nr = 0
            prev_klasse = None
            for team in teams:
                if team.klasse != prev_klasse:
                    team.break_before = True
                    team.url_download_teams = reverse('CompRayon:formulier-teams-als-bestand',
                                                      kwargs={'wedstrijd_pk': wedstrijd.pk,
                                                              'klasse_pk': team.klasse.pk})

                    prev_klasse = team.klasse
                    volg_nr = 0

                volg_nr += 1
                team.volg_nr = volg_nr
                team.ver_nr = team.vereniging.ver_nr
                team.ver_naam = team.vereniging.naam
                team.sterkte_str = "%.1f" % (team.aanvangsgemiddelde * aantal_pijlen)
                team.sterkte_str = team.sterkte_str.replace('.', ',')

                for lid in team.gekoppelde_schutters.all():
                    sporter = lid.sporterboog.sporter
                    lid.naam_str = "[%s] %s" % (sporter.lid_nr, sporter.volledige_naam())
                    lid.gem_str = lid.gemiddelde
                # for
            # for

        context['kruimels'] = (
            (reverse('Vereniging:overzicht'), 'Beheer Vereniging'),
            (reverse('CompScores:wedstrijden'), 'Competitie wedstrijden'),
            (None, 'RK programma')
        )

        menu_dynamics(self.request, context)
        return context


class FormulierIndivAlsBestandView(UserPassesTestMixin, TemplateView):

    """ Geef de HWL het ingevulde wedstrijdformulier voor een RK wedstrijd bij deze vereniging """

    # class variables shared by all instances
    raise_exception = True  # genereer PermissionDefined als test_func False terug geeft

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.rol_nu, self.functie_nu = None, None

    def test_func(self):
        """ called by the UserPassesTestMixin to verify the user has permissions to use this view """
        self.rol_nu, self.functie_nu = rol_get_huidige_functie(self.request)
        return self.functie_nu and self.rol_nu == Rollen.ROL_HWL

    def get(self, request, *args, **kwargs):
        """ Afhandelen van de GET request waarmee we een bestand terug geven. """

        try:
            wedstrijd_pk = int(kwargs['wedstrijd_pk'][:6])      # afkappen voor de veiligheid
            wedstrijd = (CompetitieWedstrijd
                         .objects
                         .select_related('vereniging')
                         .get(pk=wedstrijd_pk,
                              vereniging=self.functie_nu.nhb_ver))
        except (ValueError, CompetitieWedstrijd.DoesNotExist):
            raise Http404('Wedstrijd niet gevonden')

        try:
            klasse_pk = int(kwargs['klasse_pk'][:6])            # afkappen voor de veiligheid
            klasse = (CompetitieKlasse
                      .objects
                      .exclude(indiv=None)
                      .get(pk=klasse_pk))
        except (ValueError, CompetitieKlasse.DoesNotExist):
            raise Http404('Klasse niet gevonden')

        plannen = wedstrijd.competitiewedstrijdenplan_set.all()
        if len(plannen) == 0:
            raise Http404('Geen wedstrijden plan')
        plan = plannen[0]
        del plannen

        deelcomp = plan.deelcompetitie_set.select_related('competitie', 'nhb_rayon').all()[0]
        if deelcomp.laag != LAAG_RK:
            raise Http404('Verkeerde competitie')

        comp = deelcomp.competitie
        # TODO: check fase

        vastgesteld = timezone.localtime(timezone.now())

        klasse_str = klasse.indiv.beschrijving

        aantal_banen = 16
        if wedstrijd.locatie:
            if comp.afstand == '18':
                aantal_banen = wedstrijd.locatie.banen_18m
            else:
                aantal_banen = wedstrijd.locatie.banen_25m

        # TODO: haal de ingestelde limiet op (maximum aantal deelnemers)
        try:
            lim = DeelcompetitieKlasseLimiet(deelcompetitie=deelcomp,
                                             klasse=klasse)
        except DeelcompetitieKlasseLimiet.DoesNotExist:
            limiet = 24
        else:
            limiet = lim.limiet

        if comp.afstand == '18':
            excel_name = 'template-excel-rk-indoor-indiv.xlsm'
            ws_name = 'Voorronde'
        else:
            excel_name = 'template-excel-rk-25m1pijl-indiv.xlsm'
            ws_name = 'Wedstrijd'

        # bepaal de naam van het terug te geven bestand
        fname = "rk-programma_individueel-rayon%s_" % deelcomp.nhb_rayon.rayon_nr
        fname += klasse_str.lower().replace(' ', '-')
        fname += '.xlsm'

        # make een kopie van het RK programma in een tijdelijk bestand
        fpath = os.path.join(settings.INSTALL_PATH, 'CompRayon', 'files', excel_name)
        tmp_file = NamedTemporaryFile()
        try:
            shutil.copyfile(fpath, tmp_file.name)
        except FileNotFoundError:
            raise Http404('Kan RK programma niet vinden')

        # open de kopie, zodat we die aan kunnen passen
        try:
            prg = openpyxl.load_workbook(tmp_file, keep_vba=True)
        except (OSError, zipfile.BadZipFile, KeyError):
            raise Http404('Kan RK programma niet openen')

        # maak wijzigingen in het RK programma
        ws = prg[ws_name]

        ws['C4'] = 'Rayonkampioenschappen %s, Rayon %s, %s' % (comp.beschrijving, deelcomp.nhb_rayon.rayon_nr, klasse.indiv.beschrijving)

        ws['D5'] = wedstrijd.vereniging.naam     # organisatie
        ws['F5'] = 'Datum: ' + wedstrijd.datum_wanneer.strftime('%Y-%m-%d')
        if wedstrijd.locatie:
            ws['H5'] = wedstrijd.locatie.adres       # adres van de wedstrijdlocatie
        else:
            ws['H5'] = 'Onbekend'

        ws['A32'] = 'Deze gegevens zijn opgehaald op %s' % vastgesteld.strftime('%Y-%m-%d %H:%M:%S')

        i_font = ws['I7'].font
        i_align = ws['I7'].alignment
        i_format = ws['I7'].number_format

        d_align = ws['D7'].alignment
        g_align = ws['G7'].alignment

        deelnemers = (KampioenschapSchutterBoog
                      .objects
                      .filter(deelcompetitie=deelcomp,
                              klasse=klasse.pk)
                      .select_related('sporterboog',
                                      'sporterboog__sporter',
                                      'bij_vereniging',
                                      'bij_vereniging__regio')
                      .order_by('rank'))

        baan_nr = 1
        baan_letter = 'A'
        deelnemer_nr = 0

        row1_nr = 6
        row2_nr = 35
        for deelnemer in deelnemers:
            is_deelnemer = False
            reserve_str = ''
            if deelnemer.deelname != DEELNAME_NEE:
                deelnemer_nr += 1
                if deelnemer_nr > limiet:
                    reserve_str = ' (reserve)'
                else:
                    is_deelnemer = True

            if is_deelnemer:
                row1_nr += 1
                row = str(row1_nr)
            else:
                row2_nr += 1
                row = str(row2_nr)
                ws['D' + row].alignment = copy(d_align)
                ws['G' + row].alignment = copy(g_align)
                ws['I' + row].alignment = copy(i_align)
                ws['I' + row].font = copy(i_font)
                ws['I' + row].number_format = copy(i_format)

            if is_deelnemer:
                ws['A' + row] = baan_nr
                ws['B' + row] = baan_letter

                baan_nr += 1
                if baan_nr > aantal_banen:
                    baan_nr = 1
                    baan_letter = chr(ord(baan_letter) + 1)

            # bondsnummer
            ws['D' + row] = deelnemer.sporterboog.sporter.lid_nr

            # volledige naam
            ws['E' + row] = deelnemer.sporterboog.sporter.volledige_naam()

            # vereniging
            ver = deelnemer.bij_vereniging
            ws['F' + row] = '%s %s' % (ver.ver_nr, ver.naam)

            # regio
            ws['G' + row] = ver.regio.regio_nr

            ws['H' + row] = DEELNAME2STR[deelnemer.deelname] + reserve_str

            # gemiddelde
            ws['I' + row] = deelnemer.gemiddelde
        # for

        # geef het aangepaste RK programma aan de client
        response = HttpResponse(content_type='application/vnd.ms-excel.sheet.macroEnabled.12')
        response['Content-Disposition'] = 'attachment; filename="%s"' % fname
        prg.save(response)

        return response


class FormulierTeamsAlsBestandView(UserPassesTestMixin, TemplateView):

    """ Geef de HWL het ingevulde wedstrijdformulier voor een RK wedstrijd bij deze vereniging """

    # class variables shared by all instances
    raise_exception = True  # genereer PermissionDefined als test_func False terug geeft

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.rol_nu, self.functie_nu = None, None

    def test_func(self):
        """ called by the UserPassesTestMixin to verify the user has permissions to use this view """
        self.rol_nu, self.functie_nu = rol_get_huidige_functie(self.request)
        return self.functie_nu and self.rol_nu == Rollen.ROL_HWL

    def get(self, request, *args, **kwargs):
        """ Afhandelen van de GET request waarmee we een bestand terug geven. """

        try:
            wedstrijd_pk = int(kwargs['wedstrijd_pk'][:6])      # afkappen voor de veiligheid
            wedstrijd = (CompetitieWedstrijd
                         .objects
                         .select_related('vereniging',
                                         'locatie')
                         .get(pk=wedstrijd_pk,
                              vereniging=self.functie_nu.nhb_ver))
        except (ValueError, CompetitieWedstrijd.DoesNotExist):
            raise Http404('Wedstrijd niet gevonden')

        try:
            klasse_pk = int(kwargs['klasse_pk'][:6])            # afkappen voor de veiligheid
            klasse = (CompetitieKlasse
                      .objects
                      .exclude(team=None)
                      .get(pk=klasse_pk))
        except (ValueError, CompetitieKlasse.DoesNotExist):
            raise Http404('Klasse niet gevonden')

        plannen = wedstrijd.competitiewedstrijdenplan_set.all()
        if len(plannen) == 0:
            raise Http404('Geen wedstrijden plan')
        plan = plannen[0]
        del plannen

        deelcomp = plan.deelcompetitie_set.select_related('competitie', 'nhb_rayon').all()[0]
        if deelcomp.laag != LAAG_RK:
            raise Http404('Verkeerde competitie')

        comp = deelcomp.competitie
        # TODO: check fase

        if comp.afstand == '18':
            aantal_pijlen = 30
        else:
            aantal_pijlen = 25

        vastgesteld = timezone.localtime(timezone.now())

        klasse_str = klasse.team.beschrijving

        boog_typen = klasse.team.team_type.boog_typen.all()
        boog_pks = list(boog_typen.values_list('pk', flat=True))

        # bepaal de naam van het terug te geven bestand
        fname = "rk-programma_teams-rayon%s_" % deelcomp.nhb_rayon.rayon_nr
        fname += klasse_str.lower().replace(' ', '-')
        fname += '.xlsm'

        if comp.afstand == '18':
            excel_name = 'template-excel-rk-indoor-teams.xlsm'
        else:
            excel_name = 'template-excel-rk-25m1pijl-teams.xlsm'

        # make een kopie van het RK programma in een tijdelijk bestand
        fpath = os.path.join(settings.INSTALL_PATH, 'CompRayon', 'files', excel_name)
        tmp_file = NamedTemporaryFile()

        try:
            shutil.copyfile(fpath, tmp_file.name)
        except FileNotFoundError:
            raise Http404('Kan RK programma niet vinden')

        # open de kopie, zodat we die aan kunnen passen
        try:
            prg = openpyxl.load_workbook(tmp_file, keep_vba=True)
        except (OSError, zipfile.BadZipFile, KeyError):
            raise Http404('Kan RK programma niet openen')

        # maak wijzigingen in het RK programma
        ws = prg['Uitleg']
        ws['A5'] = 'Deze gegevens zijn opgehaald op %s' % vastgesteld.strftime('%Y-%m-%d %H:%M:%S')

        ws = prg['Deelnemers en Scores']

        ws['B1'] = 'Rayonkampioenschappen Teams Rayon %s, %s' % (deelcomp.nhb_rayon.rayon_nr, comp.beschrijving)
        ws['B4'] = 'Klasse: ' + klasse_str
        ws['D3'] = wedstrijd.datum_wanneer.strftime('%Y-%m-%d')
        ws['E3'] = wedstrijd.vereniging.naam     # organisatie
        if wedstrijd.locatie:
            ws['H3'] = wedstrijd.locatie.adres       # adres van de wedstrijdlocatie
        else:
            ws['H3'] = 'Onbekend'

        teams = (KampioenschapTeam
                 .objects
                 .filter(deelcompetitie=deelcomp,
                         klasse=klasse.pk)
                 .select_related('vereniging',
                                 'vereniging__regio')
                 .prefetch_related('gekoppelde_schutters')
                 .order_by('-aanvangsgemiddelde'))      # sterkste team bovenaan

        ver_nrs = list()

        volg_nr = 0
        for team in teams:
            row_nr = 9 + volg_nr * 6
            row = str(row_nr)

            ver = team.vereniging
            if ver.ver_nr not in ver_nrs:
                ver_nrs.append(ver.ver_nr)

            # regio
            ws['C' + row] = ver.regio.regio_nr

            # vereniging
            ws['D' + row] = '[%s] %s' % (ver.ver_nr, ver.naam)

            # team naam
            ws['F' + row] = team.team_naam

            # team sterkte
            sterkte_str = "%.1f" % (team.aanvangsgemiddelde * aantal_pijlen)
            sterkte_str = sterkte_str.replace('.', ',')
            ws['G' + row] = sterkte_str

            # vul de 4 sporters in
            aantal = 0
            for deelnemer in team.gekoppelde_schutters.all():
                row_nr += 1
                row = str(row_nr)

                sporter = deelnemer.sporterboog.sporter

                # bondsnummer
                ws['E' + row] = sporter.lid_nr

                # volledige naam
                ws['F' + row] = sporter.volledige_naam()

                # regio gemiddelde
                ws['G' + row] = deelnemer.gemiddelde

                aantal += 1
            # for

            # bij minder dan 4 sporters de overgebleven regels leegmaken
            while aantal < 4:
                row_nr += 1
                row = str(row_nr)
                ws['E' + row] = '-'         # bondsnummer
                ws['F' + row] = 'n.v.t.'    # naam
                ws['G' + row] = ''          # gemiddelde
                ws['H' + row] = ''          # score 1
                ws['I' + row] = ''          # score 2
                aantal += 1
            # while

            volg_nr += 1
        # for

        while volg_nr < 12:
            row_nr = 9 + volg_nr * 6
            row = str(row_nr)

            # vereniging leeg maken
            ws['C' + row] = '-'          # regio
            ws['D' + row] = 'n.v.t.'     # vereniging
            ws['F' + row] = 'n.v.t.'     # team naam
            ws['G' + row] = ''           # team sterkte

            # sporters leegmaken
            aantal = 0
            while aantal < 4:
                row_nr += 1
                row = str(row_nr)
                ws['E' + row] = '-'         # bondsnummer
                ws['F' + row] = '-'         # naam
                ws['G' + row] = ''          # gemiddelde
                ws['H' + row] = ''          # score 1
                ws['I' + row] = ''          # score 2
                aantal += 1
            # while

            volg_nr += 1
        # while

        ws['B82'] = 'Deze gegevens zijn opgehaald op %s' % vastgesteld.strftime('%Y-%m-%d %H:%M:%S')

        # alle gerechtigde deelnemers opnemen op een apart tabblad, met gemiddelde en boogtype
        ws = prg['Toegestane deelnemers']

        cd_font = ws['C18'].font
        c_align = ws['C18'].alignment
        c_format = ws['C18'].number_format

        d_align = ws['D18'].alignment
        d_format = ws['D18'].number_format

        efgh_font = ws['E18'].font
        e_align = ws['E18'].alignment

        f_align = ws['F18'].alignment

        g_align = ws['G18'].alignment
        g_format = ws['G18'].number_format

        row_nr = 16
        prev_ver = None
        for deelnemer in (KampioenschapSchutterBoog
                          .objects
                          .filter(deelcompetitie=deelcomp,
                                  bij_vereniging__ver_nr__in=ver_nrs,
                                  sporterboog__boogtype__pk__in=boog_pks)       # filter op toegestane boogtypen
                          .select_related('bij_vereniging__regio',
                                          'sporterboog__sporter',
                                          'sporterboog__boogtype')
                          .order_by('bij_vereniging__regio',
                                    'bij_vereniging',
                                    '-gemiddelde')):                            # hoogste eerst

            row_nr += 1
            row = str(row_nr)

            # vereniging
            ver = deelnemer.bij_vereniging
            if ver != prev_ver:
                row_nr += 1     # extra lege regel
                row = str(row_nr)
                ws['C' + row] = ver.regio.regio_nr
                if row_nr != 18:
                    ws['C' + row].font = copy(cd_font)
                    ws['C' + row].alignment = copy(c_align)
                    ws['C' + row].number_format = copy(c_format)

                ws['D' + row] = '[%s] %s' % (ver.ver_nr, ver.naam)
                if row_nr != 18:
                    ws['D' + row].font = copy(cd_font)
                    ws['D' + row].alignment = copy(d_align)
                    ws['D' + row].number_format = copy(d_format)

                prev_ver = ver

            # sporter
            sporter = deelnemer.sporterboog.sporter
            ws['E' + row] = sporter.lid_nr
            ws['F' + row] = sporter.volledige_naam()

            ws['G' + row] = deelnemer.gemiddelde
            ws['H' + row] = deelnemer.sporterboog.boogtype.beschrijving

            if row_nr != 18:
                ws['E' + row].font = copy(efgh_font)
                ws['F' + row].font = copy(efgh_font)
                ws['G' + row].font = copy(efgh_font)
                ws['H' + row].font = copy(efgh_font)

                ws['E' + row].alignment = copy(e_align)
                ws['G' + row].alignment = copy(g_align)
                ws['G' + row].number_format = copy(g_format)
        # for

        row_nr += 2
        row = str(row_nr)
        ws['B' + row] = 'Deze gegevens zijn opgehaald op %s' % vastgesteld.strftime('%Y-%m-%d %H:%M:%S')
        ws['B' + row].font = copy(efgh_font)
        ws['B' + row].alignment = copy(f_align)

        # geef het aangepaste RK programma aan de client
        response = HttpResponse(content_type='application/vnd.ms-excel.sheet.macroEnabled.12')
        response['Content-Disposition'] = 'attachment; filename="%s"' % fname
        prg.save(response)

        return response


# end of file
