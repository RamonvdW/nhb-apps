# -*- coding: utf-8 -*-

#  Copyright (c) 2019-2026 Ramon van der Winkel.
#  All rights reserved.
#  Licensed under BSD-3-Clause-Clear. See LICENSE file for details.

from django.test import TestCase
from django.utils.dateparse import parse_date
from Records.definities import LEEFTIJDSCATEGORIE, GESLACHT, MATERIAALKLASSE, DISCIPLINE
from Records.models import IndivRecord
from Sporter.models import Sporter
from TestHelpers.e2ehelpers import E2EHelpers
import datetime
import openpyxl


class TestRecordsCliImport(E2EHelpers, TestCase):
    """ unittests voor de Records applicatie, Import module """

    hdrs_outdoor = ['Index', 'Geslacht', 'Leeftijd', 'Materiaalklasse', 'Discipline', 'Soort_record',
                    'Para klasse', 'Verbeterbaar', 'Pijlen', 'Bondsnummer', 'Naam', 'Datum', 'Plaats', 'Land',
                    'Score', 'X-count', 'Ook ER', 'Ook WR', 'Notities']

    hdrs_indoor = ['Index', 'Geslacht', 'Leeftijd', 'Materiaalklasse', 'Discipline', 'Soort_record',
                   'Para klasse', 'Verbeterbaar', 'Pijlen', 'Bondsnummer', 'Naam', 'Datum', 'Plaats', 'Land',
                   'Score', 'X-count', 'Ook ER', 'Ook WR', 'Notities']

    hdrs_25m1p = ['Index', 'Geslacht', 'Leeftijd', 'Materiaalklasse', 'Discipline', 'Soort_record',
                  'Para klasse', 'Verbeterbaar', 'Pijlen', 'Bondsnummer', 'Naam', 'Datum', 'Plaats', 'Land',
                  'Score', 'Notities']

    hdrs_team = ['Geslacht', 'Leeftijd', 'Materiaalklasse', 'Discipline', 'Soort_record', 'Verbeterbaar',
                 'Pijlen', 'Bondsnummer', 'Naam', 'Datum', 'Plaats', 'Land', 'Score', 'Notities']

    def setUp(self):
        """ initialisatie van de test case """

        # NhbLib
        sporter1 = Sporter(
                    lid_nr=123456,
                    voornaam='Jan',
                    achternaam='Schutter',
                    email='jan@test.nl',
                    geboorte_datum=parse_date('1970-03-03'),
                    adres_code='Papendal',
                    geslacht='M',
                    sinds_datum=parse_date("1991-02-03"))  # Y-M-D
        sporter1.save()
        self.sporter1 = sporter1

        sporter2 = Sporter(
                    lid_nr=123457,
                    voornaam='Petra',
                    achternaam='Schutter',
                    email='petra@test.nl',
                    geboorte_datum=parse_date('1970-01-30'),
                    adres_code='Arnhem',
                    geslacht='V',
                    sinds_datum=parse_date("1991-02-05"))  # Y-M-D
        sporter2.save()
        self.sporter2 = sporter2

        # Record 42
        rec = IndivRecord(
                    volg_nr=42,
                    discipline=DISCIPLINE[0][0],   # OD
                    soort_record='WA1440',
                    geslacht=GESLACHT[0][0],   # M
                    leeftijdscategorie=LEEFTIJDSCATEGORIE[0][0],   # M
                    materiaalklasse=MATERIAALKLASSE[0][0],     # R
                    # materiaalklasse_overig
                    sporter=sporter2,
                    naam=sporter2.volledige_naam(),
                    datum=parse_date('2017-08-27'),
                    plaats='Papendal',
                    land='Nederland',
                    # score_notitie
                    # is_european_record
                    # is_world_record
                    score=1234,
                    max_score=1440,
                    x_count=56)
        rec.save()
        self.rec_42_outdoor = rec

        self.assertEqual(rec.score_str(), '1234 (56X)')
        self.assertEqual(rec.max_score_str(), '1440 (144X)')

        # Record 43
        rec = IndivRecord(
                    volg_nr=43,
                    discipline=DISCIPLINE[1][0],   # 18
                    soort_record='Test record para',
                    geslacht=GESLACHT[1][0],       # V
                    leeftijdscategorie=LEEFTIJDSCATEGORIE[1][0],   # S
                    materiaalklasse='R',           # Recurve
                    para_klasse='Open',
                    # rec.sporter,
                    naam='Top Schutter 2',
                    datum=datetime.datetime.now(),
                    plaats='Ergens Anders',
                    land='Nederland',
                    # score_notitie
                    # is_european_record
                    # is_world_record
                    score=1235)
        rec.save()
        self.rec_43_18m = rec

        self.assertIsNotNone(str(rec))
        self.assertEqual(rec.score_str(), '1235')

        # Record 44
        rec = IndivRecord(
                volg_nr=44,
                discipline=DISCIPLINE[2][0],   # 25
                soort_record='25m',
                geslacht=GESLACHT[1][0],   # V
                leeftijdscategorie=LEEFTIJDSCATEGORIE[3][0],   # C
                materiaalklasse='R',       # Recurve
                sporter=self.sporter2,
                naam=self.sporter2.volledige_naam(),
                datum=parse_date('2017-08-27'),
                plaats='Nergens',
                land='Niederland',      # noqa
                # score_notitie
                # is_european_record
                # is_world_record
                score=249)
        rec.save()
        self.rec_44_25m = rec

        self.assertIsNotNone(str(rec))

        self.tmp_fname = '/tmp/test_records.xlsx'

    def test_basics(self):
        # bestand bestaat niet
        f1, f2 = self.run_management_command('import_records', './not-existing.xlsx')
        self.assertEqual(f1.getvalue(), '')
        self.assertTrue("[ERROR] Kan bestand './not-existing.xlsx' niet lezen (" in f2.getvalue())

        # leeg bestand
        wb = openpyxl.Workbook()
        wb.save(self.tmp_fname)
        f1, f2 = self.run_management_command('import_records', self.tmp_fname, '--verbose', '--dryrun')
        # print("f2: %s" % f2.getvalue())
        self.assertEqual(f1.getvalue(), '')
        self.assertTrue("[ERROR] Kan sheet 'Data individueel outdoor' niet vinden" in f2.getvalue())
        self.assertTrue("DRY RUN" in f2.getvalue())
        self.assertTrue("Samenvatting: 0 records;" in f2.getvalue())

        # verkeerde headers
        ws = wb.create_sheet('Data individueel outdoor')
        ws.append(self.hdrs_team)     # verkeerde headers
        wb.save(self.tmp_fname)
        f1, f2 = self.run_management_command('import_records', self.tmp_fname, '--verbose', '--dryrun')
        # print("f2: %s" % f2.getvalue())
        self.assertEqual(f1.getvalue(), '')
        self.assertTrue("[ERROR] Kolom headers kloppen niet:" in f2.getvalue())

    def test_import_outdoor(self):
        wb = openpyxl.Workbook()
        ws = wb.create_sheet('Data individueel outdoor')
        ws.append(self.hdrs_outdoor)
        ws.append([])       # lege regel, wordt overgeslagen
        ws.append([
            '42',
            'H', 'M', 'R', 'Outdoor', 'WA1440', '', 'Ja', '144',
            str(self.sporter2.lid_nr), self.sporter2.volledige_naam(),
            '2017-08-27',
            self.rec_42_outdoor.plaats, self.rec_42_outdoor.land,
            '1234', 56,
        ])
        wb.save(self.tmp_fname)

        f1, f2 = self.run_management_command('import_records', self.tmp_fname, '--verbose', '--dryrun')
        # print("f1: %s" % f1.getvalue())
        # print("f2: %s" % f2.getvalue())
        self.assertEqual(f1.getvalue(), '')
        self.assertTrue('Samenvatting: 1 records; 1 ongewijzigd;' in f2.getvalue())

        # nog een keer, nu zijn er wijzigingen
        self.rec_42_outdoor.delete()
        rec = IndivRecord(
                    volg_nr=42,
                    discipline=DISCIPLINE[0][0],   # OD
                    soort_record='#WA1440',
                    geslacht=GESLACHT[1][0],       # V
                    leeftijdscategorie=LEEFTIJDSCATEGORIE[1][0],   # S
                    materiaalklasse=MATERIAALKLASSE[1][0],     # C
                    para_klasse='#',
                    # materiaalklasse_overig
                    sporter=self.sporter1,
                    naam=self.sporter1.volledige_naam(),
                    datum=parse_date('2017-08-26'),
                    plaats='#Papendal',
                    land='#Nederland',
                    verbeterbaar=False,
                    score_notitie='##',
                    is_european_record=True,
                    is_world_record=True,
                    score=1235,
                    max_score=1441,
                    x_count=55)
        rec.save()
        self.rec_42_outdoor = rec

        # niet opslaan
        self.run_management_command('import_records', self.tmp_fname, '--dryrun')

        # wel opslaan
        f1, f2 = self.run_management_command('import_records', self.tmp_fname)
        self.assertEqual(f1.getvalue(), '')
        # print("f2: %s" % f2.getvalue())
        self.assertTrue('[INFO] Wijzigingen voor record OD-42:' in f2.getvalue())
        self.assertEqual(17, f2.getvalue().count(' --> '))
        self.assertTrue('; 17 wijzigingen; ' in f2.getvalue())

    def test_nieuw(self):
        wb = openpyxl.Workbook()
        ws = wb.create_sheet('Data individueel outdoor')
        ws.append(self.hdrs_outdoor)

        # bestaand record
        ws.append([
            '42',
            'H', 'M', 'R', 'Outdoor', 'WA1440', '', 'Ja', '144',
            str(self.sporter2.lid_nr), self.sporter2.volledige_naam(),
            '2017-08-27',
            self.rec_42_outdoor.plaats, self.rec_42_outdoor.land,
            '1234', 56,
        ])

        # nieuw record
        ws.append([
            '999',
            'D', '21+', 'C', 'Outdoor', '70m (72p)', '', 'Nee', '72',
            str(self.sporter2.lid_nr), self.sporter2.volledige_naam(),
            '2026-03-27',
            'Ergens', 'Niemandsland',
            '650', 42,
            'ER', 'WR',
        ])

        wb.save(self.tmp_fname)

        f1, f2 = self.run_management_command('import_records', self.tmp_fname)
        self.assertEqual(f1.getvalue(), '')
        # print("f2: %s" % f2.getvalue())
        self.assertTrue('Record OD-999 toegevoegd' in f2.getvalue())

        rec = IndivRecord.objects.filter(volg_nr=999).first()
        self.assertEqual(rec.discipline, 'OD')
        self.assertEqual(rec.soort_record, '70m (72p)')
        self.assertEqual(rec.geslacht, 'V')
        self.assertEqual(rec.leeftijdscategorie, 's')
        self.assertEqual(rec.materiaalklasse, 'C')
        self.assertEqual(rec.para_klasse, '')
        self.assertEqual(rec.sporter.lid_nr, self.sporter2.lid_nr)
        self.assertEqual(rec.naam, self.sporter2.volledige_naam())
        self.assertEqual(rec.datum.year, 2026)
        self.assertEqual(rec.datum.month, 3)
        self.assertEqual(rec.datum.day, 27)
        self.assertEqual(rec.plaats, 'Ergens')
        self.assertEqual(rec.land, 'Niemandsland')
        self.assertEqual(rec.verbeterbaar, False)
        self.assertEqual(rec.score_notitie, '')
        self.assertEqual(rec.is_european_record, True)
        self.assertEqual(rec.is_world_record, True)
        self.assertEqual(rec.score, 650)
        self.assertEqual(rec.max_score, 720),
        self.assertEqual(rec.x_count, 42)

    def test_bad(self):
        wb = openpyxl.Workbook()
        ws = wb.create_sheet('Data individueel outdoor')
        ws.append(self.hdrs_outdoor)

        # bestaand record
        ws.append([
            '42',
            'H', 'M', 'R', 'Outdoor', 'WA1440', '', 'Ja', '144',
            str(self.sporter2.lid_nr), self.sporter2.volledige_naam(),
            '2017-08-27',
            self.rec_42_outdoor.plaats, self.rec_42_outdoor.land,
            '1234', 56,
        ])

        # slechte index (rest wordt niet overwogen)
        ws.append([
            'x',        # index is geen number
        ])

        # allemaal fouten
        ws.append([
            '998',
            '?',        # fout geslacht
            '?',        # foute leeftijdscategorie
            '?',        # foute materiaalklasse
            '?',        # foute discipline
            '?',        # foute record soort
            '?',        # foute para klasse
            'Ja',
            '?',        # fout aantal pijlen (geen getal)
            '999999',   # niet bestaand lid_nr
            '',         # naam wordt niet bekeken
            '1-2-3',    # foute datum
            self.rec_42_outdoor.plaats, self.rec_42_outdoor.land,
            '?',        # foute score (geen getal)
            '?',        # fout max score (geen getal)
            '?',        # foute ER indicatie
            '?',        # foute WR indicatie
        ])

        # nog meer fouten
        ws.append([
            '998',
            '?',
            '?',
            '?',
            '?',
            '70m',
            '?',
            'Ja',
            '36',
            '?',            # fout bondsnummer (niet 6 cijfers)
            '',   # naam
            'Onbekend',     # special case
            self.rec_42_outdoor.plaats, self.rec_42_outdoor.land,
            '999',          # foute score (hoger dan max score)
            '9999',         # fout max score (te hoog)
        ])

        wb.save(self.tmp_fname)

        f1, f2 = self.run_management_command('import_records', self.tmp_fname)
        self.assertEqual(f1.getvalue(), '')
        # print("f2: %s" % f2.getvalue())
        self.assertTrue("Samenvatting: 4 records; 1 ongewijzigd; 21 fouten; 0 verwijderd; 0 wijzigingen; 0 toegevoegd;" in f2.getvalue())

    def test_indoor(self):
        wb = openpyxl.Workbook()
        ws = wb.create_sheet('Data individueel indoor')
        ws.append(self.hdrs_indoor)

        # 50+
        ws.append([
            '1', 'D', '50+',
            'BB', 'Indoor', '30m',
            '', 'Nee', '36', '123456',
            'F. Schietnie',             # verkeerde naam
            'Onbekend', 'Plaats', 'Land',
            '360',
            '',     # X-count
        ])
        ws.append([
            '2', 'D', '50+',
            'BB', 'Indoor', '30m',
            '', 'Nee', '36', '123456',
            'Petra Schutter',
            'Onbekend', 'Plaats', 'Land',
            '360',
            '',     # X-count

        ])
        wb.save(self.tmp_fname)

        # dry-run: nieuwe record wordt niet opgeslagen
        f1, f2 = self.run_management_command('import_records', self.tmp_fname, '--dryrun')
        self.assertEqual(f1.getvalue(), '')
        # print("f2: %s" % f2.getvalue())
        self.assertTrue('[INFO] Record 18-2 toegevoegd' in f2.getvalue())
        self.assertTrue("[ERROR] Foute naam: 'F. Schietnie' maar sporter naam is 'Jan Schutter'" in f2.getvalue())

    def test_25m1pijl(self):
        wb = openpyxl.Workbook()
        ws = wb.create_sheet('Data individueel 25m1pijl')
        ws.append(self.hdrs_25m1p)

        ws.append([
            '1', 'H', 'Onder 18', 'BB', '25m1p', '50m',
            '', 'Nee', '50', '123456', 'Petra Schutter',
            'Onbekend', 'Plaats', 'Land',
            '500', 'Maximale score!'
        ])

        # Onder 21 + te oud
        ws.append([
            '2', 'H', 'Onder 21', 'BB', '25m1p', '50m',
            '', 'Nee', '50', '123456', 'Petra Schutter',
            '1900-01-01 00:00', 'Plaats', 'Land',
            '500', 'Maximale score!'
        ])

        wb.save(self.tmp_fname)

        f1, f2 = self.run_management_command('import_records', self.tmp_fname)
        self.assertEqual(f1.getvalue(), '')
        # print("f2: %s" % f2.getvalue())
        self.assertTrue('[INFO] Record 25-1 toegevoegd' in f2.getvalue())
        self.assertTrue("[ERROR] Fout in datum: '1900-01-01'" in f2.getvalue())

    def test_teams(self):
        wb = openpyxl.Workbook()
        ws = wb.create_sheet('Data team')
        ws.append(self.hdrs_team)

        wb.save(self.tmp_fname)

        f1, f2 = self.run_management_command('import_records', self.tmp_fname)
        self.assertEqual(f1.getvalue(), '')
        # print("f2: %s" % f2.getvalue())
        self.assertTrue("[INFO] Importeer van blad 'Data team'" in f2.getvalue())
        self.assertTrue('[TODO] Team records worden nog niet ondersteund' in f2.getvalue())


# end of file
