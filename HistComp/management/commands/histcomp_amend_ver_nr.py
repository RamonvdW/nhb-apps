# -*- coding: utf-8 -*-

#  Copyright (c) 2023 Ramon van der Winkel.
#  All rights reserved.
#  Licensed under BSD-3-Clause-Clear. See LICENSE file for details.

from django.core.management.base import BaseCommand
from HistComp.models import HistCompSeizoen, HistCompRegioIndiv
from NhbStructuur.models import NhbVereniging, NhbRegio
from openpyxl.utils.exceptions import InvalidFileException
import openpyxl
import zipfile


def verwijderde_verenigingen():
    for tup in [(1026, 109, 'Elshout'),
                (1028, 107, 'Riel'),
                (1058, 109, 'Vught'),
                (1066, 107, 'Prinsenbeek'),
                (1093, 111, 'Best'),
                (1147, 114, 'Venray'),
                (1152, 114, 'America'),
                (1170, 115, 'Weert'),
                (1191, 116, 'Maastricht'),
                (1226, 113, 'Asten'),
                (1310, 102, 'Culemborg'),
                (1355, 103, 'Drachten')]:

        ver_nr, regio_nr, plaats = tup

        ver = NhbVereniging(
                    ver_nr=tup[0],
                    regio=NhbRegio(regio_nr=regio_nr),
                    plaats=plaats)
        yield ver
    # for


class Command(BaseCommand):         # pragma: no cover
    help = "Check consistentie van historische uitslag, individueel"

    def __init__(self, stdout=None, stderr=None, no_color=False, force_color=False):
        super().__init__(stdout, stderr, no_color, force_color)
        self.dryrun = True
        self.verbose = True

    def add_arguments(self, parser):
        parser.add_argument('--dryrun', action='store_true')
        parser.add_argument('--verbose', action='store_true')
        parser.add_argument('seizoen', nargs=1, help="competitie seizoen: 20xx/20yy")
        parser.add_argument('comptype', nargs=1, choices=('18', '25'), help="competitie type: 18 of 25")
        parser.add_argument('bestand', type=str, help='Pad naar het Excel bestand')
        parser.add_argument('tabblad', type=str, help='Naam van het tabblad')
        parser.add_argument('col_lid_nr', type=str, help='Kolomletter waarin het sporter lid nummer staat')
        parser.add_argument('col_ver_nr', type=str, help='Kolomletter waarin het verenigingsnummer staat')
        parser.add_argument('col_naam', type=str, help='Kolomletter waarin de sporter naam staat')

    def handle(self, *args, **options):
        self.dryrun = options['dryrun']
        self.verbose = options['verbose']
        comp_type = options['comptype'][0]
        seizoen = options['seizoen'][0]
        fname = options['bestand']
        ws_name = options['tabblad']
        col_lid_nr = options['col_lid_nr']
        col_ver_nr = options['col_ver_nr']
        col_naam = options['col_naam']

        try:
            hist_seizoen = HistCompSeizoen.objects.get(seizoen=seizoen, comp_type=comp_type)
        except HistCompSeizoen.DoesNotExist:
            self.stderr.write('[ERROR] Historisch seizoen %s - %s niet gevonden' % (repr(seizoen), repr(comp_type)))
            return
        self.stdout.write('[INFO] Seizoen: %s' % hist_seizoen)

        self.stdout.write('[INFO] Lees bestand %s' % repr(fname))
        try:
            prg = openpyxl.load_workbook(fname,
                                         data_only=True)  # do not evaluate formulas; use last calculated values
        except (OSError, zipfile.BadZipFile, KeyError, InvalidFileException) as exc:
            self.stderr.write('[ERROR] Kan het excel bestand niet openen (%s)' % str(exc))
            return

        try:
            ws = prg[ws_name]
        except KeyError:        # pragma: no cover
            self.stderr.write('[ERROR] Kan tabblad %s niet vinden' % repr(ws_name))
            return

        lid_nr2indiv = dict()       # [lid_nr] = [indiv, indiv, ..]
        for indiv in HistCompRegioIndiv.objects.filter(seizoen=hist_seizoen):
            try:
                lid_nr2indiv[indiv.sporter_lid_nr].append(indiv)
            except KeyError:
                lid_nr2indiv[indiv.sporter_lid_nr] = [indiv]
        # for

        ver_nr2ver = dict()
        for ver in NhbVereniging.objects.select_related('regio'):
            ver_nr2ver[ver.ver_nr] = ver
        # for

        # al verwijderde verenigingen toevoegen
        for ver in verwijderde_verenigingen():
            ver_nr2ver[ver.ver_nr] = ver
        # for

        row_nr = 2 - 1      # skip header
        while row_nr < 5000:
            row_nr += 1
            row_str = str(row_nr)

            lid_nr = ws[col_lid_nr + row_str].value
            ver_nr = ws[col_ver_nr + row_str].value

            if lid_nr is None and ver_nr is None:
                # einde van de lijst
                break   # from the while

            if ver_nr == 1377:
                self.stderr.write('[ERROR] Sporter %s van vereniging %s mag niet in de uitslag' % (lid_nr, ver_nr))
                # TODO: verwijder uit uitslag
                continue

            try:
                indiv_list = lid_nr2indiv[lid_nr]
            except KeyError:
                # sporter staat niet in de lijst
                # dat is normaal als er geen uitslag aan hing
                #self.stderr.write('[ERROR] lid_nr %s niet gevonden op regel %s' % (lid_nr, row_nr))
                continue

            try:
                ver = ver_nr2ver[ver_nr]
            except KeyError:
                self.stderr.write('[ERROR] Kan vereniging %s niet vinden' % ver_nr)
                continue

            # indiv_list: omdat sporter met meerdere bogen in de uitslag kan staan
            for indiv in indiv_list:
                if indiv.sporter_naam == '':
                    self.stdout.write('[WARNING] Geen sporter_naam voor pk=%s: %s' % (indiv.pk, indiv))
                    naam = ws[col_naam + row_str].value
                    naam = str(naam).strip()
                    naam = naam.encode('iso-8859-1').decode('utf-8')
                    if naam == naam.lower():
                        naam = naam.title()
                        naam = naam.replace(' De ', ' de ')
                        naam = naam.replace(' Der ', ' der ')
                        naam = naam.replace(' Van ', ' van ')
                        self.stderr.write('[WARNING] Fixing capitalization: %s' % repr(naam))
                    if len(naam) < 5:
                        self.stderr.write('[ERROR] Rejected: %s' % repr(naam))
                    else:
                        self.stdout.write('          Voorstel: %s' % repr(naam))
                        indiv.sporter_naam = naam
                        if not self.dryrun:
                            indiv.save(update_fields=['sporter_naam'])

                if indiv.vereniging_nr != ver_nr:
                    self.stdout.write('ver_nr %s --> %s for %s' % (indiv.vereniging_nr, ver_nr, indiv))
                    indiv.vereniging_nr = ver_nr
                    indiv.vereniging_naam = ver.naam
                    indiv.vereniging_plaats = ver.plaats
                    indiv.regio_nr = ver.regio.regio_nr
                    if not self.dryrun:
                        indiv.save(update_fields=['vereniging_nr', 'vereniging_naam', 'vereniging_plaats'])
            # for


# end of file
