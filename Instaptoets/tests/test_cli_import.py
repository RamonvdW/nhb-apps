# -*- coding: utf-8 -*-

#  Copyright (c) 2021-2025 Ramon van der Winkel.
#  All rights reserved.
#  Licensed under BSD-3-Clause-Clear. See LICENSE file for details.

from django.test import TestCase
from django.utils import timezone
from Geo.models import Regio
from Instaptoets.models import Instaptoets
from Sporter.models import Sporter
from TestHelpers.e2ehelpers import E2EHelpers
from Vereniging.models import Vereniging
import datetime
import openpyxl


class TestInstaptoetsCli(E2EHelpers, TestCase):
    """ unittests voor de Instaptoets applicatie, management command import_instaptoets """

    hdrs = ['Vraagtekst', 'Antwoord A', 'Antwoord B', 'Antwoord C', 'Antwoord D', 'Juiste antwoord', 'Toets', 'Quiz']

    vraag_1 = ['Dit is de vraag', 'Eerste antwoord', 'Tweede antwoord', '', '', 'A', 'Ja', 'Ja']
    vraag_2 = ['Dit is de vraag', 'Eerste antwoord', 'Tweede antwoord', 'Derde', 'Vierde', 'C', 'Nee', 'Ja']
    vraag_te_kort = ['Vraag zonder antwoorden',]

    def setUp(self):
        """ initialisatie van de test case """
        self.tmp_fname = '/tmp/test_instaptoets.xlsx'

    def test_basis(self):
        # onvolledige argumenten
        f1, f2 = self.run_management_command('import_instaptoets',
                                             report_exit_code=False)
        self.assertTrue("raised CommandError('Error: the following arguments are required: filename" in f1.getvalue())

        # bestand bestaat niet
        f1, f2 = self.run_management_command('import_instaptoets', 'bestaat-niet')
        self.assertTrue("[ERROR] Kan bestand 'bestaat-niet' niet lezen" in f1.getvalue())

        # leeg bestand
        wb = openpyxl.Workbook()
        wb.save(self.tmp_fname)
        f1, f2 = self.run_management_command('import_instaptoets', self.tmp_fname)
        self.assertTrue("[ERROR] Kan correcte header niet vinden." in f1.getvalue())

        # ws met juiste headers
        ws = wb.create_sheet('Test Categorie')
        ws.append(self.hdrs)
        wb.save(self.tmp_fname)
        f1, f2 = self.run_management_command('import_instaptoets', self.tmp_fname, '--verbose')
        self.assertTrue("[INFO] Aantal vragen is nu 0" in f2.getvalue())

        ws.append(self.vraag_1)
        ws.append(self.vraag_2)
        ws.append(self.vraag_te_kort)
        wb.save(self.tmp_fname)

        f1, f2 = self.run_management_command('import_instaptoets', self.tmp_fname, '--verbose')
        # print('\nf1: %s\nf2: %s' % (f1.getvalue(), f2.getvalue()))
        self.assertTrue("[INFO] Aantal vragen is nu 2" in f2.getvalue())

        # nog een keer importeren, dan bestaan alle vragen al
        # verwijder een vraag
        wb = openpyxl.Workbook()
        ws = wb.create_sheet('Test Categorie')
        ws.append(self.hdrs)
        ws.append(self.vraag_1)
        wb.save(self.tmp_fname)
        f1, f2 = self.run_management_command('import_instaptoets', self.tmp_fname, '--verbose')
        print('\nf1: %s\nf2: %s' % (f1.getvalue(), f2.getvalue()))

# end of file
