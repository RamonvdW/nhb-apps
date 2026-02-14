# -*- coding: utf-8 -*-

#  Copyright (c) 2020-2026 Ramon van der Winkel.
#  All rights reserved.
#  Licensed under BSD-3-Clause-Clear. See LICENSE file for details.

from django.conf import settings
from django.urls import reverse
from django.http import Http404, HttpResponse
from django.utils import timezone
from django.views.generic import TemplateView
from django.contrib.auth.mixins import UserPassesTestMixin
from Competitie.definities import DEEL_RK, DEELNAME_NEE, DEELNAME2STR
from Competitie.models import (CompetitieIndivKlasse, CompetitieTeamKlasse, CompetitieMatch,
                               KampioenschapIndivKlasseLimiet, KampioenschapSporterBoog, KampioenschapTeam)
from CompKampioenschap.operations import get_url_wedstrijdformulier
from Functie.definities import Rol
from Functie.rol import rol_get_huidige_functie
from Scheidsrechter.models import MatchScheidsrechters
from Sporter.models import SporterVoorkeuren
from tempfile import NamedTemporaryFile
from copy import copy
import openpyxl
import textwrap
import zipfile
import shutil
import os


TEMPLATE_DOWNLOAD_RK_FORMULIER = 'complaagrayon/hwl-download-rk-formulier.dtl'

CONTENT_TYPE_XLSX = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'     # noqa


class DownloadRkFormulierView(UserPassesTestMixin, TemplateView):

    """ Toon de HWL of WL de waarschijnlijke deelnemerslijst voor een wedstrijd bij deze vereniging
    """

    # class variables shared by all instances
    template_name = TEMPLATE_DOWNLOAD_RK_FORMULIER
    raise_exception = True  # genereer PermissionDefined als test_func False terug geeft
    permission_denied_message = 'Geen toegang'

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.rol_nu, self.functie_nu = None, None

    def test_func(self):
        """ called by the UserPassesTestMixin to verify the user has permissions to use this view """
        self.rol_nu, self.functie_nu = rol_get_huidige_functie(self.request)
        return self.functie_nu and self.rol_nu in (Rol.ROL_SEC, Rol.ROL_HWL, Rol.ROL_WL)

    def get_context_data(self, **kwargs):
        """ called by the template system to get the context data for the template """
        context = super().get_context_data(**kwargs)

        try:
            match_pk = int(kwargs['match_pk'][:6])      # afkappen voor de veiligheid
            match = (CompetitieMatch
                     .objects
                     .select_related('vereniging',
                                     'locatie')
                     .prefetch_related('indiv_klassen',
                                       'team_klassen')
                     .get(pk=match_pk,
                          vereniging=self.functie_nu.vereniging))
        except (ValueError, CompetitieMatch.DoesNotExist):
            raise Http404('Wedstrijd niet gevonden')

        deelkamps = match.kampioenschap_set.filter(deel=DEEL_RK)
        if len(deelkamps) == 0:
            raise Http404('Geen kampioenschap')
        deelkamp = deelkamps[0]

        context['deelkamp'] = deelkamp
        context['wedstrijd'] = match
        context['vereniging'] = match.vereniging        # als we hier komen is dit altijd bekend
        context['locatie'] = match.locatie
        context['aantal_banen'] = '?'

        comp = deelkamp.competitie
        # TODO: check fase
        if comp.is_indoor():
            aantal_pijlen = 30
            if match.locatie:
                context['aantal_banen'] = match.locatie.banen_18m
        else:
            aantal_pijlen = 25
            if match.locatie:
                context['aantal_banen'] = match.locatie.banen_25m

        match.is_rk = True
        match.beschrijving = "Rayonkampioenschap"

        heeft_indiv = heeft_teams = False
        beschr = list()

        klasse_indiv_pks = list()
        klasse_team_pks = list()
        match.klassen_lijst = klassen_lijst = list()
        for klasse in match.indiv_klassen.select_related('boogtype').all():
            klassen_lijst.append(str(klasse))
            klasse_indiv_pks.append(klasse.pk)
            if not heeft_indiv:
                heeft_indiv = True
                beschr.append('Individueel')
        # for
        for klasse in match.team_klassen.all():
            klassen_lijst.append(klasse.beschrijving)
            klasse_team_pks.append(klasse.pk)
            if not heeft_teams:
                heeft_teams = True
                beschr.append('Teams')
        # for

        lid2voorkeuren = dict()  # [lid_nr] = (SporterVoorkeuren: .para_voorwerpen, .opmerking_para_sporter)
        for tup in SporterVoorkeuren.objects.select_related('sporter').values_list('sporter__lid_nr',
                                                                                   'para_voorwerpen',
                                                                                   'opmerking_para_sporter'):
            lid_nr, para_voorwerpen, opmerking_para_sporter = tup
            lid2voorkeuren[lid_nr] = (para_voorwerpen, opmerking_para_sporter)
        # for

        vastgesteld = timezone.localtime(timezone.now())
        context['vastgesteld'] = vastgesteld

        context['heeft_indiv'] = heeft_indiv
        context['heeft_teams'] = heeft_teams
        context['beschrijving'] = "%s %s" % (match.beschrijving, " en ".join(beschr))

        # zoek de deelnemers erbij
        if heeft_indiv:
            deelnemers = (KampioenschapSporterBoog
                          .objects
                          .filter(kampioenschap=deelkamp,
                                  indiv_klasse__pk__in=klasse_indiv_pks)
                          .exclude(deelname=DEELNAME_NEE)
                          .select_related('sporterboog',
                                          'sporterboog__sporter',
                                          'bij_vereniging',
                                          'indiv_klasse')
                          .order_by('indiv_klasse',
                                    'rank'))
            context['deelnemers_indiv'] = deelnemers

            prev_klasse = None
            for deelnemer in deelnemers:
                if deelnemer.indiv_klasse != prev_klasse:
                    deelnemer.break_before = True
                    deelnemer.url_open_indiv = get_url_wedstrijdformulier(comp.begin_jaar, int(comp.afstand),
                                                                          deelkamp.rayon.rayon_nr,
                                                                          deelnemer.indiv_klasse.pk,
                                                                          is_bk=False, is_teams=False)
                    prev_klasse = deelnemer.indiv_klasse

                deelnemer.ver_nr = deelnemer.bij_vereniging.ver_nr
                deelnemer.ver_naam = deelnemer.bij_vereniging.naam
                deelnemer.lid_nr = deelnemer.sporterboog.sporter.lid_nr
                deelnemer.volledige_naam = deelnemer.sporterboog.sporter.volledige_naam()
                deelnemer.gemiddelde_str = "%.3f" % deelnemer.gemiddelde
                deelnemer.gemiddelde_str = deelnemer.gemiddelde_str.replace('.', ',')

                try:
                    voorkeuren_para_voorwerpen, voorkeuren_opmerking_para_sporter = lid2voorkeuren[deelnemer.lid_nr]
                except KeyError:        # pragma: no cover
                    pass
                else:
                    if voorkeuren_para_voorwerpen:
                        if deelnemer.kampioen_label != '':
                            deelnemer.kampioen_label += ';\n'
                        deelnemer.kampioen_label += 'Sporter laat voorwerpen\nop de schietlijn staan'

                    if voorkeuren_opmerking_para_sporter:
                        if deelnemer.kampioen_label != '':
                            deelnemer.kampioen_label += ';\n'
                        deelnemer.kampioen_label += textwrap.fill(voorkeuren_opmerking_para_sporter, 30)

                deelnemer.url_wijzig = reverse('CompLaagRayon:wijzig-status-rk-deelnemer',
                                               kwargs={'deelnemer_pk': deelnemer.pk})
            # for

        if heeft_teams:
            teams = (KampioenschapTeam
                     .objects
                     .filter(kampioenschap=deelkamp,
                             team_klasse__pk__in=klasse_team_pks)
                     .select_related('vereniging',
                                     'team_klasse')
                     .prefetch_related('gekoppelde_leden')
                     .order_by('team_klasse',               # TODO: volgorde?
                               '-aanvangsgemiddelde'))      # sterkste team bovenaan
            context['deelnemers_teams'] = teams

            if not comp.klassengrenzen_vastgesteld_rk_bk:
                context['geen_klassengrenzen'] = True

            volg_nr = 0
            prev_klasse = None
            for team in teams:
                if team.team_klasse != prev_klasse:
                    team.break_before = True
                    team.url_download_teams = reverse('CompLaagRayon:formulier-teams-als-bestand',
                                                      kwargs={'match_pk': match.pk,
                                                              'klasse_pk': team.team_klasse.pk})

                    prev_klasse = team.team_klasse
                    volg_nr = 0

                volg_nr += 1
                team.volg_nr = volg_nr
                team.ver_nr = team.vereniging.ver_nr
                team.ver_naam = team.vereniging.naam
                sterkte = float(team.aanvangsgemiddelde) * aantal_pijlen
                team.sterkte_str = "%.1f" % sterkte
                team.sterkte_str = team.sterkte_str.replace('.', ',')

                team.gekoppelde_leden_lijst = list()
                for lid in team.gekoppelde_leden.select_related('sporterboog__sporter').order_by('-gemiddelde'):
                    sporter = lid.sporterboog.sporter
                    lid.naam_str = "[%s] %s" % (sporter.lid_nr, sporter.volledige_naam())
                    lid.gem_str = '%.3f' % lid.gemiddelde

                    try:
                        voorkeuren_para_voorwerpen, voorkeuren_opmerking_para_sporter = lid2voorkeuren[sporter.lid_nr]
                    except KeyError:        # pragma: no cover
                        pass
                    else:
                        if voorkeuren_para_voorwerpen or len(voorkeuren_opmerking_para_sporter) > 1:
                            lid.is_para = True
                            context['toon_para_uitleg'] = True

                    team.gekoppelde_leden_lijst.append(lid)
                # for
            # for

        if match.aantal_scheids > 0:
            match_sr = MatchScheidsrechters.objects.filter(match=match).first()
            if match_sr:
                aantal = 0
                for sr in (match_sr.gekozen_hoofd_sr, match_sr.gekozen_sr1, match_sr.gekozen_sr2):
                    if sr:
                        aantal += 1
                # for
                if aantal > 0:
                    context['aantal_sr_str'] = "%s scheidsrechter" % aantal
                    if aantal > 1:
                        context['aantal_sr_str'] += 's'

                    context['url_sr_contact'] = reverse('Scheidsrechter:match-hwl-contact',
                                                        kwargs={'match_pk': match.pk})

        context['kruimels'] = (
            (reverse('Vereniging:overzicht'), 'Beheer vereniging'),
            (reverse('CompScores:wedstrijden'), 'Competitiewedstrijden'),
            (None, 'RK programma')
        )

        return context


class FormulierTeamsAlsBestandView(UserPassesTestMixin, TemplateView):

    """ Geef de HWL het ingevulde wedstrijdformulier voor een RK teams wedstrijd bij deze vereniging """

    # class variables shared by all instances
    raise_exception = True  # genereer PermissionDefined als test_func False terug geeft
    permission_denied_message = 'Geen toegang'

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.rol_nu, self.functie_nu = None, None

        self.match = None
        self.deelkamp = None
        self.team_klasse = None
        self.klasse_str = ''
        self.titel = ''
        self.lid2voorkeuren = dict()  # [lid_nr] = (SporterVoorkeuren: .para_voorwerpen, .opmerking_para_sporter)
        self.deelnemers_ver_nrs = list()
        self.vastgesteld = timezone.localtime(timezone.now())

    def test_func(self):
        """ called by the UserPassesTestMixin to verify the user has permissions to use this view """
        self.rol_nu, self.functie_nu = rol_get_huidige_functie(self.request)
        return self.functie_nu and self.rol_nu in (Rol.ROL_HWL, Rol.ROL_WL)

    def _laad_lid_voorkeuren(self):
        self.lid2voorkeuren = dict()  # [lid_nr] = (SporterVoorkeuren: .para_voorwerpen, .opmerking_para_sporter)
        for tup in SporterVoorkeuren.objects.select_related('sporter').values_list('sporter__lid_nr',
                                                                                   'para_voorwerpen',
                                                                                   'opmerking_para_sporter'):
            lid_nr, para_voorwerpen, opmerking_para_sporter = tup
            self.lid2voorkeuren[lid_nr] = (para_voorwerpen, opmerking_para_sporter)
        # for

    def _vul_deelnemers(self, prg):
        # maak wijzigingen in het RK programma
        ws = prg['Deelnemers']

        ws['B1'] = self.titel
        ws['B2'] = self.klasse_str

        # organisatie
        if self.match.locatie:
            ws['B3'] = self.match.vereniging.naam + ', ' + self.match.locatie.plaats
        else:
            ws['B3'] = self.match.vereniging.naam

        ws['B4'] = self.match.datum_wanneer.strftime('%Y-%m-%d')

        teams = (KampioenschapTeam
                 .objects
                 .filter(kampioenschap=self.deelkamp,
                         team_klasse=self.team_klasse.pk)
                 .select_related('vereniging',
                                 'vereniging__regio')
                 .prefetch_related('gekoppelde_leden')
                 .order_by('-aanvangsgemiddelde'))      # sterkste team bovenaan

        self.deelnemers_ver_nrs = list()

        volg_nr = 0
        for team in teams:
            row_nr = 8 + volg_nr * 5
            row = str(row_nr)

            ver = team.vereniging
            if ver.ver_nr not in self.deelnemers_ver_nrs:
                self.deelnemers_ver_nrs.append(ver.ver_nr)

            # team naam
            ws['B' + row] = team.team_naam

            # ver_nr
            ws['C' + row] = str(ver.ver_nr)

            # team sterkte
            # sterkte_str = "%.1f" % (team.aanvangsgemiddelde * aantal_pijlen)
            # sterkte_str = sterkte_str.replace('.', ',')
            # ws['G' + row] = sterkte_str

            # vul de 3 sporters in
            aantal = 0
            for deelnemer in team.gekoppelde_leden.select_related('sporterboog__sporter'):
                row_nr += 1
                row = str(row_nr)

                sporter = deelnemer.sporterboog.sporter

                para_mark = False
                try:
                    voorkeuren_para_voorwerpen, voorkeuren_opmerking_para_sporter = self.lid2voorkeuren[sporter.lid_nr]
                except KeyError:        # pragma: no cover
                    pass
                else:
                    if voorkeuren_para_voorwerpen or voorkeuren_opmerking_para_sporter:
                        para_mark = True

                # volledige naam
                naam_str = sporter.volledige_naam()
                if para_mark:
                    naam_str += ' **'
                ws['B' + row] = naam_str

                # bondsnummer
                ws['C' + row] = str(sporter.lid_nr)

                # regio gemiddelde
                gem_str = '%.3f' % deelnemer.gemiddelde
                gem_str = gem_str.replace('.', ',')         # NL komma
                ws['D' + row] = gem_str

                aantal += 1
            # for

            # altijd 3 regels invullen
            while aantal < 3:
                row_nr += 1
                row = str(row_nr)
                ws['B' + row] = 'n.v.t.'    # naam
                ws['C' + row] = '-'         # bondsnummer
                ws['D' + row] = ''          # gemiddelde
                aantal += 1
            # while

            volg_nr += 1
        # for

        # 5 of meer teams?
        ws['D5'] = "Ja" if volg_nr > 4 else "Nee"

        while volg_nr < 8:
            row_nr = 8 + volg_nr * 5
            row = str(row_nr)

            # vereniging leeg maken
            ws['B' + row] = 'n.v.t.'     # vereniging
            ws['C' + row] = ''           # regio

            # sporters leegmaken
            aantal = 0
            while aantal < 3:
                row_nr += 1
                row = str(row_nr)
                ws['B' + row] = ''         # naam
                ws['C' + row] = ''         # bondsnummer
                ws['D' + row] = ''         # gemiddelde
                aantal += 1
            # while

            volg_nr += 1
        # while

        ws['A50'] = 'Deze gegevens zijn opgehaald op %s' % self.vastgesteld.strftime('%Y-%m-%d %H:%M:%S')

    def _vul_toegestane_deelnemers(self, prg):
        # alle gerechtigde deelnemers opnemen op een apart tabblad, met gemiddelde en boogtype

        ws = prg['Toegestane deelnemers']

        ws['B1'] = '%s, %s' % (self.titel, self.klasse_str)

        cd_font = ws['C18'].font
        c_align = ws['C18'].alignment
        c_format = ws['C18'].number_format

        d_align = ws['D18'].alignment
        d_format = ws['D18'].number_format

        efgh_font = ws['E18'].font      # noqa
        e_align = ws['E18'].alignment

        f_align = ws['F18'].alignment

        g_align = ws['G18'].alignment
        g_format = ws['G18'].number_format

        boog_typen = self.team_klasse.team_type.boog_typen.all()
        boog_pks = list(boog_typen.values_list('pk', flat=True))

        row_nr = 16
        prev_ver = None
        for deelnemer in (KampioenschapSporterBoog
                          .objects
                          .filter(kampioenschap=self.deelkamp,
                                  bij_vereniging__ver_nr__in=self.deelnemers_ver_nrs,
                                  sporterboog__boogtype__pk__in=boog_pks)       # filter op toegestane boogtypen
                          .select_related('bij_vereniging__regio',
                                          'sporterboog__sporter',
                                          'sporterboog__boogtype')
                          .order_by('bij_vereniging__regio',
                                    'bij_vereniging',
                                    '-gemiddelde')):                            # hoogste eerst

            row_nr += 1
            row = str(row_nr)

            # vereniging
            ver = deelnemer.bij_vereniging
            if ver != prev_ver:
                row_nr += 1     # extra lege regel
                row = str(row_nr)
                ws['C' + row] = ver.regio.regio_nr
                if row_nr != 18:
                    ws['C' + row].font = copy(cd_font)
                    ws['C' + row].alignment = copy(c_align)
                    ws['C' + row].number_format = copy(c_format)

                ws['D' + row] = '[%s] %s' % (ver.ver_nr, ver.naam)
                if row_nr != 18:
                    ws['D' + row].font = copy(cd_font)
                    ws['D' + row].alignment = copy(d_align)
                    ws['D' + row].number_format = copy(d_format)

                prev_ver = ver

            # sporter
            sporter = deelnemer.sporterboog.sporter

            para_notities = ''
            try:
                voorkeuren_para_voorwerpen, voorkeuren_opmerking_para_sporter = self.lid2voorkeuren[sporter.lid_nr]
            except KeyError:        # pragma: no cover
                pass
            else:
                if voorkeuren_para_voorwerpen:
                    para_notities = 'Sporter laat voorwerpen op de schietlijn staan. '

                if voorkeuren_opmerking_para_sporter:
                    para_notities += voorkeuren_opmerking_para_sporter
            para_notities = para_notities.strip()

            naam_str = sporter.volledige_naam()
            if para_notities:
                naam_str += ' **'
            ws['E' + row] = naam_str

            ws['F' + row] = str(sporter.lid_nr)

            gem_str = '%.3f' % deelnemer.gemiddelde
            gem_str = gem_str.replace('.', ',')  # NL komma
            ws['G' + row] = gem_str
            ws['H' + row] = deelnemer.sporterboog.boogtype.beschrijving

            if para_notities:
                ws['I' + row] = para_notities

            if row_nr != 18:
                ws['E' + row].font = copy(efgh_font)
                ws['F' + row].font = copy(efgh_font)
                ws['G' + row].font = copy(efgh_font)
                ws['H' + row].font = copy(efgh_font)
                ws['I' + row].font = copy(efgh_font)

                ws['E' + row].alignment = copy(e_align)
                ws['G' + row].alignment = copy(g_align)
                ws['G' + row].number_format = copy(g_format)
        # for

        row_nr += 2
        row = str(row_nr)
        ws['B' + row] = 'Deze gegevens zijn opgehaald op %s' % self.vastgesteld.strftime('%Y-%m-%d %H:%M:%S')
        ws['B' + row].font = copy(efgh_font)
        ws['B' + row].alignment = copy(f_align)

    def get(self, request, *args, **kwargs):
        """ Afhandelen van de GET request waarmee we een bestand terug geven. """

        try:
            match_pk = int(kwargs['match_pk'][:6])      # afkappen voor de veiligheid
            self.match = (CompetitieMatch
                          .objects
                          .select_related('vereniging',
                                          'locatie')
                          .get(pk=match_pk,
                               vereniging=self.functie_nu.vereniging))
        except (ValueError, CompetitieMatch.DoesNotExist):
            raise Http404('Wedstrijd niet gevonden')

        try:
            klasse_pk = int(kwargs['klasse_pk'][:6])            # afkappen voor de veiligheid
            self.team_klasse = (CompetitieTeamKlasse
                                .objects
                                .get(pk=klasse_pk))
        except (ValueError, CompetitieTeamKlasse.DoesNotExist):
            raise Http404('Klasse niet gevonden')

        self.deelkamp = self.match.kampioenschap_set.filter(deel=DEEL_RK).first()
        if not self.deelkamp:
            raise Http404('Geen kampioenschap')

        # comp = self.deelkamp.competitie
        # # TODO: check fase

        self._laad_lid_voorkeuren()

        self.klasse_str = self.team_klasse.beschrijving

        self.titel = 'RK Teams Rayon %s, %s' % (self.deelkamp.rayon.rayon_nr,
                                                self.deelkamp.competitie.beschrijving)

        # bepaal de naam van het terug te geven bestand
        fname = "rk-programma_teams-rayon%s_" % self.deelkamp.rayon.rayon_nr
        fname += self.klasse_str.lower().replace(' ', '-')
        fname += '.xlsx'

        excel_template_name = 'template-excel-rk-teams.xlsx'

        # make een kopie van het RK programma in een tijdelijk bestand
        fpath = os.path.join(settings.INSTALL_PATH, 'CompLaagRayon', 'files', excel_template_name)
        tmp_file = NamedTemporaryFile()

        try:
            shutil.copyfile(fpath, tmp_file.name)
        except FileNotFoundError:
            raise Http404('Kan RK programma niet vinden')

        # open de kopie, zodat we die aan kunnen passen
        try:
            prg = openpyxl.load_workbook(tmp_file)
        except (OSError, zipfile.BadZipFile, KeyError):
            raise Http404('Kan RK programma niet openen')

        self._vul_deelnemers(prg)
        self._vul_toegestane_deelnemers(prg)

        # geef het aangepaste RK programma aan de client
        response = HttpResponse(content_type=CONTENT_TYPE_XLSX)
        response['Content-Disposition'] = 'attachment; filename="%s"' % fname
        prg.save(response)      # noqa

        del prg
        tmp_file.close()

        return response


# end of file
