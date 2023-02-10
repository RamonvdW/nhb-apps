# -*- coding: utf-8 -*-

#  Copyright (c) 2022-2023 Ramon van der Winkel.
#  All rights reserved.
#  Licensed under BSD-3-Clause-Clear. See LICENSE file for details.

from django.core.management.base import BaseCommand
from Competitie.models import KampioenschapSporterBoog, KampioenschapTeam, DEEL_RK
from openpyxl.utils.exceptions import InvalidFileException
from decimal import Decimal
import openpyxl
import zipfile


class Command(BaseCommand):
    help = "Importeer uitslag team kampioenschap Indoor"

    def __init__(self, stdout=None, stderr=None, no_color=False, force_color=False):
        super().__init__(stdout, stderr, no_color, force_color)
        self.dryrun = True
        self.verbose = False
        self.deel = "?"
        self.deelnemers = dict()            # [lid_nr] = [KampioenschapSchutterBoog, ...]
        self.teams_cache = list()           # [KampioenschapTeam, ...]
        self.team_lid_nrs = dict()          # [team.pk] = [lid_nr, ...]
        self.ver_lid_nrs = dict()           # [ver_nr] = [lid_nr, ...]
        self.kamp_lid_nrs = list()          # [lid_nr, ...]     iedereen die geplaatst is voor de kampioenschappen
        self.deelnemende_teams = dict()     # [team naam] = KampioenschapTeam

    def add_arguments(self, parser):
        parser.add_argument('--dryrun', action='store_true')
        parser.add_argument('--verbose', action='store_true')
        parser.add_argument('bestand', type=str,
                            help='Pad naar het Excel bestand')

    def _bepaal_laag(self):
        # TODO: aan de hand van de competitie fase bepalen of dit een RK of BK uitslag moet zijn
        self.deel = DEEL_RK

    def _deelnemers_ophalen(self):
        for deelnemer in (KampioenschapSporterBoog
                          .objects
                          .filter(kampioenschap__competitie__afstand='18',
                                  kampioenschap__deel=self.deel)
                          .select_related('kampioenschap',
                                          'kampioenschap__nhb_rayon',
                                          'sporterboog__sporter',
                                          'sporterboog__boogtype',
                                          'indiv_klasse')):

            lid_nr = deelnemer.sporterboog.sporter.lid_nr
            ver_nr = deelnemer.bij_vereniging.ver_nr

            try:
                self.deelnemers[lid_nr].append(deelnemer)
            except KeyError:
                self.deelnemers[lid_nr] = [deelnemer]

            try:
                self.ver_lid_nrs[ver_nr].append(lid_nr)
            except KeyError:
                self.ver_lid_nrs[ver_nr] = [lid_nr]

            self.kamp_lid_nrs.append(lid_nr)
        # for

    def _teams_ophalen(self):
        for team in (KampioenschapTeam
                     .objects
                     .filter(kampioenschap__competitie__afstand='18',
                             kampioenschap__deel=self.deel)
                     .select_related('kampioenschap',
                                     'kampioenschap__nhb_rayon',
                                     'vereniging',
                                     'team_type',
                                     'team_klasse')
                     .prefetch_related('gekoppelde_leden',
                                       'feitelijke_leden')):

            self.teams_cache.append(team)
            self.team_lid_nrs[team.pk] = [deelnemer.sporterboog.sporter.lid_nr for deelnemer in team.gekoppelde_leden.all()]
        # for

    def _sort_op_gemiddelde(self, lid_nrs):
        gem = list()
        for lid_nr in lid_nrs:
            deelnemer_all = self.deelnemers[lid_nr]
            if len(deelnemer_all) == 1:
                deelnemer = deelnemer_all[0]
            else:
                self.stderr.write('[WARNING] TODO: bepaal juiste deelnemer uit %s' % repr(deelnemer_all))
                deelnemer = deelnemer_all[0]
            tup = (deelnemer.gemiddelde, lid_nr)
            gem.append(tup)
        # for
        gem.sort(reverse=True)
        return gem

    def _get_deelnemer(self, lid_nr, lid_ag):
        deelnemer_all = self.deelnemers[lid_nr]
        for deelnemer in deelnemer_all:
            if abs(deelnemer.gemiddelde - lid_ag) < 0.0001:
                return deelnemer
        # for

        self.stderr.write('[WARNING] TODO: bepaal juiste deelnemer met ag=%s uit\n%s' % (
                            lid_ag,
                            "\n".join(["%s / %s / %s" % (deelnemer,
                                                         deelnemer.sporterboog.boogtype.afkorting,
                                                         deelnemer.gemiddelde) for deelnemer in deelnemer_all])))
        deelnemer = deelnemer_all[0]
        return deelnemer

    def _get_team(self, team_naam, ver_nr, row_nr, team_klasse):
        if self.verbose:
            self.stdout.write('[DEBUG] get_team: team_klasse=%s' % repr(team_klasse))
            
        up_naam = team_naam.upper()
        sel_teams = list()
        for team in self.teams_cache:
            # self.stdout.write('[DEBUG] cached team: %s' % repr(team))
            if team.vereniging.ver_nr == ver_nr:
                if team.team_naam.upper() == up_naam:
                    if team_klasse is None or team.team_klasse == team_klasse:
                        sel_teams.append(team)
        # for

        kamp_team = None
        if len(sel_teams) == 1:
            kamp_team = sel_teams[0]
        elif len(sel_teams) > 1:
            self.stderr.write('[ERROR] Kan team %s van vereniging %s op regel %s niet kiezen uit\n%s' % (
                repr(team_naam), ver_nr, row_nr, "\n".join([str(team) for team in sel_teams])))

        if kamp_team is None:
            self.stderr.write('[ERROR] Kan team %s van vereniging %s op regel %s niet vinden' % (
                repr(team_naam), ver_nr, row_nr))

        return kamp_team

    def _importeer_voorronde(self, ws):
        col_ver_naam = 'D'
        col_team_naam = 'F'
        col_lid_nr = 'E'
        col_lid_ag = 'G'
        col_score1 = 'H'
        col_score2 = 'I'

        # doorloop alle regels van het excel blad en ga op zoek naar bondsnummers
        row_nr = 8
        nix_count = 0
        team_klasse = None
        kamp_teams = list()
        while nix_count < 10:
            row_nr += 1
            row = str(row_nr)

            # vind een vereniging + team naam
            ver_naam = ws[col_ver_naam + row].value
            team_naam = ws[col_team_naam + row].value

            if ver_naam is None:
                nix_count += 1
                continue

            nix_count = 0
            ver_nr = -1

            try:
                if ver_naam[0] == '[' and ver_naam[5:5+2] == '] ':
                    ver_nr = int(ver_naam[1:1+4])
            except ValueError:
                pass

            if self.verbose:
                self.stdout.write('[DEBUG] regel %s: ver_nr=%s, ver_naam=%s, team_naam=%s' % (
                                    row, ver_nr, repr(ver_naam), repr(team_naam)))

            if ver_nr < 0:
                continue

            # zoek het team erbij
            kamp_team = self._get_team(team_naam, ver_nr, row_nr, team_klasse)
            if kamp_team is None:
                continue

            if team_klasse is None:
                team_klasse = kamp_team.team_klasse
            else:
                if team_klasse != kamp_team.team_klasse:
                    self.stderr.write('[ERROR] Inconsistente team klasse op regel %s: %s (eerdere teams: %s)' % (
                                        row_nr, kamp_team.team_klasse, team_klasse))
                    continue

            self.deelnemende_teams[team_naam] = kamp_team

            ver_lid_nrs = self.ver_lid_nrs[ver_nr]
            team_lid_nrs = self.team_lid_nrs[kamp_team.pk]
            aantal_gekoppeld = len(team_lid_nrs)
            feitelijke_deelnemers = list()

            # haal de teamleden op
            gevonden_lid_nrs = list()
            for lp in range(10):            # pragma: no branch
                row_nr += 1
                row = str(row_nr)

                lid_nr = ws[col_lid_nr + row].value
                if lid_nr is None:
                    break
                if str(lid_nr) == '-':
                    continue

                if lid_nr not in self.kamp_lid_nrs:
                    self.stderr.write('[ERROR] Lid %s is niet gekwalificeerd voor dit kampioenschap!' % lid_nr)
                elif lid_nr not in ver_lid_nrs:
                    self.stderr.write('[ERROR] Lid %s is niet van vereniging %s!' % (lid_nr, ver_nr))
                else:
                    try:
                        lid_ag = round(Decimal(ws[col_lid_ag + row].value), 3)  # 3 cijfers achter de komma
                        score1 = int(ws[col_score1 + row].value)
                        score2 = int(ws[col_score2 + row].value)
                    except (ValueError, TypeError):
                        self.stderr.write('[ERROR] Probleem met de scores op regel %s' % row_nr)
                    else:
                        if score1 + score2 == 0:
                            # sporter heeft niet meegedaan
                            self.stdout.write('[WARNING] Geen scores voor sporter %s op regel %s' % (lid_nr, row_nr))
                        else:
                            deelnemer = self._get_deelnemer(lid_nr, lid_ag)
                            deelnemer.result_teamscore_1 = score1
                            deelnemer.result_teamscore_2 = score2
                            feitelijke_deelnemers.append(deelnemer)
                            gevonden_lid_nrs.append(lid_nr)
            # for

            # haal de verwachte lid_nrs eruit
            feitelijke_lid_nrs = [lid_nr for lid_nr in gevonden_lid_nrs if lid_nr in team_lid_nrs]
            if self.verbose:
                self.stdout.write('[DEBUG] feitelijke_lid_nrs van team [%s] %s: %s' % (ver_nr, team_naam, repr(feitelijke_lid_nrs)))

            if len(feitelijke_lid_nrs) == 0:
                self.stdout.write('[WARNING] Team %s van vereniging %s heeft niet meegedaan (geen scores)' % (ver_nr, team_naam))
                del self.deelnemende_teams[team_naam]
            else:
                for lid_nr in feitelijke_lid_nrs:
                    gevonden_lid_nrs.remove(lid_nr)
                    team_lid_nrs.remove(lid_nr)
                # for

                if len(gevonden_lid_nrs):
                    uitvallers = self._sort_op_gemiddelde(team_lid_nrs)
                    invallers = self._sort_op_gemiddelde(gevonden_lid_nrs)
                    afgekeurd = list()

                    if len(invallers) > len(uitvallers):
                        self.stderr.write('[ERROR] Te veel invallers voor team %s met max %s sporters (vereniging %s)' % (
                                            repr(team_naam), aantal_gekoppeld, ver_nr))
                        feitelijke_deelnemers = list()
                    else:
                        while len(invallers) > 0:
                            gemiddelde_in, lid_nr_in = invallers.pop(0)
                            gemiddelde_uit, lid_nr_uit = uitvallers[0]
                            if gemiddelde_in > gemiddelde_uit:
                                # mag niet invallen
                                afgekeurd.append(lid_nr_in)
                                self.stderr.write(
                                    '[ERROR] Te hoog gemiddelde %s voor invaller %s voor team %s van vereniging %s' % (
                                            gemiddelde_in, lid_nr_in, team_naam, ver_nr))
                                for gemiddelde, lid_nr in uitvallers:
                                    self.stderr.write('        Uitvaller %s heeft gemiddelde %s' % (lid_nr, gemiddelde))
                            else:
                                # mag wel invaller en vervangt de hoogste uitvaller
                                uitvallers.pop(0)
                        # while

                        feitelijke_deelnemers = [deelnemer for deelnemer in feitelijke_deelnemers if deelnemer.sporterboog.sporter.lid_nr not in afgekeurd]

                if len(feitelijke_deelnemers) > 0:
                    kamp_team.result_teamscore = 0
                    deelnemer_totalen = list()

                    for deelnemer in feitelijke_deelnemers:
                        if not self.dryrun:
                            # uitgestelde save actie
                            deelnemer.save(update_fields=['result_teamscore_1', 'result_teamscore_2'])
                        deelnemer_totaal = deelnemer.result_teamscore_1 + deelnemer.result_teamscore_2
                        deelnemer_totalen.append(deelnemer_totaal)
                    # for
                    deelnemer_totalen.sort(reverse=True)                        # hoogste eerst
                    kamp_team.result_teamscore = sum(deelnemer_totalen[:3])     # de 3 hoogste gebruiken
                    kamp_team.feitelijke_leden.set(feitelijke_deelnemers)

                    deelnemer_totalen.insert(0, kamp_team.pk)                   # backup voor sorteren
                    deelnemer_totalen.insert(0, kamp_team.result_teamscore)     # resultaat vooraan
                    deelnemer_totalen.append(kamp_team)                         # team record laatst
                    kamp_teams.append(deelnemer_totalen)
        # while

        kamp_teams.sort(reverse=True)               # hoogste eerste
        rank = 0
        for tup in kamp_teams:
            rank += 1
            kamp_team = tup[-1]
            kamp_team.result_rank = rank
            kamp_team.result_volgorde = rank
            if not self.dryrun:
                kamp_team.save(update_fields=['result_rank', 'result_volgorde', 'result_teamscore'])
        # for

    def _zet_rank_en_volgorde(self, goud, zilver, brons, vierde, vijfden):

        # remove empty names
        if isinstance(brons, list):
            brons = [team_naam for team_naam in brons if team_naam]
            if len(brons) == 1:
                brons = brons[0]

        if self.verbose:
            self.stdout.write('[DEBUG] goud: %s' % repr(goud))
            self.stdout.write('[DEBUG] zilver: %s' % repr(zilver))
            self.stdout.write('[DEBUG] brons: %s' % repr(brons))
            self.stdout.write('[DEBUG] vierde: %s' % repr(vierde))
            self.stdout.write('[DEBUG] vijfden: %s' % repr(vijfden))

        rank = 1
        for team_naam in (goud, zilver, brons, vierde):
            if isinstance(team_naam, list):
                volgorde = rank
                for naam in team_naam:
                    kamp_team = self.deelnemende_teams[naam]
                    kamp_team.result_rank = rank
                    kamp_team.result_volgorde = volgorde
                    if not self.dryrun:
                        kamp_team.save(update_fields=['result_rank', 'result_volgorde'])
                    volgorde += 1
                # for
                rank += 1  # skip 4
            elif team_naam:
                kamp_team = self.deelnemende_teams[team_naam]
                kamp_team.result_rank = rank
                kamp_team.result_volgorde = rank
                if not self.dryrun:
                    kamp_team.save(update_fields=['result_rank', 'result_volgorde'])
            rank += 1
        # for

        # de rest is 5e
        # result_volgorde is al gezet, gebaseerd op aflopende scores
        for team_naam in vijfden:
            kamp_team = self.deelnemende_teams[team_naam]
            kamp_team.result_rank = 5
            if not self.dryrun:
                kamp_team.save(update_fields=['result_rank'])
        # for

    def _importeer_finales_8(self, ws):
        """ Lees de uitslag van het blad 'Finales 8 teams' """

        final_8 = list()
        for row_nr in (12, 14, 16, 18, 20, 22, 24, 26):
            team_naam = ws['B' + str(row_nr)].value
            if team_naam in (None, 'n.v.t.', 'nvt', 'n.v.t'): team_naam = ''
            if team_naam:
                if team_naam in self.deelnemende_teams.keys():
                    final_8.append(team_naam)
                else:
                    self.stdout.write('[WARNING] Kwartfinale team %s op finale blad wordt niet herkend' % repr(team_naam))
        # for
        # self.stdout.write('final_8: %s' % repr(final_8))

        # gouden finale
        team_naam_1 = ws['Q18'].value
        if team_naam_1 in (None, 'n.v.t.', 'nvt', 'n.v.t'): team_naam_1 = ''
        team_naam_2 = ws['Q20'].value
        if team_naam_2 in (None, 'n.v.t.', 'nvt', 'n.v.t'): team_naam_2 = ''
        if ws['V18'].value == '2e':
            goud = team_naam_2
            zilver = team_naam_1
        else:
            goud = team_naam_1
            zilver = team_naam_2

        # bronzen finale
        team_naam_1 = ws['Q30'].value
        if team_naam_1 in (None, 'n.v.t.', 'nvt', 'n.v.t'): team_naam_1 = ''
        team_naam_2 = ws['Q32'].value
        if team_naam_2 in (None, 'n.v.t.', 'nvt', 'n.v.t'): team_naam_2 = ''
        if ws['V30'].value == '3e':
            brons = team_naam_1
            vierde = team_naam_2
        elif ws['V30'].value == '4e':
            brons = team_naam_2
            vierde = team_naam_1
        else:
            # gelijk geëindigd
            brons = [team_naam_1, team_naam_2]
            vierde = None

        for team_naam in (goud, zilver, brons, vierde):
            if team_naam in final_8:
                final_8.remove(team_naam)
        # for

        self._zet_rank_en_volgorde(goud, zilver, brons, vierde, final_8)

    def _importeer_finales_4(self, ws):
        """ Lees de uitslag van het blad 'Finales 4 teams' """

        # gouden finale
        team_naam_1 = ws['M15'].value
        if team_naam_1 in (None, 'n.v.t.', 'nvt', 'n.v.t'): team_naam_1 = ''
        team_naam_2 = ws['M17'].value
        if team_naam_2 in (None, 'n.v.t.', 'nvt', 'n.v.t'): team_naam_2 = ''
        if ws['R15'].value == '2e':
            goud = team_naam_2
            zilver = team_naam_1
        else:
            goud = team_naam_1
            zilver = team_naam_2

        # bronzen finale
        team_naam_1 = ws['M30'].value
        if team_naam_1 in (None, 'n.v.t.', 'nvt', 'n.v.t'): team_naam_1 = ''
        team_naam_2 = ws['M32'].value
        if team_naam_2 in (None, 'n.v.t.', 'nvt', 'n.v.t'): team_naam_2 = ''
        if ws['R30'].value == '3e':
            brons = team_naam_1
            vierde = team_naam_2
        elif ws['R30'].value == '4e':
            brons = team_naam_2
            vierde = team_naam_1
        else:
            # gelijk geëindigd
            brons = [team_naam_1, team_naam_2]
            vierde = None

        self._zet_rank_en_volgorde(goud, zilver, brons, vierde, [])

    def handle(self, *args, **options):

        self.dryrun = options['dryrun']
        self.verbose = options['verbose']

        # open de kopie, zodat we die aan kunnen passen
        fname = options['bestand']
        self.stdout.write('[INFO] Lees bestand %s' % repr(fname))
        try:
            prg = openpyxl.load_workbook(fname,
                                         data_only=True)        # do not evaluate formulas; use last calculated values
        except (OSError, zipfile.BadZipFile, KeyError, InvalidFileException) as exc:
            self.stderr.write('[ERROR] Kan het excel bestand niet openen (%s)' % str(exc))
            return

        blad = 'Deelnemers en Scores'
        try:
            ws = prg[blad]
        except KeyError:
            self.stderr.write('[ERROR] Kan blad %s niet vinden' % repr(blad))
            return

        self._bepaal_laag()
        self._deelnemers_ophalen()
        self._teams_ophalen()

        self._importeer_voorronde(ws)

        if len(self.deelnemende_teams) == 0:
            self.stdout.write('[INFO] Geen deelnemende teams, dus geen kampioen')
        else:
            blad = 'Finales 8 teams'
            try:
                ws = prg[blad]
            except KeyError:
                self.stderr.write('[ERROR] Kan blad %s niet vinden' % repr(blad))
                return

            winnaar = ws['V18'].value
            if not winnaar:
                # None or leeg
                # dit blad is niet gebruikt - probeer het blad met finales met 4 teams
                blad = 'Finales 4 teams'
                try:
                    ws = prg[blad]
                except KeyError:
                    self.stderr.write('[ERROR] Kan blad %s niet vinden' % repr(blad))
                    return

                winnaar = ws['R15'].value
                if not winnaar:
                    # None or leeg
                    self.stderr.write('[ERROR] Kan juiste finale blad niet bepalen (geen WINNAAR)')
                    return

                self.stdout.write('[INFO] Uitslag wordt van blad %s gehaald' % repr(blad))
                self._importeer_finales_4(ws)
            else:
                self.stdout.write('[INFO] Uitslag wordt van blad %s gehaald' % repr(blad))
                self._importeer_finales_8(ws)

            result = list()
            for team_naam, team in self.deelnemende_teams.items():
                tup = (team.result_rank, team.result_volgorde, team_naam)
                result.append(tup)
            # for
            result.sort()
            self.stdout.write('Resultaat:')
            for rank, volgorde, naam in result:
                self.stdout.write('%s %s %s' % (rank, volgorde, naam))

# end of file
