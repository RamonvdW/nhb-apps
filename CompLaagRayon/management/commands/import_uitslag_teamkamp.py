# -*- coding: utf-8 -*-

#  Copyright (c) 2022 Ramon van der Winkel.
#  All rights reserved.
#  Licensed under BSD-3-Clause-Clear. See LICENSE file for details.

from django.core.management.base import BaseCommand
from Competitie.models import KampioenschapSporterBoog, KampioenschapTeam, DEEL_RK
from openpyxl.utils.exceptions import InvalidFileException
from decimal import Decimal
import openpyxl
import zipfile


class Command(BaseCommand):
    help = "Importeer uitslag team kampioenschap"

    def __init__(self, stdout=None, stderr=None, no_color=False, force_color=False):
        super().__init__(stdout, stderr, no_color, force_color)
        self.deelnemers = dict()        # [lid_nr] = [KampioenschapSchutterBoog, ...]
        self.teams_cache = list()       # [KampioenschapTeam, ...]
        self.team_lid_nrs = dict()      # [team.pk] = [lid_nr, ...]
        self.ver_lid_nrs = dict()       # [ver_nr] = [lid_nr, ...]
        self.kamp_lid_nrs = list()      # [lid_nr, ...]     iedereen die geplaatst is voor de kampioenschappen
        self.deel = "?"

    def add_arguments(self, parser):
        parser.add_argument('--dryrun', action='store_true')
        parser.add_argument('afstand', type=str,
                            help='Competitie afstand (18/25)')
        parser.add_argument('bestand', type=str,
                            help='Pad naar het Excel bestand')
        parser.add_argument('blad', type=str,
                            help='Naam van het blad met resultaten')
        parser.add_argument('kolommen', type=str, nargs='+',
                            help='Kolom letters: verenigingsnaam, teamnaam, bondsnummer, score1, score2')

    def _bepaal_laag(self, afstand):
        # TODO: aan de hand van de competitie fase bepalen of dit een RK of BK uitslag moet zijn
        self.deel = DEEL_RK

    def _deelnemers_ophalen(self, afstand):
        for deelnemer in (KampioenschapSporterBoog
                          .objects
                          .filter(kampioenschap__competitie__afstand=afstand,
                                  kampioenschap__deel=self.deel)
                          .select_related('kampioenschap',
                                          'kampioenschap__nhb_rayon',
                                          'sporterboog__sporter',
                                          'sporterboog__boogtype',
                                          'indiv_klasse')):

            lid_nr = deelnemer.sporterboog.sporter.lid_nr
            ver_nr = deelnemer.bij_vereniging.ver_nr

            try:
                self.deelnemers[lid_nr].append(deelnemer)
            except KeyError:
                self.deelnemers[lid_nr] = [deelnemer]

            try:
                self.ver_lid_nrs[ver_nr].append(lid_nr)
            except KeyError:
                self.ver_lid_nrs[ver_nr] = [lid_nr]

            self.kamp_lid_nrs.append(lid_nr)
        # for

    def _teams_ophalen(self, afstand):
        for team in (KampioenschapTeam
                     .objects
                     .filter(kampioenschap__competitie__afstand=afstand,
                             kampioenschap__deel=self.deel)
                     .select_related('kampioenschap',
                                     'kampioenschap__nhb_rayon',
                                     'vereniging',
                                     'team_type',
                                     'team_klasse')
                     .prefetch_related('gekoppelde_leden',
                                       'feitelijke_leden')):

            self.teams_cache.append(team)
            self.team_lid_nrs[team.pk] = [deelnemer.sporterboog.sporter.lid_nr for deelnemer in team.gekoppelde_leden.all()]
        # for

    def _sort_op_gemiddelde(self, lid_nrs):
        gem = list()
        for lid_nr in lid_nrs:
            deelnemer_all = self.deelnemers[lid_nr]
            if len(deelnemer_all) == 1:
                deelnemer = deelnemer_all[0]
            else:
                self.stderr.write('[WARNING] TODO: bepaal juiste deelnemer uit %s' % repr(deelnemer_all))
                deelnemer = deelnemer_all[0]
            tup = (deelnemer.gemiddelde, lid_nr)
            gem.append(tup)
        # for
        gem.sort(reverse=True)
        return gem

    def _get_deelnemer(self, lid_nr, lid_ag):
        deelnemer_all = self.deelnemers[lid_nr]
        for deelnemer in deelnemer_all:
            if abs(deelnemer.gemiddelde - lid_ag) < 0.0001:
                return deelnemer
        # for

        self.stderr.write('[WARNING] TODO: bepaal juiste deelnemer met ag=%s uit\n%s' % (
                            lid_ag,
                            "\n".join(["%s / %s / %s" % (deelnemer,
                                                         deelnemer.sporterboog.boogtype.afkorting,
                                                         deelnemer.gemiddelde) for deelnemer in deelnemer_all])))
        deelnemer = deelnemer_all[0]
        return deelnemer

    def _get_team(self, team_naam, ver_nr, row_nr, team_klasse):
        up_naam = team_naam.upper()
        sel_teams = list()
        for team in self.teams_cache:
            if team.vereniging.ver_nr == ver_nr:
                if team.team_naam.upper() == up_naam:
                    if team_klasse is None or team.team_klasse == team_klasse:
                        sel_teams.append(team)
        # for

        kamp_team = None
        if len(sel_teams) == 1:
            kamp_team = sel_teams[0]
        elif len(sel_teams) > 1:
            self.stderr.write('[ERROR] Kan team %s van vereniging %s op regel %s niet kiezen uit\n%s' % (
                repr(team_naam), ver_nr, row_nr, "\n".join([str(team) for team in sel_teams])))

        if kamp_team is None:
            self.stderr.write('[ERROR] Kan team %s van vereniging %s op regel %s niet vinden' % (
                repr(team_naam), ver_nr, row_nr))

        return kamp_team

    def handle(self, *args, **options):

        dryrun = options['dryrun']

        afstand = options['afstand']
        if afstand not in ('18', '25'):
            self.stderr.write('[ERROR] Afstand moet 18 of 25 zijn')
            return

        # open de kopie, zodat we die aan kunnen passen
        fname = options['bestand']
        self.stdout.write('[INFO] Lees bestand %s' % repr(fname))
        try:
            prg = openpyxl.load_workbook(fname,
                                         data_only=True)        # do not evaluate formulas; use last calculated values
        except (OSError, zipfile.BadZipFile, KeyError, InvalidFileException) as exc:
            self.stderr.write('[ERROR] Kan het excel bestand niet openen (%s)' % str(exc))
            return

        blad = options['blad']
        try:
            ws = prg[blad]
        except KeyError:
            self.stderr.write('[ERROR] Kan blad %s niet vinden' % repr(blad))
            return

        cols = options['kolommen']
        if afstand == '25':
            if len(cols) != 6:
                self.stderr.write('[ERROR] Vereiste kolommen: verenigingsnaam, teamnaam, bondsnummer, ag, score1, score2')
                return

            col_ver_naam = cols[0]
            col_team_naam = cols[1]
            col_lid_nr = cols[2]
            col_lid_ag = cols[3]
            col_score1 = cols[4]
            col_score2 = cols[5]
        else:
            # indoor nog niet ondersteund
            self.stderr.write('[ERROR] Indoor nog niet ondersteund')
            return

        self._bepaal_laag(afstand)
        self._deelnemers_ophalen(afstand)
        self._teams_ophalen(afstand)

        # doorloop alle regels van het excel blad en ga op zoek naar bondsnummers
        row_nr = 8
        nix_count = 0
        team_klasse = None
        kamp_teams = list()
        while nix_count < 10:
            row_nr += 1
            row = str(row_nr)

            # vind een vereniging + team naam
            ver_naam = ws[col_ver_naam + row].value
            team_naam = ws[col_team_naam + row].value

            if ver_naam is None:
                nix_count += 1
                continue

            nix_count = 0
            ver_nr = -1

            try:
                if ver_naam[0] == '[' and ver_naam[5:5+2] == '] ':
                    ver_nr = int(ver_naam[1:1+4])
            except ValueError:
                pass

            self.stdout.write('[DEBUG] regel %s: ver_nr=%s, ver_naam=%s, team_naam=%s' % (row, ver_nr, repr(ver_naam), repr(team_naam)))

            if ver_nr < 0:
                continue

            # zoek het team erbij
            kamp_team = self._get_team(team_naam, ver_nr, row_nr, team_klasse)
            if kamp_team is None:
                continue

            if team_klasse is None:
                team_klasse = kamp_team.team_klasse
            else:
                if team_klasse != kamp_team.team_klasse:
                    self.stderr.write('[ERROR] Inconsistente team klasse op regel %s: %s (eerdere teams: %s)' % (row_nr, kamp_team.team_klasse, team_klasse))
                    continue

            ver_lid_nrs = self.ver_lid_nrs[ver_nr]
            team_lid_nrs = self.team_lid_nrs[kamp_team.pk]
            aantal_gekoppeld = len(team_lid_nrs)
            feitelijke_deelnemers = list()

            # haal de teamleden op
            gevonden_lid_nrs = list()
            for lp in range(10):            # pragma: no branch
                row_nr += 1
                row = str(row_nr)

                lid_nr = ws[col_lid_nr + row].value
                if lid_nr is None:
                    break
                if str(lid_nr) == '-':
                    continue

                if lid_nr not in self.kamp_lid_nrs:
                    self.stderr.write('[ERROR] Lid %s is niet gekwalificeerd voor dit kampioenschap!' % lid_nr)
                elif lid_nr not in ver_lid_nrs:
                    self.stderr.write('[ERROR] Lid %s is niet van vereniging %s!' % (lid_nr, ver_nr))
                else:
                    try:
                        lid_ag = round(Decimal(ws[col_lid_ag + row].value), 3)  # 3 cijfers achter de komma
                        score1 = int(ws[col_score1 + row].value)
                        score2 = int(ws[col_score2 + row].value)
                    except (ValueError, TypeError):
                        self.stderr.write('[ERROR] Probleem met de scores op regel %s' % row_nr)
                    else:
                        if score1 + score2 == 0:
                            # sporter heeft niet meegedaan
                            self.stdout.write('[WARNING] Geen scores voor sporter %s op regel %s' % (lid_nr, row_nr))
                        else:
                            deelnemer = self._get_deelnemer(lid_nr, lid_ag)
                            deelnemer.result_teamscore_1 = score1
                            deelnemer.result_teamscore_2 = score2
                            feitelijke_deelnemers.append(deelnemer)
                            gevonden_lid_nrs.append(lid_nr)
            # for

            # haal de verwachte lid_nrs eruit
            feitelijke_lid_nrs = [lid_nr for lid_nr in gevonden_lid_nrs if lid_nr in team_lid_nrs]
            for lid_nr in feitelijke_lid_nrs:
                gevonden_lid_nrs.remove(lid_nr)
                team_lid_nrs.remove(lid_nr)
            # for

            if len(gevonden_lid_nrs):
                uitvallers = self._sort_op_gemiddelde(team_lid_nrs)
                invallers = self._sort_op_gemiddelde(gevonden_lid_nrs)
                afgekeurd = list()

                if len(invallers) > len(uitvallers):
                    self.stderr.write('[ERROR] Te veel invallers voor team %s met max %s sporters (vereniging %s)' % (
                                        repr(team_naam), aantal_gekoppeld, ver_nr))
                    feitelijke_deelnemers = list()
                else:
                    while len(invallers) > 0:
                        gemiddelde_in, lid_nr_in = invallers.pop(0)
                        gemiddelde_uit, lid_nr_uit = uitvallers[0]
                        if gemiddelde_in > gemiddelde_uit:
                            # mag niet invallen
                            afgekeurd.append(lid_nr_in)
                            self.stderr.write(
                                '[ERROR] Te hoog gemiddelde %s voor invaller %s voor team %s van vereniging %s' % (
                                        gemiddelde_in, lid_nr_in, team_naam, ver_nr))
                            for gemiddelde, lid_nr in uitvallers:
                                self.stderr.write('        Uitvaller %s heeft gemiddelde %s' % (lid_nr, gemiddelde))
                        else:
                            # mag wel invaller en vervangt de hoogste uitvaller
                            uitvallers.pop(0)
                    # while

                    feitelijke_deelnemers = [deelnemer for deelnemer in feitelijke_deelnemers if deelnemer.sporterboog.sporter.lid_nr not in afgekeurd]

            if len(feitelijke_deelnemers) > 0:
                kamp_team.result_teamscore = 0
                deelnemer_totalen = list()

                for deelnemer in feitelijke_deelnemers:
                    if not dryrun:
                        # uitgestelde save actie
                        deelnemer.save(update_fields=['result_teamscore_1', 'result_teamscore_2'])
                    deelnemer_totaal = deelnemer.result_teamscore_1 + deelnemer.result_teamscore_2
                    deelnemer_totalen.append(deelnemer_totaal)
                # for
                deelnemer_totalen.sort(reverse=True)                        # hoogste eerst
                kamp_team.result_teamscore = sum(deelnemer_totalen[:3])     # de 3 hoogste gebruiken
                kamp_team.feitelijke_leden.set(feitelijke_deelnemers)

                deelnemer_totalen.insert(0, kamp_team.result_teamscore)     # resultaat vooraan
                deelnemer_totalen.append(kamp_team)                         # team record laatst
                kamp_teams.append(deelnemer_totalen)
        # while

        kamp_teams.sort(reverse=True)               # hoogste eerste
        rank = 0
        for tup in kamp_teams:
            rank += 1
            kamp_team = tup[-1]
            kamp_team.result_rank = rank
            if not dryrun:
                kamp_team.save(update_fields=['result_rank', 'result_teamscore'])
        # for

# end of file
