# -*- coding: utf-8 -*-

#  Copyright (c) 2022-2025 Ramon van der Winkel.
#  All rights reserved.
#  Licensed under BSD-3-Clause-Clear. See LICENSE file for details.

from django.core.management.base import BaseCommand
from Competitie.definities import DEEL_RK
from Competitie.models import KampioenschapSporterBoog, KampioenschapTeam
from openpyxl.utils.exceptions import InvalidFileException
from decimal import Decimal
import openpyxl
import zipfile


class Command(BaseCommand):
    help = "Importeer uitslag team kampioenschap 25m1pijl"

    def __init__(self, stdout=None, stderr=None, no_color=False, force_color=False):
        super().__init__(stdout, stderr, no_color, force_color)
        self.dryrun = True
        self.verbose = False
        self.deelnemers = dict()        # [lid_nr] = [KampioenschapSporterBoog, ...]
        self.teams_cache = list()       # [KampioenschapTeam, ...]
        self.team_lid_nrs = dict()      # [team.pk] = [lid_nr, ...]
        self.ver_lid_nrs = dict()       # [ver_nr] = [lid_nr, ...]
        self.kamp_lid_nrs = list()      # [lid_nr, ...]     iedereen die geplaatst is voor de kampioenschappen

    def add_arguments(self, parser):
        parser.add_argument('--dryrun', action='store_true')
        parser.add_argument('--verbose', action='store_true')
        parser.add_argument('bestand', type=str,
                            help='Pad naar het Excel bestand')

    def _deelnemers_ophalen(self):
        for deelnemer in (KampioenschapSporterBoog
                          .objects
                          .filter(kampioenschap__competitie__afstand='25',
                                  kampioenschap__deel=DEEL_RK)
                          .select_related('kampioenschap',
                                          'kampioenschap__rayon',
                                          'sporterboog__sporter',
                                          'sporterboog__boogtype',
                                          'indiv_klasse')):

            lid_nr = deelnemer.sporterboog.sporter.lid_nr
            ver_nr = deelnemer.bij_vereniging.ver_nr

            try:
                self.deelnemers[lid_nr].append(deelnemer)
            except KeyError:
                self.deelnemers[lid_nr] = [deelnemer]

            try:
                self.ver_lid_nrs[ver_nr].append(lid_nr)
            except KeyError:
                self.ver_lid_nrs[ver_nr] = [lid_nr]

            self.kamp_lid_nrs.append(lid_nr)
        # for

    def _filter_deelnemers(self, team_klasse):
        # reduceer het aantal KampioenschapSporterBoog aan de hand van de bogen toegestaan in deze wedstrijdklasse
        afkortingen = list(team_klasse.boog_typen.values_list('afkorting', flat=True))
        self.stdout.write('[INFO] Toegestane bogen in team klasse %s = %s' % (team_klasse, ",".join(afkortingen)))

        # count1 = sum([len(deelnemers) for deelnemers in self.deelnemers.values()])

        for lid_nr in self.deelnemers.keys():
            deelnemers = self.deelnemers[lid_nr]
            self.deelnemers[lid_nr] = [deelnemer
                                       for deelnemer in deelnemers
                                       if deelnemer.sporterboog.boogtype.afkorting in afkortingen]
        # for

        # count2 = sum([len(deelnemers) for deelnemers in self.deelnemers.values()])
        # self.stdout.write('[INFO] Aantal KampioenschapSporterBoog verwijderd: %s' % (count1 - count2))

    def _teams_ophalen(self):
        for team in (KampioenschapTeam
                     .objects
                     .filter(kampioenschap__competitie__afstand='25',
                             kampioenschap__deel=DEEL_RK)
                     .select_related('kampioenschap',
                                     'kampioenschap__rayon',
                                     'vereniging',
                                     'team_type',
                                     'team_klasse')
                     .prefetch_related('gekoppelde_leden',
                                       'feitelijke_leden')):

            self.teams_cache.append(team)
            self.team_lid_nrs[team.pk] = [deelnemer.sporterboog.sporter.lid_nr
                                          for deelnemer in team.gekoppelde_leden.all()]
        # for

    def _sort_op_gemiddelde(self, lid_nrs):
        gem = list()
        for lid_nr in lid_nrs:
            deelnemer_all = self.deelnemers[lid_nr]
            if len(deelnemer_all) == 1:
                deelnemer = deelnemer_all[0]
            else:
                self.stderr.write('[WARNING] TODO: bepaal juiste deelnemer uit %s' % repr(deelnemer_all))
                deelnemer = deelnemer_all[0]
            tup = (deelnemer.gemiddelde, lid_nr)
            gem.append(tup)
        # for
        gem.sort(reverse=True)
        return gem

    def _get_deelnemer(self, lid_nr, lid_ag):
        deelnemer_all = self.deelnemers[lid_nr]
        for deelnemer in deelnemer_all:
            if abs(deelnemer.gemiddelde - lid_ag) < 0.0001:
                return deelnemer
        # for

        self.stderr.write('[WARNING] TODO: bepaal juiste deelnemer met ag=%s uit\n%s' % (
                            lid_ag,
                            "\n".join(["%s / %s / %s" % (deelnemer,
                                                         deelnemer.sporterboog.boogtype.afkorting,
                                                         deelnemer.gemiddelde)
                                       for deelnemer in deelnemer_all])))
        deelnemer = deelnemer_all[0]
        return deelnemer

    def _get_team(self, team_naam, ver_nr, row_nr, team_klasse):
        up_naam = team_naam.upper()
        sel_teams = list()
        for team in self.teams_cache:
            if team.vereniging.ver_nr == ver_nr:
                if team.team_naam.upper() == up_naam:
                    if team_klasse is None or team.team_klasse == team_klasse:
                        sel_teams.append(team)
        # for

        kamp_team = None
        if len(sel_teams) == 1:
            kamp_team = sel_teams[0]
        elif len(sel_teams) > 1:
            self.stderr.write('[ERROR] Kan team %s van vereniging %s op regel %s niet kiezen uit\n%s' % (
                repr(team_naam), ver_nr, row_nr, "\n".join([str(team) for team in sel_teams])))

        if kamp_team is None:
            self.stderr.write('[ERROR] Kan team %s van vereniging %s op regel %s niet vinden' % (
                repr(team_naam), ver_nr, row_nr))

        return kamp_team

    def handle(self, *args, **options):

        self.dryrun = options['dryrun']
        self.verbose = options['verbose']

        # open de kopie, zodat we die aan kunnen passen
        fname = options['bestand']
        self.stdout.write('[INFO] Lees bestand %s' % repr(fname))
        try:
            prg = openpyxl.load_workbook(fname,
                                         data_only=True)        # do not evaluate formulas; use last calculated values
        except (OSError, zipfile.BadZipFile, KeyError, InvalidFileException) as exc:
            self.stderr.write('[ERROR] Kan het excel bestand niet openen (%s)' % str(exc))
            return

        blad = "Deelnemers en Scores"
        try:
            ws = prg[blad]
        except KeyError:            # pragma: no cover
            self.stderr.write('[ERROR] Kan blad %s niet vinden' % repr(blad))
            return

        col_ver_naam = 'D'
        col_team_naam = 'F'
        col_lid_nr = 'E'
        col_lid_ag = 'G'
        col_score1 = 'H'
        col_score2 = 'I'

        self._deelnemers_ophalen()
        self._teams_ophalen()

        # doorloop alle regels van het excel blad en ga op zoek naar bondsnummers
        row_nr = 9 - 1
        nix_count = 0
        team_klasse = None
        kamp_teams = list()
        while nix_count < 10:
            row_nr += 1
            row = str(row_nr)

            # vind een vereniging + team naam
            ver_naam = ws[col_ver_naam + row].value
            team_naam = ws[col_team_naam + row].value

            if ver_naam is None:
                nix_count += 1
                continue

            nix_count = 0
            ver_nr = -1

            try:
                if ver_naam[0] == '[' and ver_naam[5:5+2] == '] ':
                    ver_nr = int(ver_naam[1:1+4])
            except ValueError:
                pass

            # self.stdout.write('[DEBUG] regel %s: ver_nr=%s, ver_naam=%s, team_naam=%s' % (
            #                       row, ver_nr, repr(ver_naam), repr(team_naam)))

            if ver_nr < 0:
                continue

            # zoek het team erbij
            kamp_team = self._get_team(team_naam, ver_nr, row_nr, team_klasse)
            if kamp_team is None:
                self.stdout.write('[WARNING] Team %s op regel %s niet herkend voor klasse %s' % (
                                    repr(team_naam), row, team_klasse))
                continue

            if team_klasse is None:
                team_klasse = kamp_team.team_klasse
                self._filter_deelnemers(team_klasse)

            ver_lid_nrs = self.ver_lid_nrs[ver_nr]
            team_lid_nrs = self.team_lid_nrs[kamp_team.pk]
            aantal_gekoppeld = len(team_lid_nrs)
            feitelijke_deelnemers = list()

            # haal de teamleden op
            gevonden_lid_nrs = list()
            for lp in range(10):            # pragma: no branch
                row_nr += 1
                row = str(row_nr)

                lid_nr = ws[col_lid_nr + row].value
                if lid_nr is None:
                    break
                if str(lid_nr) == '-':
                    continue

                if lid_nr not in self.kamp_lid_nrs:
                    self.stderr.write('[ERROR] Lid %s is niet gekwalificeerd voor dit kampioenschap!' % lid_nr)
                elif lid_nr not in ver_lid_nrs:
                    self.stderr.write('[ERROR] Lid %s is niet van vereniging %s!' % (lid_nr, ver_nr))
                else:
                    try:
                        lid_ag = round(Decimal(ws[col_lid_ag + row].value), 3)  # 3 cijfers achter de komma
                        score1 = int(ws[col_score1 + row].value)
                        score2 = int(ws[col_score2 + row].value)
                    except (ValueError, TypeError):
                        self.stderr.write('[ERROR] Probleem met de scores op regel %s' % row_nr)
                    else:
                        if score1 + score2 == 0:
                            # sporter heeft niet meegedaan
                            self.stdout.write('[WARNING] Geen scores voor sporter %s op regel %s' % (lid_nr, row_nr))
                        else:
                            deelnemer = self._get_deelnemer(lid_nr, lid_ag)
                            deelnemer.result_rk_teamscore_1 = score1
                            deelnemer.result_rk_teamscore_2 = score2
                            feitelijke_deelnemers.append(deelnemer)
                            gevonden_lid_nrs.append(lid_nr)
            # for

            # haal de verwachte lid_nrs eruit
            feitelijke_lid_nrs = [lid_nr for lid_nr in gevonden_lid_nrs if lid_nr in team_lid_nrs]
            for lid_nr in feitelijke_lid_nrs:
                gevonden_lid_nrs.remove(lid_nr)
                team_lid_nrs.remove(lid_nr)
            # for

            if len(gevonden_lid_nrs):
                uitvallers = self._sort_op_gemiddelde(team_lid_nrs)
                invallers = self._sort_op_gemiddelde(gevonden_lid_nrs)
                afgekeurd = list()

                if len(invallers) > len(uitvallers):
                    self.stderr.write('[ERROR] Te veel invallers voor team %s met max %s sporters (vereniging %s)' % (
                                        repr(team_naam), aantal_gekoppeld, ver_nr))
                    feitelijke_deelnemers = list()
                else:
                    while len(invallers) > 0:
                        gemiddelde_in, lid_nr_in = invallers.pop(0)
                        gemiddelde_uit, lid_nr_uit = uitvallers[0]
                        if gemiddelde_in > gemiddelde_uit:
                            # mag niet invallen
                            afgekeurd.append(lid_nr_in)
                            self.stderr.write(
                                '[ERROR] Te hoog gemiddelde %s voor invaller %s voor team %s van vereniging %s' % (
                                        gemiddelde_in, lid_nr_in, team_naam, ver_nr))
                            for gemiddelde, lid_nr in uitvallers:
                                self.stderr.write('        Uitvaller %s heeft gemiddelde %s' % (lid_nr, gemiddelde))
                        else:
                            # mag wel invaller en vervangt de hoogste uitvaller
                            uitvallers.pop(0)
                    # while

                    feitelijke_deelnemers = [deelnemer
                                             for deelnemer in feitelijke_deelnemers
                                             if deelnemer.sporterboog.sporter.lid_nr not in afgekeurd]

            if len(feitelijke_deelnemers) > 0:
                kamp_team.result_teamscore = 0
                deelnemer_totalen = list()

                for deelnemer in feitelijke_deelnemers:
                    if not self.dryrun:
                        # uitgestelde save actie
                        deelnemer.save(update_fields=['result_rk_teamscore_1', 'result_rk_teamscore_2'])
                    deelnemer_totaal = deelnemer.result_rk_teamscore_1 + deelnemer.result_rk_teamscore_2
                    deelnemer_totalen.append(deelnemer_totaal)
                # for
                deelnemer_totalen.sort(reverse=True)                        # hoogste eerst
                kamp_team.result_teamscore = sum(deelnemer_totalen[:3])     # de 3 hoogste gebruiken
                kamp_team.feitelijke_leden.set(feitelijke_deelnemers)

                deelnemer_totalen.insert(0, kamp_team.result_teamscore)     # resultaat vooraan
                deelnemer_totalen.append(kamp_team)                         # team record laatst
                kamp_teams.append(deelnemer_totalen)
        # while

        kamp_teams.sort(reverse=True)               # hoogste eerste
        rank = 0
        for tup in kamp_teams:
            kamp_team = tup[-1]
            if kamp_team.result_teamscore > 0:
                rank += 1
                kamp_team.result_rank = rank
                kamp_team.result_volgorde = rank
            else:
                kamp_team.result_rank = 0
                kamp_team.result_volgorde = 0
            if self.verbose:
                self.stdout.write("%s (%s) %s" % (kamp_team.result_rank, kamp_team.result_teamscore, kamp_team))
            if not self.dryrun:
                kamp_team.save(update_fields=['result_rank', 'result_teamscore'])
        # for

# end of file
