# -*- coding: utf-8 -*-

#  Copyright (c) 2022 Ramon van der Winkel.
#  All rights reserved.
#  Licensed under BSD-3-Clause-Clear. See LICENSE file for details.

# bereken de team sterktes opnieuw
# door een foutje was deze berekend over de 3 laagste gemiddelden in plaats van de 3 hoogste

from django.core.management.base import BaseCommand
from Competitie.models import KampioenschapSchutterBoog, DEELNAME_NEE
from openpyxl.utils.exceptions import InvalidFileException
import openpyxl
import zipfile


class Command(BaseCommand):
    help = "Importeer uitslag kampioenschap"

    def __init__(self, stdout=None, stderr=None, no_color=False, force_color=False):
        super().__init__(stdout, stderr, no_color, force_color)
        self.deelnemers = dict()        # [lid_nr] = [KampioenschapSchutterBoog, ...]

    def add_arguments(self, parser):
        parser.add_argument('--dryrun', action='store_true')
        parser.add_argument('afstand', type=str,
                            help='Competitie afstand (18/25)')
        parser.add_argument('bestand', type=str,
                            help='Pad naar het Excel bestand')
        parser.add_argument('blad', type=str,
                            help='Naam van het blad met resultaten')
        parser.add_argument('kolommen', type=str, nargs='+',
                            help='Kolom letters: bondsnummer, score1, score2, #10, #9, #8')

    def deelnemers_ophalen(self, afstand):
        for deelnemer in (KampioenschapSchutterBoog
                          .objects
                          .filter(deelcompetitie__competitie__afstand=afstand)
                          .select_related('deelcompetitie',
                                          'deelcompetitie__nhb_rayon',
                                          'sporterboog__sporter',
                                          'sporterboog__boogtype',
                                          'indiv_klasse')):

            lid_nr = deelnemer.sporterboog.sporter.lid_nr
            try:
                self.deelnemers[lid_nr].append(deelnemer)
            except KeyError:
                self.deelnemers[lid_nr] = [deelnemer]
        # for

    def handle(self, *args, **options):

        dryrun = options['dryrun']

        afstand = options['afstand']
        if afstand not in ('18', '25'):
            self.stderr.write('[ERROR] Afstand moet 18 of 25 zijn')
            return

        # open de kopie, zodat we die aan kunnen passen
        fname = options['bestand']
        self.stdout.write('[INFO] Lees bestand %s' % repr(fname))
        try:
            prg = openpyxl.load_workbook(fname,
                                         data_only=True)        # do not evaluate formulas; use last calculated values
        except (OSError, zipfile.BadZipFile, KeyError, InvalidFileException) as exc:
            self.stderr.write('[ERROR] Kan het excel bestand niet openen (%s)' % str(exc))
            return

        blad = options['blad']
        try:
            ws = prg[blad]
        except KeyError:
            self.stderr.write('[ERROR] Kan blad %s niet vinden' % repr(blad))
            return

        cols = options['kolommen']
        if afstand == '25':
            if len(cols) != 6:
                self.stderr.write('[ERROR] Vereiste kolommen: lid_nr, score1, score2, tienen, negens, achten')
                return

            col_lid_nr = cols[0]
            col_score1 = cols[1]
            col_score2 = cols[2]
            col_10s = cols[3]
            col_9s = cols[4]
            col_8s = cols[5]
        else:
            # indoor nog niet ondersteund
            self.stderr.write('[ERROR] Indoor nog niet ondersteund')
            return

        self.deelnemers_ophalen(afstand)

        klasse_pks = list()
        deelcomp_pks = list()

        # doorloop alle regels van het excel blad en ga op zoek naar bondsnummers
        row_nr = 0
        nix_count = 0
        klasse_pk = -1
        rank = 0
        prev_totaal = 999
        prev_counts_str = ""
        while nix_count < 10:
            row_nr += 1
            row = str(row_nr)

            lid_nr_str = ws[col_lid_nr + row].value
            if lid_nr_str:
                nix_count = 0
                try:
                    lid_nr = int(lid_nr_str)
                except ValueError:
                    pass
                else:
                    try:
                        deelnemers = self.deelnemers[lid_nr]
                    except KeyError:
                        self.stderr.write('[ERROR] Geen RK deelnemer op regel %s: %s' % (row, lid_nr))
                    else:
                        if len(deelnemers) == 1:
                            deelnemer = deelnemers[0]
                        else:
                            deelnemer = None
                            if klasse_pk:                                               # pragma: no branch
                                for kandidaat in deelnemers:
                                    if kandidaat.indiv_klasse.pk == klasse_pk:
                                        deelnemer = kandidaat
                                # for
                            if deelnemer is None:
                                # probeer het nog een keer, maar kijk nu ook naar afgemeld status
                                for kandidaat in deelnemers:
                                    if kandidaat.deelname != DEELNAME_NEE:
                                        deelnemer = kandidaat
                                # for
                        if deelnemer is None:
                            self.stderr.write('[ERROR] Kan deelnemer niet bepalen voor regel %s. Keuze uit %s' % (row, repr(deelnemers)))
                            continue    # met de while

                        dupe_check = False
                        if deelnemer.result_rank > 0:
                            dupe_check = True

                        if klasse_pk != deelnemer.indiv_klasse.pk:
                            self.stdout.write('[INFO] Klasse: %s' % deelnemer.indiv_klasse)
                            klasse_pk = deelnemer.indiv_klasse.pk
                            rank = 0
                            prev_totaal = 999
                            if klasse_pk not in klasse_pks:                             # pragma: no branch
                                klasse_pks.append(klasse_pk)
                                deelcomp_pk = deelnemer.deelcompetitie.pk
                                if deelcomp_pk not in deelcomp_pks:                     # pragma: no branch
                                    deelcomp_pks.append(deelnemer.deelcompetitie.pk)

                        score1 = ws[col_score1 + row].value
                        score2 = ws[col_score2 + row].value
                        c10 = ws[col_10s + row].value
                        c9 = ws[col_9s + row].value
                        c8 = ws[col_8s + row].value

                        try:
                            score1 = int(score1)
                            score2 = int(score2)
                        except (TypeError, ValueError):
                            # if deelnemer.deelname != DEELNAME_NEE:      # afgemeld?
                            if score1 is None and score2 is None:
                                self.stdout.write('[WARNING] Regel %s wordt overgeslagen (geen scores)' % row)
                            else:
                                self.stderr.write('[ERROR] Probleem met scores op regel %s: %s en %s' % (row, repr(score1), repr(score2)))
                        else:
                            counts = list()
                            try:
                                if c10:
                                    c10 = int(c10)
                                    counts.append('%sx10' % c10)
                                if c9:
                                    c9 = int(c9)
                                    counts.append('%sx9' % c9)
                                if c8:
                                    c8 = int(c8)
                                    counts.append('%sx8' % c8)
                            except (TypeError, ValueError) as err:
                                self.stderr.write('[ERROR] Probleem met 10/9/8 count op regel %s: %s' % (row, str(err)))
                                counts_str = ''
                            else:
                                counts_str = ", ".join(counts)

                            totaal = score1 + score2
                            if totaal > 0:                  # soms wordt 0,0 ingevuld bij niet aanwezig
                                if totaal == prev_totaal and counts_str == prev_counts_str:
                                    # zelfde score, zelf rank
                                    pass
                                elif totaal > prev_totaal:
                                    self.stderr.write('[ERROR] Score is niet aflopend op regel %s' % row)
                                else:
                                    rank += 1
                                prev_totaal = totaal
                                prev_counts_str = counts_str

                                self.stdout.write('%s: %s, scores: %s %s %s' % (rank, deelnemer, score1, score2, counts_str))

                                do_report = False
                                if dupe_check:
                                    # allow result_rank change
                                    is_dupe = (deelnemer.result_score_1 == score1
                                               and deelnemer.result_score_2 == score2
                                               and deelnemer.result_counts == counts_str)
                                    if not is_dupe:
                                        do_report = True

                                if do_report:
                                    self.stderr.write('[ERROR] Deelnemer pk=%s heeft al andere resultaten! (%s)' % (deelnemer.pk, deelnemer))
                                else:
                                    deelnemer.result_rank = rank
                                    deelnemer.result_score_1 = score1
                                    deelnemer.result_score_2 = score2
                                    deelnemer.result_counts = counts_str

                        if not dryrun:
                            deelnemer.save(update_fields=['result_rank',
                                                          'result_score_1', 'result_score_2',
                                                          'result_counts'])
            else:
                nix_count += 1
        # while

        # zet deelnemers die niet meegedaan hebben op een hoog rank nummer
        for deelnemers in self.deelnemers.values():
            for deelnemer in deelnemers:
                if deelnemer.deelcompetitie.pk in deelcomp_pks:
                    if deelnemer.indiv_klasse.pk in klasse_pks:
                        if deelnemer.result_rank in (0, 32000, 32001):
                            if deelnemer.deelname != DEELNAME_NEE:
                                self.stdout.write('[WARNING] No-show voor deelnemer %s' % deelnemer)
                                deelnemer.result_rank = 32000
                            else:
                                deelnemer.result_rank = 32001
                            deelnemer.save(update_fields=['result_rank'])
            # for
        # for

# end of file
