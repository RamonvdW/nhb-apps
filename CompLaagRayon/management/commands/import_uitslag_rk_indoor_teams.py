# -*- coding: utf-8 -*-

#  Copyright (c) 2022-2026 Ramon van der Winkel.
#  All rights reserved.
#  Licensed under BSD-3-Clause-Clear. See LICENSE file for details.

from django.core.management.base import BaseCommand
from Competitie.definities import DEEL_RK, KAMP_RANK_NO_SHOW
from Competitie.models import KampioenschapSporterBoog, KampioenschapTeam
from openpyxl.utils.exceptions import InvalidFileException
from decimal import Decimal, InvalidOperation
import openpyxl
import zipfile


class Command(BaseCommand):
    help = "Importeer uitslag RK Indoor Teams"

    def __init__(self, stdout=None, stderr=None, no_color=False, force_color=False):
        super().__init__(stdout, stderr, no_color, force_color)
        self.dryrun = True
        self.verbose = False
        self.has_error = False
        self.rayon_nr = 0
        self.team_klasse = None

        self.deelnemers = dict()            # [lid_nr] = [KampioenschapSporterBoog, ...]
        self.teams_cache = list()           # [KampioenschapTeam, ...]
        self.team_lid_nrs = dict()          # [team.pk] = [lid_nr, ...]
        self.ver_lid_nrs = dict()           # [ver_nr] = [lid_nr, ...]
        self.kamp_lid_nrs = list()          # [lid_nr, ...]     iedereen die geplaatst is voor de kampioenschappen
        self.deelnemende_teams = dict()     # [team naam] = KampioenschapTeam
        self.toegestane_bogen = list()

    def add_arguments(self, parser):
        parser.add_argument('--dryrun', action='store_true')
        parser.add_argument('--verbose', action='store_true')
        parser.add_argument('bestand', type=str,
                            help='Pad naar het Excel bestand')

    def _zet_team_klasse(self, klasse):
        self.team_klasse = klasse
        self.toegestane_bogen = list(klasse.boog_typen.values_list('afkorting', flat=True))
        if self.verbose:
            self.stdout.write('[DEBUG] toegestane bogen: %s' % repr(self.toegestane_bogen))

    def _deelnemers_ophalen(self):
        for deelnemer in (KampioenschapSporterBoog
                          .objects
                          .filter(kampioenschap__competitie__afstand='18',
                                  kampioenschap__deel=DEEL_RK)
                          .select_related('kampioenschap',
                                          'kampioenschap__rayon',
                                          'sporterboog__sporter',
                                          'sporterboog__boogtype',
                                          'indiv_klasse')):

            lid_nr = deelnemer.sporterboog.sporter.lid_nr
            ver_nr = deelnemer.bij_vereniging.ver_nr

            try:
                self.deelnemers[lid_nr].append(deelnemer)
            except KeyError:
                self.deelnemers[lid_nr] = [deelnemer]

            try:
                self.ver_lid_nrs[ver_nr].append(lid_nr)
            except KeyError:
                self.ver_lid_nrs[ver_nr] = [lid_nr]

            self.kamp_lid_nrs.append(lid_nr)
        # for

    def _teams_ophalen(self):
        for team in (KampioenschapTeam
                     .objects
                     .filter(kampioenschap__competitie__afstand='18',
                             kampioenschap__deel=DEEL_RK)
                     .select_related('kampioenschap',
                                     'kampioenschap__rayon',
                                     'vereniging',
                                     'team_type',
                                     'team_klasse')
                     .prefetch_related('gekoppelde_leden',
                                       'feitelijke_leden')):

            self.teams_cache.append(team)
            self.team_lid_nrs[team.pk] = [deelnemer.sporterboog.sporter.lid_nr
                                          for deelnemer in team.gekoppelde_leden.all()]
        # for

    def _sort_op_gemiddelde(self, lid_nrs):
        gem = list()
        for lid_nr in lid_nrs:
            deelnemer_all = [deelnemer
                             for deelnemer in self.deelnemers[lid_nr]
                             if deelnemer.sporterboog.boogtype.afkorting in self.toegestane_bogen]

            if len(deelnemer_all) == 1:
                deelnemer = deelnemer_all[0]
            else:
                if len(deelnemer_all) == 0:
                    self.stderr.write('[ERROR] Geen deelnemers gevonden voor lid %s' % lid_nr)
                    self.has_error = True
                    continue

                self.stderr.write('[WARNING] Neem de eerste want kan niet kiezen uit: %s' % repr(deelnemer_all))
                deelnemer = deelnemer_all[0]

            tup = (deelnemer.gemiddelde, lid_nr)
            gem.append(tup)
        # for
        gem.sort(reverse=True)
        return gem

    def _get_deelnemer(self, lid_nr, lid_ag):
        deelnemer_all = self.deelnemers[lid_nr]
        for deelnemer in deelnemer_all:
            if abs(deelnemer.gemiddelde - lid_ag) < 0.0001:
                return deelnemer
        # for

        self.stderr.write('[WARNING] TODO: bepaal juiste deelnemer met ag=%s uit\n%s' % (
                            lid_ag,
                            "\n".join(["%s / %s / %s" % (deelnemer,
                                                         deelnemer.sporterboog.boogtype.afkorting,
                                                         deelnemer.gemiddelde) for deelnemer in deelnemer_all])))
        deelnemer = deelnemer_all[0]
        return deelnemer

    def _get_team(self, team_naam: str, ver_nr: int, row_nr: int):
        # if self.verbose:
        #     self.stdout.write('[DEBUG] get_team: team_klasse=%s' % repr(self.team_klasse))

        up_naam = team_naam.upper()
        sel_teams = list()
        for team in self.teams_cache:
            # self.stdout.write('[DEBUG] cached team: %s' % repr(team))
            if team.vereniging.ver_nr == ver_nr:
                if team.team_naam.upper() == up_naam:
                    if team.team_klasse == self.team_klasse:
                        sel_teams.append(team)
        # for

        if len(sel_teams) == 0:
            # niet gevonden, dus waarschijnlijk is de naam aangepast
            for team in self.teams_cache:
                # self.stdout.write('[DEBUG] cached team: %s' % repr(team))
                if team.vereniging.ver_nr == ver_nr:
                    team_up_naam = team.team_naam.upper()
                    if team_up_naam in up_naam or up_naam in team_up_naam:
                        if team.team_klasse == self.team_klasse:
                            sel_teams.append(team)
                            self.stdout.write('[WARNING] Aangepaste team naam: %s --> %s' % (
                                                repr(team.team_naam), repr(team_naam)))
            # for

        kamp_team = None
        if len(sel_teams) == 1:
            kamp_team = sel_teams[0]

        elif len(sel_teams) > 1:
            self.stderr.write('[ERROR] Kan team %s van vereniging %s op regel %s niet kiezen uit\n%s' % (
                repr(team_naam), ver_nr, row_nr, "\n".join([str(team) for team in sel_teams])))
            self.has_error = True

        if kamp_team is None:
            self.stderr.write('[ERROR] Kan team %s van vereniging %s op regel %s niet vinden' % (
                repr(team_naam), ver_nr, row_nr))
            self.has_error = True

        return kamp_team

    @staticmethod
    def _lees_team_naam(ws, cell):
        team_naam = ws[cell].value
        if team_naam is None:
            team_naam = ''
        else:
            team_naam = str(team_naam)
            if str(team_naam).upper() in ('N.V.T.', 'NVT', 'BYE', '#N/A', '0'):
                team_naam = ''
        return team_naam

    def _bepaal_klasse_en_rayon(self, ws):
        mogelijke_klassen = list()

        # loop alle teams af
        for row_nr in range(8, 43+1, 5):
            # team naam
            team_naam = self._lees_team_naam(ws, 'B' + str(row_nr))
            if not team_naam:
                continue

            # ver_nr
            try:
                ver_nr = int(ws['C' + str(row_nr)].value)
            except ValueError:
                continue

            # zoek de teams van de vereniging erbij
            up_naam = team_naam.upper()
            ver_teams = list()
            for team in self.teams_cache:
                # self.stdout.write('[DEBUG] cached team: %s' % repr(team))
                if team.vereniging.ver_nr == ver_nr:
                    if team.team_naam.upper() == up_naam:
                        ver_teams.append(team)
            # for

            if len(ver_teams) == 1:
                # duidelijk
                team_klasse = ver_teams[0].team_klasse
                if team_klasse not in mogelijke_klassen:
                    mogelijke_klassen.append(team_klasse)

                self.rayon_nr = ver_teams[0].vereniging.regio.rayon_nr
        # for team row

        if len(mogelijke_klassen) == 0:
            self.stderr.write('[ERROR] Team klasse niet kunnen bepalen (geen teams)')
            self.has_error = True
            return

        if len(mogelijke_klassen) != 1:
            self.stderr.write('[ERROR] Kan team klasse niet kiezen uit: %s' % repr(mogelijke_klassen))
            self.has_error = True
            return

        self._zet_team_klasse(mogelijke_klassen[0])

        self.stdout.write('[INFO] Rayon: %s' % self.rayon_nr)
        self.stdout.write('[INFO] Klasse: %s' % self.team_klasse)

    def _importeer_teams_en_sporters(self, ws):
        for row_nr in range(8, 43+1, 5):

            # team naam
            team_naam = self._lees_team_naam(ws, 'B' + str(row_nr))
            if not team_naam:
                continue

            # ver_nr
            try:
                ver_nr = int(ws['C' + str(row_nr)].value)
            except ValueError:
                continue

            # zoek het team erbij
            kamp_team = self._get_team(team_naam, ver_nr, row_nr)
            if kamp_team is None:
                continue

            self.deelnemende_teams[team_naam] = kamp_team
            self.stdout.write('[INFO] Gevonden team: %s van ver %s' % (repr(team_naam), ver_nr))

            if self.team_klasse != kamp_team.team_klasse:
                self.stderr.write('[ERROR] Inconsistente team klasse op regel %s: %s (eerdere teams: %s)' % (
                                    row_nr, kamp_team.team_klasse, self.team_klasse))
                self.has_error = True
                continue

            # lees de deelnemers van dit team
            ver_lid_nrs = self.ver_lid_nrs[ver_nr]
            team_lid_nrs = self.team_lid_nrs[kamp_team.pk]
            aantal_gekoppeld = len(team_lid_nrs)
            gevonden_lid_nrs = list()
            feitelijke_deelnemers = list()

            for sporter_row in range(row_nr+1, row_nr+4):
                lid_nr_str = ws['C' + str(sporter_row)].value
                if lid_nr_str is None:
                    continue

                lid_ag_str = ws['D' + str(sporter_row)].value
                lid_ag_str = str(lid_ag_str).replace(',', '.')       # decimale komma naar punt

                try:
                    lid_nr = int(lid_nr_str)
                except ValueError:
                    self.stderr.write('[ERROR] Geen valide lid_nr %s op regel %s' % (repr(lid_nr_str), sporter_row))
                    self.has_error = True
                    continue
                else:
                    if lid_nr not in self.kamp_lid_nrs:
                        self.stderr.write('[ERROR] Lid %s is niet gekwalificeerd voor dit kampioenschap!' % lid_nr)
                        self.has_error = True
                        continue

                    if lid_nr not in ver_lid_nrs:
                        self.stderr.write('[ERROR] Lid %s is niet van vereniging %s!' % (lid_nr, ver_nr))
                        self.has_error = True
                        continue

                    try:
                        lid_ag = round(Decimal(lid_ag_str), 3)  # 3 cijfers achter de komma
                    except (TypeError, InvalidOperation):
                        self.stderr.write('[ERROR] Geen valide AG %s op regel %s' % (repr(lid_ag_str), sporter_row))
                        self.has_error = True
                        continue

                    deelnemer = self._get_deelnemer(lid_nr, lid_ag)
                    self.stdout.write('[INFO] team lid: %s' % deelnemer.sporterboog.sporter.lid_nr_en_volledige_naam())

                    feitelijke_deelnemers.append(deelnemer)
                    gevonden_lid_nrs.append(lid_nr)
            # for sporter

            # haal de verwachte lid_nrs eruit
            feitelijke_lid_nrs = [lid_nr
                                  for lid_nr in gevonden_lid_nrs
                                  if lid_nr in team_lid_nrs]
            if self.verbose:
                self.stdout.write('[DEBUG] feitelijke_lid_nrs van team [%s] %s: %s' % (
                                        ver_nr, team_naam, repr(feitelijke_lid_nrs)))

            for lid_nr in feitelijke_lid_nrs:
                gevonden_lid_nrs.remove(lid_nr)
                team_lid_nrs.remove(lid_nr)
            # for

            if len(gevonden_lid_nrs):
                uitvallers = self._sort_op_gemiddelde(team_lid_nrs)
                invallers = self._sort_op_gemiddelde(gevonden_lid_nrs)
                afgekeurd = list()

                if len(invallers) > len(uitvallers):
                    self.stderr.write('[ERROR] Te veel invallers voor team %s met max %s sporters (ver: %s)' % (
                                        repr(team_naam), aantal_gekoppeld, ver_nr))
                    feitelijke_deelnemers = list()
                else:
                    while len(invallers) > 0:
                        gemiddelde_in, lid_nr_in = invallers.pop(0)
                        gemiddelde_uit, lid_nr_uit = uitvallers[0]
                        if gemiddelde_in > gemiddelde_uit:
                            # mag niet invallen
                            afgekeurd.append(lid_nr_in)
                            self.stderr.write(
                                '[ERROR] Te hoog gemiddelde %s voor invaller %s voor team %s van vereniging %s' % (
                                        gemiddelde_in, lid_nr_in, team_naam, ver_nr))
                            for gemiddelde, lid_nr in uitvallers:
                                self.stderr.write('        Uitvaller %s heeft gemiddelde %s' % (lid_nr, gemiddelde))
                            self.has_error = True
                        else:
                            # mag wel invaller en vervangt de hoogste uitvaller
                            uitvallers.pop(0)
                    # while

                    feitelijke_deelnemers = [deelnemer
                                             for deelnemer in feitelijke_deelnemers
                                             if deelnemer.sporterboog.sporter.lid_nr not in afgekeurd]

            if len(feitelijke_deelnemers) != 3:
                self.stderr.write('[ERROR] Maar %s deelnemers in team %s' % (
                                        len(feitelijke_deelnemers), repr(kamp_team.team_naam)))
                self.has_error = True
                return

            if not self.dryrun:
                if len(feitelijke_deelnemers) > 0:
                    kamp_team.feitelijke_leden.set(feitelijke_deelnemers)

        # for team row

    def _importeer_eindstand(self, ws):

        if len(self.deelnemende_teams) == 0:
            self.stderr.write('[ERROR] Geen deelnemende teams gevonden!')
            self.has_error = True
            return

        first_team = next(iter(self.deelnemende_teams.values()))
        deelkamp = first_team.kampioenschap
        self.stdout.write('[INFO] Rayon: %s' % deelkamp.rayon.rayon_nr)

        # begin met alle teams in deze klasse op "no-show" te zetten
        # let op: dit kunnen ook reserve teams zijn (>8 deelnemers)
        expected_teams = list()
        for kamp_team in self.teams_cache:
            if kamp_team.kampioenschap == deelkamp and kamp_team.team_klasse == self.team_klasse:
                if self.verbose:
                    self.stdout.write('[DEBUG] Verwacht team in deze klasse: %s' % kamp_team)
                kamp_team.result_rank = KAMP_RANK_NO_SHOW
                kamp_team.result_volgorde = 50 + kamp_team.volgorde
                kamp_team.result_teamscore = 0
                expected_teams.append(kamp_team)
        # for

        rank = 0

        for row_nr in range(8, 15+1):
            # team naam
            team_naam = self._lees_team_naam(ws, 'B' + str(row_nr))
            if not team_naam:
                continue

            kamp_team = self.deelnemende_teams[team_naam]

            matchpunten_str = ws['D' + str(row_nr)].value
            try:
                matchpunten = int(matchpunten_str)
            except ValueError:
                self.stderr.write('[ERROR] Geen valide matchpunten %s op regel %s' % (repr(matchpunten_str), row_nr))
                self.has_error = True
                continue

            rank += 1
            self.stdout.write('[INFO] Rank %s: %s punten, team %s' % (rank,
                                                                      matchpunten,
                                                                      repr(team_naam)))

            # 100=blanco, 32000=no show, 32001=reserve
            kamp_team.result_rank = rank
            kamp_team.result_volgorde = rank
            kamp_team.result_teamscore = matchpunten
            expected_teams.remove(kamp_team)

            if not (self.dryrun or self.has_error):
                kamp_team.save(update_fields=['result_rank', 'result_volgorde', 'result_teamscore'])
        # for

        # rapporteer de no-shows
        for kamp_team in expected_teams:
            self.stderr.write('[WARNING] Team %s van ver %s staat niet in de uitslag --> no-show' % (
                                repr(kamp_team.team_naam), kamp_team.vereniging.ver_nr))
            if not (self.dryrun or self.has_error):
                kamp_team.save(update_fields=['result_rank', 'result_volgorde', 'result_teamscore'])
        # for

    def handle(self, *args, **options):

        self.dryrun = options['dryrun']
        self.verbose = options['verbose']

        fname = options['bestand']
        self.stdout.write('[INFO] Lees bestand %s' % repr(fname))
        try:
            prg = openpyxl.load_workbook(fname,
                                         read_only=True,        # avoids warnings
                                         data_only=True)        # do not evaluate formulas; use last calculated values
        except (OSError, zipfile.BadZipFile, KeyError, InvalidFileException) as exc:
            self.stderr.write('[ERROR] Kan het excel bestand niet openen (%s)' % str(exc))
            return

        blad = 'Deelnemers'
        try:
            ws_deelnemers = prg[blad]
        except KeyError:        # pragma: no cover
            self.stderr.write('[ERROR] Kan blad %s niet vinden' % repr(blad))
            return

        blad = 'Stand'
        try:
            ws_stand = prg[blad]
        except KeyError:        # pragma: no cover
            self.stderr.write('[ERROR] Kan blad %s niet vinden' % repr(blad))
            return

        self._deelnemers_ophalen()
        self._teams_ophalen()

        self._bepaal_klasse_en_rayon(ws_deelnemers)

        if not self.has_error:
            self._importeer_teams_en_sporters(ws_deelnemers)

        if not self.has_error:
            self._importeer_eindstand(ws_stand)

# end of file
