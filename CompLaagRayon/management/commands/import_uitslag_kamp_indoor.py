# -*- coding: utf-8 -*-

#  Copyright (c) 2022-2023 Ramon van der Winkel.
#  All rights reserved.
#  Licensed under BSD-3-Clause-Clear. See LICENSE file for details.

from django.core.management.base import BaseCommand
from Competitie.models import (Competitie, DEEL_RK, DEEL_BK,
                               KampioenschapSporterBoog, DEELNAME_NEE, DEELNAME_JA,
                               KAMP_RANK_UNKNOWN, KAMP_RANK_NO_SHOW, KAMP_RANK_RESERVE)
from openpyxl.utils.exceptions import InvalidFileException
import openpyxl
import zipfile
import sys


class Command(BaseCommand):
    help = "Importeer uitslag Indoor kampioenschap"

    blad_voorronde = 'Voorronde'
    blad_finales = 'Finale'

    def __init__(self, stdout=None, stderr=None, no_color=False, force_color=False):
        super().__init__(stdout, stderr, no_color, force_color)
        self.deel = DEEL_RK
        self.deelnemers = dict()        # [lid_nr] = [KampioenschapSporterBoog, ...]
        self.rk_deelnemers = list()     # [KampioenschapSporterBoog, ...]
        self.dryrun = True
        self.verbose = False

    def add_arguments(self, parser):
        parser.add_argument('--dryrun', action='store_true')
        parser.add_argument('--verbose', action='store_true')
        parser.add_argument('bestand', type=str,
                            help='Pad naar het Excel bestand')

    def _bepaal_deel(self):
        """ bepaal of we de RK of BK gaan importeren """
        comp = Competitie.objects.filter(afstand='18').order_by('begin_jaar')[0]    # pak de oudeste
        if comp.alle_rks_afgesloten:
            self.deel = DEEL_BK

    def _deelnemers_ophalen(self):
        for deelnemer in (KampioenschapSporterBoog
                          .objects
                          .filter(kampioenschap__competitie__afstand='18',
                                  kampioenschap__deel=self.deel)
                          .select_related('kampioenschap',
                                          'kampioenschap__nhb_rayon',
                                          'sporterboog__sporter',
                                          'sporterboog__boogtype',
                                          'indiv_klasse')):

            # print(deelnemer, deelnemer.indiv_klasse)
            lid_nr = deelnemer.sporterboog.sporter.lid_nr
            try:
                self.deelnemers[lid_nr].append(deelnemer)
            except KeyError:
                self.deelnemers[lid_nr] = [deelnemer]
        # for

    def _importeer_resultaten(self, ws):
        col_lid_nr = 'D'
        col_score1 = 'J'
        col_score2 = 'K'
        col_result = 'Q'

        klasse_pk = None
        deelkamp_pk = None

        # doorloop alle regels van het excel blad en ga op zoek naar bondsnummers
        row_nr = 4
        nix_count = 0
        while nix_count < 10:
            row_nr += 1
            row = str(row_nr)

            stop = ws['A' + row].value
            if stop == 'Afgemeld en reserven:':
                break   # from the while

            lid_nr_str = ws[col_lid_nr + row].value
            if lid_nr_str:
                nix_count = 0
                try:
                    lid_nr = int(lid_nr_str)
                except ValueError:
                    pass
                else:
                    try:
                        deelnemers = self.deelnemers[lid_nr]
                    except KeyError:
                        self.stderr.write('[ERROR] Geen RK deelnemer op regel %s: %s' % (row, lid_nr))
                    else:
                        if len(deelnemers) == 1:
                            deelnemer = deelnemers[0]
                        else:
                            deelnemer = None
                            if klasse_pk:
                                for kandidaat in deelnemers:
                                    if kandidaat.indiv_klasse.pk == klasse_pk:
                                        deelnemer = kandidaat
                                # for
                            if deelnemer is None:
                                # probeer het nog een keer, maar kijk nu ook naar afgemeld status
                                for kandidaat in deelnemers:
                                    if kandidaat.deelname != DEELNAME_NEE:
                                        if deelnemer:
                                            deelnemer = None
                                            break
                                        else:
                                            deelnemer = kandidaat
                                # for
                        if deelnemer is None:
                            self.stderr.write('[ERROR] Kan deelnemer niet bepalen voor regel %s. Keuze uit %s' % (
                                                row, repr(deelnemers)))
                            sys.exit(1)

                        if deelnemer.deelname == DEELNAME_NEE:
                            self.stdout.write('[WARNING] Was afgemeld maar wordt deelnemer: %s' % deelnemer)
                            deelnemer.deelname = DEELNAME_JA
                            if not self.dryrun:
                                deelnemer.save(update_fields=['deelname'])

                        dupe_check = False
                        if deelnemer.result_rank > 0:
                            dupe_check = True

                        if klasse_pk != deelnemer.indiv_klasse.pk:
                            if klasse_pk:
                                self.stderr.write('[ERROR] Deelnemer %s hoort in klasse: %s' % (
                                                    deelnemer, deelnemer.indiv_klasse))
                            else:
                                klasse_pk = deelnemer.indiv_klasse.pk
                                deelkamp_pk = deelnemer.kampioenschap.pk
                                self.stdout.write('[INFO] Klasse: %s' % deelnemer.indiv_klasse)

                        score1 = ws[col_score1 + row].value
                        score2 = ws[col_score2 + row].value
                        result = ws[col_result + row].value

                        try:
                            score1 = int(score1)
                            score2 = int(score2)
                        except (TypeError, ValueError):
                            # if deelnemer.deelname != DEELNAME_NEE:      # afgemeld?
                            if score1 is None and score2 is None:
                                self.stdout.write('[WARNING] Regel %s wordt overgeslagen (geen scores)' % row)
                            else:
                                self.stderr.write('[ERROR] Probleem met scores op regel %s: %s en %s' % (
                                                    row, repr(score1), repr(score2)))
                        else:
                            totaal = score1 + score2
                            if totaal > 0:                  # soms wordt 0,0 ingevuld bij niet aanwezig
                                if result and len(str(result)) > 1:
                                    # dit is de rayonkampioen
                                    rank = 1
                                else:
                                    # wel meegedaan, maar met het huidige RK programma weten we niet genoeg
                                    rank = KAMP_RANK_UNKNOWN

                                if self.verbose:
                                    self.stdout.write('[DEBUG] Voorronde: %s: %s, scores: %s %s, result: %s' % (
                                                        rank, deelnemer, score1, score2, result))

                                do_report = False
                                if dupe_check:
                                    # allow result_rank change
                                    is_dupe = (deelnemer.result_score_1 == score1
                                               and deelnemer.result_score_2 == score2)
                                    if not is_dupe:
                                        do_report = True

                                if do_report:
                                    self.stderr.write('[ERROR] Deelnemer pk=%s heeft al andere resultaten! (%s)' % (
                                                        deelnemer.pk, deelnemer))
                                else:
                                    deelnemer.result_rank = rank
                                    deelnemer.result_score_1 = score1
                                    deelnemer.result_score_2 = score2

                        if not self.dryrun:
                            deelnemer.save(update_fields=['result_rank', 'result_score_1', 'result_score_2'])

                        self.deelnemers[lid_nr] = [deelnemer]
            else:
                nix_count += 1
        # while

        # zet deelnemers die niet meegedaan hebben op een hoog rank nummer
        for deelnemers in self.deelnemers.values():
            for deelnemer in deelnemers:
                if deelnemer.kampioenschap.pk == deelkamp_pk and deelnemer.indiv_klasse.pk == klasse_pk:
                    if deelnemer.result_rank in (0, KAMP_RANK_NO_SHOW, KAMP_RANK_RESERVE):
                        if deelnemer.deelname != DEELNAME_NEE:
                            # noteer: we weten niet welke reserves opgeroepen waren
                            # noteer: nog geen oplossing voor reserves die toch niet mee kunnen doen
                            self.stdout.write('[WARNING] Mogelijke no-show voor deelnemer %s' % deelnemer)
                            deelnemer.result_rank = KAMP_RANK_NO_SHOW
                        else:
                            deelnemer.result_rank = KAMP_RANK_RESERVE
                        if not self.dryrun:
                            deelnemer.save(update_fields=['result_rank'])
                    else:
                        self.rk_deelnemers.append(deelnemer)
            # for
        # for

    def _importeer_finales(self, ws):
        """ Zoek uit wie er in de finale zaten """

        final_16 = list()
        final_8 = list()
        final_4 = list()
        final_2 = list()
        goud = 0
        brons = 0

        for row in (5, 7, 9, 11, 13, 15, 17, 19, 24, 26, 28, 30, 32, 34, 36, 38):
            lid_nr_str = ws['B%s' % row].value
            if lid_nr_str:
                try:
                    lid_nr = int(lid_nr_str)
                except ValueError:
                    pass
                else:
                    final_16.append(lid_nr)
        # for

        for row in (6, 10, 14, 18, 25, 29, 33, 37):
            lid_nr_str = ws['K%s' % row].value
            if lid_nr_str:
                try:
                    lid_nr = int(lid_nr_str)
                except ValueError:
                    pass
                else:
                    final_8.append(lid_nr)
        # for

        for row in (8, 16, 27, 35):
            lid_nr_str = ws['T%s' % row].value
            if lid_nr_str:
                try:
                    lid_nr = int(lid_nr_str)
                except ValueError:
                    pass
                else:
                    final_4.append(lid_nr)
        # for

        for row in (12, 31):
            lid_nr_str = ws['AC%s' % row].value
            if lid_nr_str:
                try:
                    lid_nr = int(lid_nr_str)
                except ValueError:
                    pass
                else:
                    final_2.append(lid_nr)
        # for

        lid_nr_str = ws['AC22'].value
        if lid_nr_str:
            try:
                lid_nr = int(lid_nr_str)
            except ValueError:
                pass
            else:
                goud = lid_nr

        lid_nr_str = ws['T22'].value
        if lid_nr_str:
            try:
                lid_nr = int(lid_nr_str)
            except ValueError:
                pass
            else:
                brons = lid_nr

        if self.verbose:
            self.stdout.write('[DEBUG] Aantal rk deelnemers: %s' % len(self.rk_deelnemers))

        if goud == 0:
            self.stderr.write('[ERROR] Kan winnaar van gouden finale niet vaststellen')
            return

        if brons == 0 and len(self.rk_deelnemers) > 2:
            self.stderr.write('[ERROR] Kan winnaar van bronzen finale niet vaststellen')
            return

        if self.verbose:
            print('final_16: %s' % repr(final_16))
            print('final_8: %s' % repr(final_8))
            print('final_4: %s' % repr(final_4))
            print('final_2: %s' % repr(final_2))
            print('goud: %s' % repr(goud))
            print('brons: %s' % repr(brons))

        for lid_nr in final_2:
            final_4.remove(lid_nr)
            if lid_nr in final_8:
                final_8.remove(lid_nr)
            if lid_nr in final_16:
                final_16.remove(lid_nr)
        # for

        for lid_nr in final_4:
            if lid_nr in final_8:
                final_8.remove(lid_nr)
            if lid_nr in final_16:
                final_16.remove(lid_nr)
        # for

        for lid_nr in final_8:
            if lid_nr in final_16:
                final_16.remove(lid_nr)
        # for

        lid_nr2deelnemer = dict()
        for deelnemer in self.rk_deelnemers:
            lid_nr2deelnemer[deelnemer.sporterboog.sporter.lid_nr] = deelnemer
            deelnemer.result_rank = 99
            deelnemer.result_volgorde = 99
        # for

        # 1
        lid_nr = goud
        deelnemer = lid_nr2deelnemer[lid_nr]
        deelnemer.result_rank = 1
        deelnemer.result_volgorde = 1
        if not self.dryrun:
            deelnemer.save(update_fields=['result_rank', 'result_volgorde'])
        final_2.remove(goud)

        # 2
        if len(final_2) == 0:
            return
        lid_nr = final_2[0]
        deelnemer = lid_nr2deelnemer[lid_nr]
        deelnemer.result_rank = 2
        deelnemer.result_volgorde = 2
        if not self.dryrun:
            deelnemer.save(update_fields=['result_rank', 'result_volgorde'])

        # 3
        if brons == 0:
            return
        lid_nr = brons
        deelnemer = lid_nr2deelnemer[lid_nr]
        deelnemer.result_rank = 3
        deelnemer.result_volgorde = 3
        if not self.dryrun:
            deelnemer.save(update_fields=['result_rank', 'result_volgorde'])
        final_4.remove(brons)

        # 4
        if len(final_4) == 0:
            return
        lid_nr = final_4[0]
        deelnemer = lid_nr2deelnemer[lid_nr]
        deelnemer.result_rank = 4
        deelnemer.result_volgorde = 4
        if not self.dryrun:
            deelnemer.save(update_fields=['result_rank', 'result_volgorde'])

        result = 5

        # 5
        sort_scores = list()
        for lid_nr in final_8:
            deelnemer = lid_nr2deelnemer[lid_nr]
            tup = (deelnemer.result_score_1 + deelnemer.result_score_2, deelnemer.volgorde, lid_nr, deelnemer)
            sort_scores.append(tup)
        # for
        sort_scores.sort(reverse=True)  # hoogste eerst
        for _, _, _, deelnemer in sort_scores:
            deelnemer.result_rank = 5
            deelnemer.result_volgorde = result
            result += 1
            if not self.dryrun:
                deelnemer.save(update_fields=['result_rank', 'result_volgorde'])
        # for

        # 9
        sort_scores = list()
        for lid_nr in final_16:
            deelnemer = lid_nr2deelnemer[lid_nr]
            tup = (deelnemer.result_score_1 + deelnemer.result_score_2, deelnemer.volgorde, lid_nr, deelnemer)
            sort_scores.append(tup)
        # for
        sort_scores.sort(reverse=True)  # hoogste eerst
        for _, _, _, deelnemer in sort_scores:
            deelnemer.result_rank = 9
            deelnemer.result_volgorde = result
            result += 1
            if not self.dryrun:
                deelnemer.save(update_fields=['result_rank', 'result_volgorde'])
        # for

        # alle overige deelnemers
        sort_scores = list()
        for deelnemer in lid_nr2deelnemer.values():
            if deelnemer.result_rank == 99:
                tup = (deelnemer.result_score_1 + deelnemer.result_score_2, deelnemer.volgorde, lid_nr, deelnemer)
                sort_scores.append(tup)
        # for
        sort_scores.sort(reverse=True)  # hoogste eerst
        for _, _, _, deelnemer in sort_scores:
            deelnemer.result_rank = result
            deelnemer.result_volgorde = result
            result += 1
            if not self.dryrun:
                deelnemer.save(update_fields=['result_rank', 'result_volgorde'])
        # for

    def handle(self, *args, **options):

        self.dryrun = options['dryrun']
        self.verbose = options['verbose']

        self._bepaal_deel()

        # open de kopie, zodat we die aan kunnen passen
        fname = options['bestand']
        self.stdout.write('[INFO] Lees bestand %s' % repr(fname))
        try:
            prg = openpyxl.load_workbook(fname,
                                         data_only=True)  # do not evaluate formulas; use last calculated values
        except (OSError, zipfile.BadZipFile, KeyError, InvalidFileException) as exc:
            self.stderr.write('[ERROR] Kan het excel bestand niet openen (%s)' % str(exc))
            return

        try:
            ws_voorronde = prg[self.blad_voorronde]
        except KeyError:        # pragma: no cover
            self.stderr.write('[ERROR] Kan blad %s niet vinden' % repr(self.blad_voorronde))
            return

        try:
            ws_finale = prg[self.blad_finales]
        except KeyError:        # pragma: no cover
            self.stderr.write('[ERROR] Kan blad %s niet vinden' % repr(self.blad_finales))
            return

        self._deelnemers_ophalen()
        self._importeer_resultaten(ws_voorronde)
        self._importeer_finales(ws_finale)

        # toon het resultaat
        unsorted = list()
        for deelnemer in self.rk_deelnemers:
            msg = 'Volgorde=%s, Rank=%s, Q-scores=%s, %s, deelnemer=%s' % (
                    deelnemer.result_volgorde, deelnemer.result_rank,
                    deelnemer.result_score_1, deelnemer.result_score_2, deelnemer)
            tup = (deelnemer.result_volgorde, msg)
            unsorted.append(tup)
        # for
        unsorted.sort()
        for _, msg in unsorted:
            self.stdout.write(msg)

# end of file
