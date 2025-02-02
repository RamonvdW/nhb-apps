# -*- coding: utf-8 -*-

#  Copyright (c) 2022-2024 Ramon van der Winkel.
#  All rights reserved.
#  Licensed under BSD-3-Clause-Clear. See LICENSE file for details.

from django.core.management.base import BaseCommand
from Competitie.definities import (DEEL_RK,
                                   DEELNAME_NEE, DEELNAME_JA,
                                   KAMP_RANK_NO_SHOW, KAMP_RANK_RESERVE)
from Competitie.models import KampioenschapSporterBoog
from openpyxl.utils.exceptions import InvalidFileException
import openpyxl
import zipfile
import sys


class Command(BaseCommand):
    help = "Importeer uitslag RK Indoor Individueel"

    blad_voorronde = 'Voorronde'

    def __init__(self, stdout=None, stderr=None, no_color=False, force_color=False):
        super().__init__(stdout, stderr, no_color, force_color)
        self.max_score = 300
        self.deelnemers = dict()        # [lid_nr] = [KampioenschapSporterBoog, ...]
        self.rk_deelnemers = list()     # [KampioenschapSporterBoog, ...]
        self.dryrun = True
        self.verbose = False
        self.deelkamp_pk = None
        self.klasse_pk = None

    def add_arguments(self, parser):
        parser.add_argument('--dryrun', action='store_true')
        parser.add_argument('--verbose', action='store_true')
        parser.add_argument('bestand', type=str,
                            help='Pad naar het Excel bestand')

    def _deelnemers_ophalen(self):
        for deelnemer in (KampioenschapSporterBoog
                          .objects
                          .filter(kampioenschap__competitie__afstand='18',
                                  kampioenschap__deel=DEEL_RK)
                          .select_related('kampioenschap',
                                          'kampioenschap__rayon',
                                          'sporterboog__sporter',
                                          'sporterboog__boogtype',
                                          'indiv_klasse')):

            # print(deelnemer, deelnemer.indiv_klasse)
            lid_nr = deelnemer.sporterboog.sporter.lid_nr
            try:
                self.deelnemers[lid_nr].append(deelnemer)
            except KeyError:
                self.deelnemers[lid_nr] = [deelnemer]
        # for

    def _importeer_resultaten(self, ws):
        col_lid_nr = 'D'
        col_score1 = 'J'
        col_score2 = 'K'

        # doorloop alle regels van het excel blad en ga op zoek naar bondsnummers
        row_nr = 7 - 1
        nix_count = 0
        while nix_count < 10:
            row_nr += 1
            row = str(row_nr)

            stop = ws['A' + row].value
            if stop == 'Afgemeld en reserven:':
                break   # from the while

            lid_nr_str = ws[col_lid_nr + row].value
            if lid_nr_str:
                nix_count = 0
                try:
                    lid_nr = int(lid_nr_str)
                except ValueError:
                    pass
                else:
                    try:
                        deelnemers = self.deelnemers[lid_nr]
                    except KeyError:
                        self.stderr.write('[ERROR] Geen RK deelnemer op regel %s: %s' % (row, lid_nr))
                    else:
                        if len(deelnemers) == 1:
                            deelnemer = deelnemers[0]
                        else:
                            deelnemer = None
                            if self.klasse_pk:
                                for kandidaat in deelnemers:
                                    if kandidaat.indiv_klasse.pk == self.klasse_pk:
                                        deelnemer = kandidaat
                                # for
                            if deelnemer is None:
                                # probeer het nog een keer, maar kijk nu ook naar afgemeld status
                                for kandidaat in deelnemers:
                                    if kandidaat.deelname != DEELNAME_NEE:
                                        if deelnemer:
                                            deelnemer = None
                                            break
                                        else:
                                            deelnemer = kandidaat
                                # for
                        if deelnemer is None:
                            self.stderr.write('[ERROR] Kan deelnemer niet bepalen voor regel %s. Keuze uit %s' % (
                                                row, repr(deelnemers)))
                            sys.exit(1)

                        if deelnemer.deelname == DEELNAME_NEE:
                            self.stdout.write('[WARNING] Was afgemeld maar wordt deelnemer: %s' % deelnemer)
                            deelnemer.deelname = DEELNAME_JA
                            if not self.dryrun:
                                deelnemer.save(update_fields=['deelname'])

                        dupe_check = False
                        if deelnemer.result_rank > 0:
                            dupe_check = True

                        if self.klasse_pk != deelnemer.indiv_klasse.pk:
                            if self.klasse_pk:
                                self.stderr.write('[ERROR] Deelnemer %s hoort in klasse: %s' % (
                                                    deelnemer, deelnemer.indiv_klasse))
                            else:
                                self.klasse_pk = deelnemer.indiv_klasse.pk
                                self.deelkamp_pk = deelnemer.kampioenschap.pk
                                self.stdout.write('[INFO] Klasse: %s' % deelnemer.indiv_klasse)

                        score1 = ws[col_score1 + row].value
                        score2 = ws[col_score2 + row].value

                        try:
                            score1 = int(score1)
                            score2 = int(score2)
                        except (TypeError, ValueError):
                            if score1 is None and score2 is None:
                                self.stdout.write('[WARNING] Regel %s wordt overgeslagen (geen scores)' % row)
                            else:
                                self.stderr.write('[ERROR] Probleem met scores op regel %s: %s en %s' % (
                                                    row, repr(score1), repr(score2)))
                        else:
                            totaal = score1 + score2
                            if score1 > self.max_score or score2 > self.max_score:
                                self.stderr.write('[ERROR] Te hoge score op regel %s: %s + %s (max is %s)' % (
                                                    row_nr, score1, score2, self.max_score))

                            if totaal > 0:                  # soms wordt 0,0 ingevuld bij niet aanwezig
                                if self.verbose:
                                    self.stdout.write('[DEBUG] Voorronde: %s, scores: %s %s' % (
                                                        deelnemer, score1, score2))

                                do_report = False
                                if dupe_check:
                                    # allow result_rank change
                                    is_dupe = (deelnemer.result_score_1 == score1
                                               and deelnemer.result_score_2 == score2)
                                    if not is_dupe:
                                        do_report = True

                                if do_report:
                                    self.stderr.write('[ERROR] Deelnemer pk=%s heeft al andere resultaten! (%s)' % (
                                                        deelnemer.pk, deelnemer))
                                else:
                                    deelnemer.result_score_1 = score1
                                    deelnemer.result_score_2 = score2

                        if not self.dryrun:
                            deelnemer.save(update_fields=['result_score_1', 'result_score_2'])

                        self.deelnemers[lid_nr] = [deelnemer]
                        self.rk_deelnemers.append(deelnemer)
            else:
                nix_count += 1
        # while

    def _report_no_shows(self):
        # zet deelnemers die niet meegedaan hebben op een hoog rank nummer
        max_deelnemers = 24
        for deelnemers in self.deelnemers.values():
            for deelnemer in deelnemers:
                if deelnemer.kampioenschap.pk == self.deelkamp_pk and deelnemer.indiv_klasse.pk == self.klasse_pk:
                    totaal = deelnemer.result_score_1 + deelnemer.result_score_2
                    if deelnemer.result_rank in (0, KAMP_RANK_NO_SHOW, KAMP_RANK_RESERVE) or totaal == 0:
                        if deelnemer.deelname != DEELNAME_NEE and deelnemer.rank < max_deelnemers:
                            # noteer: we weten niet welke reserves opgeroepen waren
                            # noteer: nog geen oplossing voor reserves die toch niet mee kunnen doen
                            self.stdout.write('[WARNING] Mogelijke no-show voor deelnemer %s' % deelnemer)
                            deelnemer.result_rank = KAMP_RANK_NO_SHOW
                        else:
                            deelnemer.result_rank = KAMP_RANK_RESERVE
                        if not self.dryrun:
                            deelnemer.save(update_fields=['result_rank'])
            # for
        # for

    def _get_lid_nr(self, ws, col_str, row_str):
        lid_nr_str = ws[col_str + row_str].value
        if lid_nr_str:
            try:
                lid_nr = int(lid_nr_str[1:1 + 6])  # [123456] Naam
            except ValueError:
                pass
            else:
                return lid_nr

        if lid_nr_str != 'BYE':
            self.stdout.write('[WARNING] Geen valide bondsnummer op regel %s in %s' % (row_str, repr(lid_nr_str)))
        return None

    def _importeer_finales(self, ws, aantal_finalisten):
        """ Zoek uit wie er in de finale zaten """

        if len(self.rk_deelnemers) == 0:
            return

        rows_16 = ('8', '10', '14', '16', '20', '22', '26', '28', '32', '34', '38', '40', '44', '46', '50', '52')
        rows_8 = ('11', '13', '23', '25', '35', '37', '47', '49')
        rows_4 = ('17', '19', '41', '43')
        rows_f12 = ('25', '27')
        rows_f34 = ('33', '35')

        col_16 = 'B'
        col_8 = 'H'
        col_4 = 'N'
        col_f = 'T'
        col_r = 'X'

        if aantal_finalisten == 8:
            col_8 = 'B'
            col_4 = 'H'
            col_f = 'N'
            col_r = 'R'
        elif aantal_finalisten == 4:
            col_4 = 'B'
            col_f = 'H'
            col_r = 'L'

        final_16 = list()
        final_8 = list()
        final_4 = list()
        final_2 = list()
        goud = 0
        brons = 0

        if aantal_finalisten == 16:
            # 1/8 finales (final 16)
            for row_str in rows_16:
                lid_nr = self._get_lid_nr(ws, col_16, row_str)
                if lid_nr:
                    final_16.append(lid_nr)
            # for

        if aantal_finalisten >= 8:
            # 1/4 finales (final 8)
            for row_str in rows_8:
                lid_nr = self._get_lid_nr(ws, col_8, row_str)
                if lid_nr:
                    final_8.append(lid_nr)
            # for

        # 1/2 finales (final 4)
        for row_str in rows_4:
            lid_nr = self._get_lid_nr(ws, col_4, row_str)
            if lid_nr:
                final_4.append(lid_nr)
        # for

        # gouden finale (final 2)
        for row_str in rows_f12:
            lid_nr = self._get_lid_nr(ws, col_f, row_str)
            if lid_nr:
                final_2.append(lid_nr)

                # heeft deze sporter goud?
                result = ws[col_r + row_str].value
                if result == 'Goud':
                    goud = lid_nr

        # for

        # bronzen finale (plek 3, 4)
        for row_str in rows_f34:
            lid_nr = self._get_lid_nr(ws, col_f, row_str)
            if lid_nr:
                # heeft deze sporter brons?
                result = ws[col_r + row_str].value
                if result == 'Brons':
                    brons = lid_nr

        # for

        # if lid_nr_str == 'nvt':
        #     # bewust niet geschoten --> 3e en 4e plek krijgen beide rank=3
        #     self.stdout.write('[INFO] Geen bronzen finale geschoten')
        #     brons = -1

        if self.verbose:
            self.stdout.write('[DEBUG] Aantal RK deelnemers: %s' % len(self.rk_deelnemers))

        # bepaal het aantal finalisten
        deelnemers_in_finale = 0
        for deelnemer in self.rk_deelnemers:
            if deelnemer.result_score_1 > 0 and deelnemer.result_score_2 > 0:
                deelnemers_in_finale += 1
        # for
        if self.verbose:
            self.stdout.write('[DEBUG] Aantal deelnemers in finale: %s' % deelnemers_in_finale)

        if goud == 0:
            self.stderr.write('[ERROR] Kan winnaar van gouden finale niet vaststellen')
            return

        if brons == 0 and deelnemers_in_finale > 2:
            self.stderr.write('[ERROR] Kan winnaar van bronzen finale niet vaststellen')
            return

        if self.verbose:
            self.stdout.write('[DEBUG] final_16: %s' % repr(final_16))
            self.stdout.write('[DEBUG] final_8: %s' % repr(final_8))
            self.stdout.write('[DEBUG] final_4: %s' % repr(final_4))
            self.stdout.write('[DEBUG] final_2: %s' % repr(final_2))
            self.stdout.write('[DEBUG] brons: %s' % repr(brons))
            self.stdout.write('[DEBUG] goud: %s' % repr(goud))

        for lid_nr in final_2:
            final_4.remove(lid_nr)
            if lid_nr in final_8:
                final_8.remove(lid_nr)
            if lid_nr in final_16:
                final_16.remove(lid_nr)
        # for

        for lid_nr in final_4:
            if lid_nr in final_8:
                final_8.remove(lid_nr)
            if lid_nr in final_16:
                final_16.remove(lid_nr)
        # for

        for lid_nr in final_8:
            if lid_nr in final_16:
                final_16.remove(lid_nr)
        # for

        lid_nr2deelnemer = dict()
        for deelnemer in self.rk_deelnemers:
            lid_nr2deelnemer[deelnemer.sporterboog.sporter.lid_nr] = deelnemer
            deelnemer.result_rank = 99
            deelnemer.result_volgorde = 99
        # for

        # 1
        lid_nr = goud
        deelnemer = lid_nr2deelnemer[lid_nr]
        deelnemer.result_rank = 1
        deelnemer.result_volgorde = 1
        if not self.dryrun:
            deelnemer.save(update_fields=['result_rank', 'result_volgorde'])
        final_2.remove(goud)

        # 2
        if len(final_2) == 0:
            return
        lid_nr = final_2[0]
        deelnemer = lid_nr2deelnemer[lid_nr]
        deelnemer.result_rank = 2
        deelnemer.result_volgorde = 2
        if not self.dryrun:
            deelnemer.save(update_fields=['result_rank', 'result_volgorde'])

        # 3
        if brons == -1:
            # beide bronzen finalisten krijgen plek 3
            # sorteer daarbinnen op kwalificatie score
            sort_scores = list()
            for lid_nr in final_4:
                deelnemer = lid_nr2deelnemer[lid_nr]
                tup = (deelnemer.result_score_1 + deelnemer.result_score_2, deelnemer.volgorde, lid_nr, deelnemer)
                sort_scores.append(tup)
            # for
            sort_scores.sort(reverse=True)  # hoogste eerst
            volgorde = 3
            for _, _, _, deelnemer in sort_scores:
                deelnemer.result_rank = 3
                deelnemer.result_volgorde = volgorde
                volgorde += 1
                if not self.dryrun:
                    deelnemer.save(update_fields=['result_rank', 'result_volgorde'])
            # for
        else:
            if brons == 0:
                return
            lid_nr = brons
            deelnemer = lid_nr2deelnemer[lid_nr]
            deelnemer.result_rank = 3
            deelnemer.result_volgorde = 3
            if not self.dryrun:
                deelnemer.save(update_fields=['result_rank', 'result_volgorde'])
            final_4.remove(brons)

            # 4
            if len(final_4) == 0:
                return
            lid_nr = final_4[0]
            deelnemer = lid_nr2deelnemer[lid_nr]
            deelnemer.result_rank = 4
            deelnemer.result_volgorde = 4
            if not self.dryrun:
                deelnemer.save(update_fields=['result_rank', 'result_volgorde'])

        result = 5

        # 5
        sort_scores = list()
        for lid_nr in final_8:
            deelnemer = lid_nr2deelnemer[lid_nr]
            tup = (deelnemer.result_score_1 + deelnemer.result_score_2, deelnemer.volgorde, lid_nr, deelnemer)
            sort_scores.append(tup)
        # for
        sort_scores.sort(reverse=True)  # hoogste eerst
        for _, _, _, deelnemer in sort_scores:
            deelnemer.result_rank = 5
            deelnemer.result_volgorde = result
            result += 1
            if not self.dryrun:
                deelnemer.save(update_fields=['result_rank', 'result_volgorde'])
        # for

        # 9
        sort_scores = list()
        for lid_nr in final_16:
            deelnemer = lid_nr2deelnemer[lid_nr]
            tup = (deelnemer.result_score_1 + deelnemer.result_score_2, deelnemer.volgorde, lid_nr, deelnemer)
            sort_scores.append(tup)
        # for
        sort_scores.sort(reverse=True)  # hoogste eerst
        for _, _, _, deelnemer in sort_scores:
            deelnemer.result_rank = 9
            deelnemer.result_volgorde = result
            result += 1
            if not self.dryrun:
                deelnemer.save(update_fields=['result_rank', 'result_volgorde'])
        # for

        # alle overige deelnemers
        sort_scores = list()
        for deelnemer in lid_nr2deelnemer.values():
            if deelnemer.result_rank == 99:
                tup = (deelnemer.result_score_1 + deelnemer.result_score_2, deelnemer.volgorde, lid_nr, deelnemer)
                sort_scores.append(tup)
        # for
        sort_scores.sort(reverse=True)  # hoogste eerst
        rank_doorlopend = result
        rank = rank_doorlopend
        prev_totaal = -1
        for totaal, _, _, deelnemer in sort_scores:
            if totaal != prev_totaal:
                rank = rank_doorlopend
            # else: zelfde score, zelfde rank
            deelnemer.result_rank = rank
            deelnemer.result_volgorde = rank_doorlopend
            rank_doorlopend += 1
            prev_totaal = totaal
            if not self.dryrun:
                deelnemer.save(update_fields=['result_rank', 'result_volgorde'])
        # for

    def _vind_finales_blad(self, prg):
        ws = None
        max_finalisten = 0

        for blad_naam, col in (("Finales 16", "X"), ("Finales 8", "R"), ("Finales 4", "L")):
            ws = prg[blad_naam]
            labels = list()
            for row in ("25", "27", "33", "35"):
                label = ws[col + row].value
                if label:
                    labels.append(label)
            # for

            # self.stdout.write('[DEBUG] blad_naam: %s, labels: %s' % (repr(blad_naam), repr(labels)))
            if len(labels) > 1:
                self.stdout.write('[INFO] Finale blad: %s' % repr(blad_naam))
                break
            ws = None
        # for

        if ws:
            if blad_naam.endswith("16"):
                max_finalisten = 16
            elif blad_naam.endswith("8"):
                max_finalisten = 8
            else:
                max_finalisten = 4

        return ws, max_finalisten

    def handle(self, *args, **options):

        self.dryrun = options['dryrun']
        self.verbose = options['verbose']

        # open de kopie, zodat we die aan kunnen passen
        fname = options['bestand']
        self.stdout.write('[INFO] Lees bestand %s' % repr(fname))
        try:
            prg = openpyxl.load_workbook(fname,
                                         data_only=True)  # do not evaluate formulas; use last calculated values
        except (OSError, zipfile.BadZipFile, KeyError, ValueError, InvalidFileException) as exc:
            self.stderr.write('[ERROR] Kan het excel bestand niet openen (%s)' % str(exc))
            return

        try:
            ws_voorronde = prg[self.blad_voorronde]
        except KeyError:        # pragma: no cover
            self.stderr.write('[ERROR] Kan blad %s niet vinden' % repr(self.blad_voorronde))
            return

        ws_finale, max_finalisten = self._vind_finales_blad(prg)
        if not ws_finale:
            self.stderr.write('[ERROR] Kan finales blad niet vinden')
            return

        self._deelnemers_ophalen()
        self._importeer_resultaten(ws_voorronde)
        self._importeer_finales(ws_finale, max_finalisten)
        self._report_no_shows()

        # toon het resultaat
        if self.verbose > 0:
            unsorted = list()
            for deelnemer in self.rk_deelnemers:
                msg = 'Volgorde=%s, Rank=%s, Q-scores=%s, %s, deelnemer=%s' % (
                        deelnemer.result_volgorde, deelnemer.result_rank,
                        deelnemer.result_score_1, deelnemer.result_score_2, deelnemer)
                tup = (deelnemer.result_volgorde, msg)
                unsorted.append(tup)
            # for
            unsorted.sort()
            for _, msg in unsorted:
                self.stdout.write(msg)

# end of file
